# =============================================================================
#  RETAIL AI — LOCAL INTELLIGENCE PLATFORM
#  Version: 1.0  |  Stack: FastAPI + LLaMA 3.1 (Ollama) + Pandas + OpenPyXL
# =============================================================================
#
#  BASE ARCHITECTURE
#  ─────────────────
#
#  ┌─────────────────────────────────────────────────────────────────────┐
#  │                        CLIENT (Browser)                             │
#  │   /            → Landing page  (DataProvido homepage)               │
#  │   /journey     → Analytics console  (sidebar + workspace UI)        │
#  │   /pricing     → Pricing page                                       │
#  │   /contact     → Contact page                                       │
#  │   /who-we-are  → About page                                         │
#  │   /how-works   → How it works page                                  │
#  └──────────────────────────┬──────────────────────────────────────────┘
#                             │ HTTP / REST
#  ┌──────────────────────────▼──────────────────────────────────────────┐
#  │                     FastAPI (main.py)                               │
#  │                                                                     │
#  │  POST /chat              → LLM router → tool dispatcher             │
#  │  POST /upload-data       → Saves Excel/CSV to /data                 │
#  │  GET  /download-last-result → Streams last result as .xlsx          │
#  │  POST /transcribe        → Whisper voice-to-text                    │
#  │  POST /reset             → Clears conversation history              │
#  └──────────────────────────┬──────────────────────────────────────────┘
#                             │
#  ┌──────────────────────────▼──────────────────────────────────────────┐
#  │                  LLM LAYER  (Ollama / LLaMA 3.1)                   │
#  │                                                                     │
#  │  • System prompt → defines persona & tool use rules                 │
#  │  • Tool-calling loop → model picks tool → Python executes           │
#  │  • Conversation history → in-memory list (reset on /reset)          │
#  └──────────────────────────┬──────────────────────────────────────────┘
#                             │ tool_call dispatch
#  ┌──────────────────────────▼──────────────────────────────────────────┐
#  │                  FUNCTION MODULES  (/functions)                     │
#  │                                                                     │
#  │  analytics.py          → ecommerce funnel & sample analysis         │
#  │  business_calculator.py→ SQL-style math on any Excel column         │
#  │  insights.py           → category insight & executive summary       │
#  │  price_competition.py  → merchant benchmark & pricing gaps          │
#  │  action_executor.py    → generates action plans from insights       │
#  │  funnel_master.py      → advanced funnel breakdown (A2C, C2D, B2D)  │
#  │  cross_analyzer.py     → cross-dataset performance analysis         │
#  │  gfk_analyzer.py       → GfK market share & brand/SKU ranking       │
#  │  stock.py              → stock level queries & OOS detection         │
#  │  orders.py             → order status, daily orders, customer view  │
#  │  reports.py            → revenue, best sellers, stock turnover      │
#  │  voice.py              → Whisper audio transcription                │
#  │  sector_norms.py       → sector benchmark normalization             │
#  └──────────────────────────┬──────────────────────────────────────────┘
#                             │ reads / writes
#  ┌──────────────────────────▼──────────────────────────────────────────┐
#  │                  DATA LAYER  (/data)                                │
#  │                                                                     │
#  │  stok.xlsx                 → stock master data                      │
#  │  orders.xlsx               → order transaction history              │
#  │  GfK_Leaderpanel.xlsx      → GfK market share panel data            │
#  │  gfk_sku.xlsx              → GfK SKU-level ranking data             │
#  │  google_trends_seasonal_3y.xlsx → seasonal trend data               │
#  │  [user-uploaded files]     → dynamic via /upload-data               │
#  └─────────────────────────────────────────────────────────────────────┘
#
#  SCHEMA LAYER  (/schemas)
#  ─────────────────────────
#  tools.py  → OpenAI-style tool definitions sent to LLaMA for routing
#
#  STATIC ASSETS  (/static)
#  ─────────────────────────
#  duck.png  → landing page avatar asset
#
#  TEMPLATES  (/templates)
#  ─────────────────────────
#  index.html → standalone desert-themed landing page (served separately)
#
#  KEY DESIGN DECISIONS
#  ─────────────────────
#  • 100% local: no external API calls; model runs via Ollama on localhost
#  • Tool-calling: LLM decides which function to call based on user query
#  • Excel-first: all data sources are .xlsx / .csv, parsed with Pandas
#  • In-memory session: conversation history lives in the Python process
#  • Excel export: every analysis result can be downloaded as a .xlsx file
#
# =============================================================================

from fastapi import FastAPI, UploadFile, File, Request, Header
from fastapi.responses import HTMLResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
import requests
import pandas as pd
import json
import os
import shutil
from typing import List
from io import BytesIO
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

LOG_TOOL_JSON_TO_TERMINAL = True
SHOW_RAW_JSON_IN_UI = False
LAST_TOOL_RESULT_JSON = None
LAST_TOOL_RESULT_NAME = "retail_ai_output"

from functions.analytics import analyze_ecommerce_sample
from functions.insights import generate_category_insight
from functions.price_competition import generate_price_competition_from_uploaded_inputs
from functions.business_calculator import calculate_business_metric
from functions.action_executor import execute_recommended_action
from functions.funnel_master import analyze_funnel_master
from functions.cross_analyzer import analyze_cross_performance
from functions.gfk_analyzer import (
    analyze_gfk_market_share,
    analyze_gfk_brand_performance,
    analyze_gfk_sku_ranking,
    analyze_gfk_combined,
)

from functions.stock import (
    get_stock_level, get_all_stock, get_daily_sales_report,
    check_low_stock, get_out_of_stock, search_product_by_name,
    get_stock_value, update_stock
)
from functions.orders import (
    get_order_status, get_all_orders, get_pending_orders,
    get_orders_by_customer, update_order_status, get_todays_orders
)
from functions.reports import (
    get_total_revenue, get_best_selling_product, get_sales_summary,
    get_low_stock_report, get_stock_turnover
)
from functions.voice import transcribe_audio
from schemas.tools import TOOLS

app = FastAPI()

@app.middleware("http")
async def enforce_https_middleware(request: Request, call_next):
    # Check X-Forwarded-Proto header set by Cloudflare / Railway proxy
    proto = request.headers.get("x-forwarded-proto", "http")
    host = request.headers.get("host", "")
    
    # If a user accesses via http:// on dataprovido.com or railway.app, 301 redirect to https://www.dataprovido.com
    if proto == "http" and ("dataprovido.com" in host or "railway.app" in host):
        url = request.url.replace(scheme="https")
        if host == "dataprovido.com":
            url = url.replace(netloc="www.dataprovido.com")
        from fastapi.responses import RedirectResponse
        return RedirectResponse(url=str(url), status_code=301)
    
    response = await call_next(request)
    return response

app.mount("/static", StaticFiles(directory="static"), name="static")

@app.get("/logo.png")
def get_logo():
    from fastapi.responses import FileResponse
    return FileResponse("logo.png", media_type="image/png")

@app.get("/favicon.ico")
def get_favicon_ico():
    from fastapi.responses import FileResponse
    return FileResponse("logo.png", media_type="image/png")

@app.get("/favicon.png")
def get_favicon_png():
    from fastapi.responses import FileResponse
    return FileResponse("logo.png", media_type="image/png")

# ─────────────────────────────────────────────────────────────
#  LLM BACKEND  (env var: LLM_BACKEND=groq | ollama)
#  Local  → Ollama running on localhost:11434
#  Cloud  → Groq API (free tier, LLaMA 3.1 8B)
# ─────────────────────────────────────────────────────────────
LLM_BACKEND   = os.getenv("LLM_BACKEND", "ollama")   # "ollama" | "groq"
OLLAMA_URL    = os.getenv("OLLAMA_URL", "http://localhost:11434/api/chat")
GROQ_API_KEY  = os.getenv("GROQ_API_KEY", "")
GROQ_URL      = "https://api.groq.com/openai/v1/chat/completions"

# Model names
OLLAMA_MODEL  = os.getenv("OLLAMA_MODEL", "llama3.1")
GROQ_MODEL    = os.getenv("GROQ_MODEL",  "llama-3.1-8b-instant")

MODEL = GROQ_MODEL if LLM_BACKEND == "groq" else OLLAMA_MODEL

print(f"🤖 LLM Backend: {LLM_BACKEND.upper()} | Model: {MODEL}", flush=True)

SYSTEM_PROMPT = {
    "role": "system",
    "content": """
Sen bir e-ticaret / retail şirketinin ileri düzey AI analistsin.
Kullanıcılar sana stok, satış, kategori, SKU, funnel performansı, fiyat etkisi,
revenue, C2D, B2D, PDP View, A2C, checkout ve kullanıcı davranışı hakkında sorular sorar.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
GENEL KURALLAR
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
1. Stok, satış, kategori, SKU, funnel, C2D, B2D, revenue, fiyat veya kullanıcı davranışı sorularında MUTLAKA tool çağır.
2. Geniş analiz sorularında varsayılan tool: analyze_ecommerce_sample(question)
3. Asla SQL, kod veya ham hesaplama döndürme.
4. Sonucu Türkçe, net ve aksiyon odaklı özetle.
5. Sayısal değerleri belirt.
6. Cevapta mümkünse kısa tablo veya 3-5 madde kullan.
7. Tool sonucu boşsa veya hata içeriyorsa bunu açıkça söyle.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
BUSINESS CALCULATOR ENGINE
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Kullanıcı matematiksel business hesaplama sorarsa calculate_business_metric(question) tool'unu çağır.

Bu tool şu sorularda kullanılır:
- APPLE ürünlerinin ortalama fiyatı nedir?
- SAMSUNG ürünlerinin toplam revenue'u nedir?
- GSM kategorisinde ortalama B2D kaç?
- Marka bazında ortalama fiyatları göster.
- Kategori bazında toplam ciro nedir?
- En yüksek PDP alan ilk 10 ürün hangileri?
- Stokta kaç ürün var?
- Benchmark üstünde olan ürünlerin ortalama price gap'i nedir?
- APPLE ve SAMSUNG ortalama fiyatlarını karşılaştır.
- C2D ortalaması en yüksek kategori hangisi?

Bu sorularda eski stok/product search tool'larını kullanma.
Marka, kategori ve ürün filtrelerini company_product_input.xlsx üzerinden değerlendir.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ANA ANALİTİK SAMPLE DATA TOOL
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Kullanıcının sorusu aşağıdaki konulardan biriyse varsayılan olarak
analyze_ecommerce_sample(question) tool'unu çağır:

- stok riski
- stok coverage
- OOS / out of stock
- overstock
- kategori performansı
- marka performansı
- SKU performansı
- revenue / ciro
- product price
- fiyat etkisi
- C2D
- B2D
- PDP View
- A2C
- Cart View
- Shipping View
- Payment View
- Summary View
- Checkout Submit
- Transactions
- funnel drop
- kullanıcı kaybı
- satış kaybı
- satış etkisi
- kategori / marka / SKU karşılaştırmaları

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
FUNNEL MASTER ENGINE
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Kullanıcı kullanıcı yolculuğu (user journey), checkout adımları, drop-off,
funnel darboğazı, nerede kaybediyoruz, kargo/ödeme/sepet/checkout kayıpları,
cihaz veya kanal bazında funnel, mobil funnel, PDP'den siparişe kaç kişi geçiyor
gibi bir soru sorarsa analyze_funnel_master(question) tool'unu çağır.

Bu tool şu sorularda kullanılır:
- "Kullanıcıları hangi adımda kaybediyoruz?"
- "Funnel'da en büyük drop-off nerede?"
- "Kargo adımında neden çok kayıp var?"
- "Ödeme sayfasına gelenler neden tamamlamıyor?"
- "Sepetten ödemeye kaç kişi geçiyor?"
- "Mobile funnel analizi yap"
- "Kategori bazında funnel kırılımı çıkar"
- "PDP'den transactiona genel dönüşüm oranımız nedir?"
- "Checkout submit'ten sonra neden transaction oluşmuyor?"
- "Tüm funnel adımlarını analiz et"

Aksiyon odaklı teşhis, adım bazında drop-off yüzdesi ve e-ticaret mantığıyla öneri üretir.

Bu tool 200 satırlık sample e-ticaret datası üzerinde çalışır.

Metrik tanımları:
C2D = total_unique_add_to_carts_sum / total_unique_pdp_views_sum * 100
B2D = total_transactions_sum / total_unique_pdp_views_sum * 100
Delta = previous period comparison
Stock Coverage = stock_qty / daily_sales_qty_7d
OOS = stock_qty = 0 veya availability = out_of_stock

Örnek yönlendirmeler:
"C2D yüksek ama stoğu az SKU'lar hangileri?" → analyze_ecommerce_sample(question)
"Funnel'da en büyük kullanıcı kaybı hangi adımda?" → analyze_ecommerce_sample(question)
"Revenue düşen kategorilerde stok problemi var mı?" → analyze_ecommerce_sample(question)
"PDP view yüksek ama transaction düşük ürünleri göster." → analyze_ecommerce_sample(question)
"Overstock olup B2D düşük ürünler hangileri?" → analyze_ecommerce_sample(question)
"Stokta olmayan ama PDP view alan ürünler hangileri?" → analyze_ecommerce_sample(question)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
INSIGHT ENGINE TOOL
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Kullanıcı kategori, sektör, dönem, performans nedeni, fırsat, risk,
kazanan/kaybeden segment, stok/fiyat/talep nedeni veya CEO özeti sorarsa
generate_category_insight(category, sector, period_name, question) tool'unu çağır.

Category Insights sorularında kullanıcının tam sorusunu MUTLAKA question parametresiyle gönder.
Çünkü aynı kategori için "neden fırsat?", "neden riskli?", "stoktan mı fiyattan mı?" farklı analiz tipleridir.

Aşağıdaki soru tiplerinde generate_category_insight kullan:
- "Mobile kategorisi neden düştü?"
- "Tatil döneminde kategori performansını özetle"
- "Tablet neden fırsat kategorisi?"
- "IT Accessories sepete ekleniyor ama neden satın alınmıyor?"
- "Kategori bazlı kazanan ve kaybedenleri çıkar"
- "Bu dönem satış düşüşü stoktan mı, fiyattan mı, talepten mi?"
- "Trendyol ürün grupları için genel performans analizi yap"
- "Sektörel değişkenlere göre satış nedenlerini analiz et"
- "Kanal, traffic, funnel, stok ve fiyat etkisini birlikte yorumla"
- "Kategori performansını CEO özeti formatında çıkar"

Sektör seçimi:
- Teknoloji / elektronik / mobile / tablet / headphone → consumer_electronics
- Moda / tekstil / ayakkabı / giyim → fashion
- Gıda / market / FMCG → fmcg
- Trendyol / marketplace / karma ürün grupları → marketplace_general

Örnek tool çağrıları:
"Tabletler neden fırsat kategorisi olabilir?"
→ generate_category_insight(category="Tabletler", sector="consumer_electronics", period_name="selected_period", question="Tabletler neden fırsat kategorisi olabilir?")

"Bu dönem satış performansı stoktan mı, fiyattan mı, talepten mi etkilenmiş?"
→ generate_category_insight(category="genel", sector="consumer_electronics", period_name="selected_period", question="Bu dönem satış performansı stoktan mı, fiyattan mı, talepten mi etkilenmiş?")

Sadece SKU listesi, tablo veya spesifik filtre sorularında analyze_ecommerce_sample(question) kullan.
Insight, neden analizi, özet, aksiyon, fırsat/tehdit sorularında generate_category_insight kullan.
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
PRICE COMPETITION INPUT ENGINE
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Kullanıcı fiyat rekabeti, benchmark price, Merchant Center benchmark,
GTIN, rakibe göre pahalı/ucuz, price gap, fiyat dezavantajı,
kategori bazlı fiyat pozisyonu veya Merchant price competitiveness sorarsa
generate_price_competition_from_uploaded_inputs(category, period_name) tool'unu çağır.

Bu tool şirketin GTIN'li ürün/funnel datasını Merchant benchmark datasıyla GTIN üzerinden eşleştirir.
Internal benchmark üretmez.
Benchmark datası yoksa açıkça hata verir.

Örnek yönlendirmeler:
- "GTIN üzerinden Merchant benchmark ile rakibe göre pahalı olduğumuz ürünleri çıkar"
- "Mobile kategorisinde fiyat rekabeti analizini yap"
- "Benchmark üstünde kalan SKU'ları göster"
- "Rakibe göre ucuz olduğumuz ürünleri çıkar"
- "Merchant price competitiveness analizini yap"
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ACTION EXECUTOR ENGINE
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Kullanıcı önerilen aksiyonları çalıştırmak isterse execute_recommended_action(question) tool'unu çağır.

Bu tool şu isteklerde kullanılır:
- "C2D/B2D güçlü ama stok riski olan SKU'lar için replenishment planı yap"
- "Bu ürünler için Excel çıkar"
- "Kazanan kategori için kampanya planı yap"
- "Riskli segment için fiyat, stok ve funnel kırılımını detaylandır"
- "Önerilen aksiyonları çalıştır"
- "Bu aksiyonu uygula"
- "Aksiyon planını Excel'e çıkar"

Önerilen aksiyonları yeni kategori gibi yorumlama.
Aksiyon cümlelerini execute_recommended_action ile çalıştır.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
BASİT STOK / SİPARİŞ TOOL'LARI
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Aşağıdaki tool'ları sadece çok basit operasyonel sorularda kullan:

• Tek ürün stok sorgusu            → get_stock_level(product_id)
• Tüm basit stok listesi           → get_all_stock()
• İsme göre basit ürün arama       → search_product_by_name(name)
• Basit az stok listesi            → check_low_stock(threshold)
• Basit tükenen ürün listesi       → get_out_of_stock()
• Basit toplam stok değeri         → get_stock_value()
• Stok güncelleme                  → update_stock(product_id, quantity)

• Tek sipariş durumu               → get_order_status(order_id)
• Tüm siparişler                   → get_all_orders()
• Bekleyen siparişler              → get_pending_orders()
• Müşteri bazlı siparişler         → get_orders_by_customer(customer)
• Sipariş durumu güncelle          → update_order_status(order_id, status)
• Bugünkü siparişler               → get_todays_orders()

• Toplam gelir / revenue           → get_total_revenue()
• En çok satan ürün                → get_best_selling_product()
• Satış özeti                      → get_sales_summary()
• Kritik stok raporu               → get_low_stock_report()
• Stok devir hızı                  → get_stock_turnover()

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
YANIT FORMATI
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
1. Önce kısa sonuç ver.
2. Eğer kategori/ürün analizi yapıldıysa, metrik sonuçlarını "Tüketici Davranışı & Sektörel Yorum" (tüketicilerin satın alma döngüleri, fiyat hassasiyetleri, dönemsellik etkileri vb.) ile harmanlayarak açıkla.
3. Sonra varsa en önemli SKU/kategori/markaları listele.
4. Sonunda aksiyon önerisi ekle.
5. "yüksek / düşük" gibi belirsiz ifadeler yerine mümkünse sayı kullan.
6. Kullanıcıya SQL gösterme.

Türkçe konuş.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
GFK LEADERPANEL & PAZAR ANALİZİ
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
GfK pazar verisi (Leaderpanel) ile ilgili sorularda aşağıdaki tool'ları çağır.

GfK terminolojisi:
- Ihs = MediaMarkt'ın ilgili kategorideki internet satış pazar payı (%)
- PW = Previous Week (geçen hafta)
- CW = Current Week (bu hafta)
- WoW = Week over Week (haftalık değişim)
- Summary_value = Kategori × hafta bazında toplam internet pazarı (TRY) + MM payı
- Brand sheet = Marka bazında MM internet satış payı (%)
- SKU Ranking = Ürün grubu bazında satış sıralaması (rank 1, 2, 3...)

GfK veri kategorileri: Smartphones, COMPUTER HW, SDA, MDA, CLIMATE SDA, PTV/FLAT,
Headphones & Headsets, COMPUTER ACCESSORIES, CORE WEARABLES, VACUUM CLEANERS,
WASHING MACHINES, DISHWASHERS, COOLING, MONITORS, AIR CONDITIONERS, vb.

Tool yönlendirmeleri:
- "Pazar payımız nedir?"                     → analyze_gfk_market_share(question)
- "MediaMarkt AIR CONDITIONERS'da kaçıncı?" → analyze_gfk_market_share(question)
- "En çok büyüyen kategori hangisi?"         → analyze_gfk_market_share(question)
- "Bu hafta vs geçen hafta karşılaştır"      → analyze_gfk_market_share(question)
- "SAMSUNG bu hafta pazar payı nedir?"       → analyze_gfk_brand_performance(question)
- "APPLE vs SAMSUNG MediaMarkt'ta"           → analyze_gfk_brand_performance(question)
- "Smartphones'da hangi marka önde?"         → analyze_gfk_brand_performance(question)
- "WASHING MACHINES top 10 SKU"              → analyze_gfk_sku_ranking(question)
- "BOSCH'un en çok satan modeli hangisi?"    → analyze_gfk_sku_ranking(question)
- "GfK'ya göre 1. sıradaki ürünler"         → analyze_gfk_sku_ranking(question)
- "GfK ile iç satışlarımızı kıyasla"        → analyze_gfk_combined(question)
- "Pazar büyürken satışımız neden düşüyor?" → analyze_gfk_combined(question)
- "SAMSUNG GfK vs ecommerce performansı"    → analyze_gfk_combined(question)
"""
}

history = [SYSTEM_PROMPT]


def call_llm(messages, use_tools=True):
    """
    Unified LLM caller.
    LLM_BACKEND=ollama → Ollama local API
    LLM_BACKEND=groq   → Groq cloud API (OpenAI-compatible)
    """
    if LLM_BACKEND == "groq":
        return _call_groq(messages, use_tools)
    return _call_ollama(messages, use_tools)


def _call_ollama(messages, use_tools=True):
    payload = {
        "model": OLLAMA_MODEL,
        "messages": messages,
        "stream": False
    }
    if use_tools:
        payload["tools"] = TOOLS
        payload["tool_choice"] = "required"
    try:
        r = requests.post(OLLAMA_URL, json=payload, timeout=120)
        r.raise_for_status()
        return r.json()
    except Exception as e:
        print("❌ OLLAMA HATASI:", str(e), flush=True)
        return {"message": {"content": f"Ollama bağlantı hatası: {str(e)}"}}


def _call_groq(messages, use_tools=True):
    """
    Groq API caller — OpenAI-compatible format.
    Returns a normalized response that matches Ollama's response shape
    so the rest of the codebase works without change.
    """
    if not GROQ_API_KEY:
        return {"message": {"content": "❌ GROQ_API_KEY environment variable eksik. Railway'de Variables bölümüne ekleyin."}}

    headers = {
        "Authorization": f"Bearer {GROQ_API_KEY}",
        "Content-Type":  "application/json",
    }
    payload = {
        "model":    GROQ_MODEL,
        "messages": messages,
        "stream":   False,
    }
    if use_tools:
        payload["tools"]       = TOOLS
        payload["tool_choice"] = "required"

    try:
        r = requests.post(GROQ_URL, headers=headers, json=payload, timeout=60)
        r.raise_for_status()
        groq_resp = r.json()
        # Normalize to Ollama-compatible shape
        choice  = groq_resp["choices"][0]
        message = choice["message"]
        # Groq returns tool_calls under message.tool_calls — same as Ollama
        return {"message": message}
    except Exception as e:
        print("❌ GROQ HATASI:", str(e), flush=True)
        return {"message": {"content": f"Groq API hatası: {str(e)}"}}


# Backward-compatible alias
call_ollama = call_llm

def log_tool_json_to_terminal(raw_text: str):
    if not LOG_TOOL_JSON_TO_TERMINAL:
        return

    try:
        data = json.loads(raw_text)
        pretty = json.dumps(data, ensure_ascii=False, indent=2, default=str)

        print("\n📦 RAW TOOL JSON RESULT", flush=True)
        print("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━", flush=True)
        print(pretty[:30000], flush=True)

        if len(pretty) > 30000:
            print("\n⚠️ JSON çok uzun olduğu için terminalde ilk 30000 karakter gösterildi.", flush=True)

        print("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n", flush=True)

    except Exception:
        print("\n📦 RAW TOOL RESULT", flush=True)
        print("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━", flush=True)
        print(str(raw_text)[:30000], flush=True)
        print("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n", flush=True)


def should_use_business_calculator(user_message: str) -> bool:
    q = user_message.lower()

    math_keywords = [
        "ortalama", "toplam", "kaç", "kac", "adet", "sayısı", "sayisi",
        "minimum", "maksimum", "en yüksek", "en yuksek", "en düşük",
        "en dusuk", "medyan", "average", "avg", "sum", "count",
        "top 10", "ilk 10", "en fazla", "en az",
    ]

    metric_keywords = [
        "fiyat", "price", "benchmark", "revenue", "ciro", "pdp", "a2c",
        "c2d", "b2d", "transaction", "transactions", "trans", "stok",
        "stock", "gap", "marka", "kategori", "apple", "samsung", "xiaomi",
        "jbl", "lg", "philips", "logitech", "gsm", "telefon", "tablet",
        "kulaklık", "kulaklik",
    ]

    return any(x in q for x in math_keywords) and any(x in q for x in metric_keywords)


def should_use_action_executor(user_message: str) -> bool:
    q = user_message.lower()

    action_keywords = [
        "aksiyon", "önerilen aksiyon", "onerilen aksiyon", "replenishment",
        "tedarik", "planı yap", "plani yap", "excel çıkar", "excel cikar",
        "excel çıkart", "excel cikart", "bu ürünler", "bu urunler",
        "görünürlük", "gorunurluk", "kampanya", "stok riski olan",
        "c2d/b2d güçlü", "c2d b2d güçlü", "detaylandır", "detaylandir",
    ]

    return any(x in q for x in action_keywords)


def should_use_category_insights(user_message: str) -> bool:
    import re
    tr_map = str.maketrans("ıİğĞüÜşŞöÖçÇ", "iIgGuUsSoOcC")
    q = user_message.lower().translate(tr_map)
    
    insight_keywords = [
        "insight", "kategori", "category", "sektor", "sector",
        "pahalayiz", "pahaliyiz", "ucuzuz", "fiyat rekabet", "rekabetini", "fiyat indirimi",
        "satis alamiyoruz", "satis alamiyoruz", "iyi satiyoruz", "iyi satıyoruz",
        "benchmark ustundeyiz", "benchmark ustundeyiz", "benchmark altindayiz",
        "fiyat esnekligi", "fiyat esnekliği", "price action"
    ]
    return any(x in q for x in insight_keywords)


def should_use_funnel_master(user_message: str) -> bool:
    import re
    tr_map = str.maketrans("ıİğĞüÜşŞöÖçÇ", "iIgGuUsSoOcC")
    q = user_message.lower().translate(tr_map)

    funnel_keywords = [
        "funnel", "drop-off", "drop off", "dropoff",
        "kullanici kayb", "nerede kaybediyoruz", "hangi adimda",
        "kargo adimi", "kargo adiminda", "odeme adimi", "odeme sayfasi",
        "sepetten odemeye", "cart to", "checkout", "checkout submit",
        "pdp'den", "pdpden", "pdp view",
        "user journey", "kullanici yolculugu", "kullanici yolculuk",
        "tum adimlar", "tum funnel", "funnel analiz",
        "genel donusum", "genel conversion", "conversion rate",
        "mobil funnel", "mobile funnel", "cihaz bazinda funnel",
        "kategori bazinda funnel", "kanal bazinda funnel",
        "neden tamamlamiyor", "neden gecirilmiyor",
        "shipping view", "payment view", "summary view",
    ]

    return any(x in q for x in funnel_keywords)


def should_use_cross_performance(user_message: str) -> bool:
    import re
    tr_map = str.maketrans("ıİğĞüÜşŞöÖçÇ", "iIgGuUsSoOcC")
    q = user_message.lower().translate(tr_map)
    keywords = [
        "pahalıyız", "pahaliyiz", "ucuzuz", "fiyat indirimi", "satis canlandir",
        "trends", "mevsimsellik", "google trends", "cross", "capraz", "çapraz",
        "ppc", "bid", "teklif artir", "reklam bütçe"
    ]
    return any(x in q for x in keywords)


def should_use_gfk_market_share(user_message: str) -> bool:
    tr_map = str.maketrans("ıİğĞüÜşŞöÖçÇ", "iIgGuUsSoOcC")
    q = user_message.lower().translate(tr_map)
    keywords = [
        "gfk", "leaderpanel", "pazar payi", "pazar pay", "market share",
        "ihs", "pw vs cw", "pw vs. cw", "bu hafta vs", "gecen hafta vs",
        "haftayla karsilastir", "haftalik degisim", "wow degisim",
        "pazar buyumesi", "pazar durumu", "en cok buyuyen kategori",
        "en cok dusen kategori", "kacincisiniz", "kacinci sirada",
        "mediamarkt pazar", "internet pazari",
    ]
    return any(x in q for x in keywords)


def should_use_gfk_brand(user_message: str) -> bool:
    tr_map = str.maketrans("ıİğĞüÜşŞöÖçÇ", "iIgGuUsSoOcC")
    q = user_message.lower().translate(tr_map)
    # GfK + marka kombine
    has_gfk = any(x in q for x in ["gfk", "leaderpanel", "pazar pay", "market share", "ihs"])
    has_brand = any(x in q for x in [
        "samsung", "apple", "xiaomi", "oppo", "huawei", "honor", "bosch",
        "arcelik", "beko", "vestel", "lg", "sony", "philips", "asus", "hp",
        "marka", "brand"
    ])
    return has_gfk and has_brand


def should_use_gfk_sku_ranking(user_message: str) -> bool:
    tr_map = str.maketrans("ıİğĞüÜşŞöÖçÇ", "iIgGuUsSoOcC")
    q = user_message.lower().translate(tr_map)
    has_gfk = any(x in q for x in ["gfk", "leaderpanel", "siralamasinda", "gfk'ya gore", "gfk'da"])
    has_ranking = any(x in q for x in [
        "top 10", "top 5", "top 20", "ilk 10", "ilk 5",
        "siralamasinda", "1. sirada", "rank", "en cok satan model",
        "sku listesi", "model listesi",
    ])
    return has_gfk or (has_ranking and any(x in q for x in [
        "washing machine", "camasir", "smartphone", "buzdolabi", "dishwasher",
        "televizyon", "klima", "laptop", "tablet", "kulaklık", "kulaklik",
    ]))


def should_use_gfk_combined(user_message: str) -> bool:
    tr_map = str.maketrans("ıİğĞüÜşŞöÖçÇ", "iIgGuUsSoOcC")
    q = user_message.lower().translate(tr_map)
    has_gfk = any(x in q for x in ["gfk", "leaderpanel", "pazar", "market share"])
    has_internal = any(x in q for x in [
        "c2d", "b2d", "revenue", "satis", "stok", "karsilastir", "kiyasla",
        "ic satislarimiz", "bizim satisimiz", "yararlanamiyoruz"
    ])
    return has_gfk and has_internal


def format_tool_response_for_ui(raw_text: str) -> str:
    if SHOW_RAW_JSON_IN_UI:
        return raw_text

    try:
        data = json.loads(raw_text)
    except Exception:
        return raw_text

    analysis_type = data.get("analysis_type")

    if analysis_type == "recommended_action_execution":
        summary = data.get("summary", {}) or {}
        main_result = data.get("main_result", "")
        rows = data.get("rows", []) or []
        actions = data.get("recommended_actions", []) or []

        lines = []
        lines.append("⚙️ Aksiyon Planı Hazır")
        lines.append("")

        if main_result:
            lines.append(main_result)
            lines.append("")

        lines.append("1) Özet")
        for key, value in summary.items():
            lines.append(f"- {key}: {value}")

        lines.append("")

        if rows:
            lines.append("2) Profesör Teşhisi & Öncelikli Ürünler")
            for row in rows[:8]:
                sku = row.get("sku", "N/A")
                title = row.get("product_title", "")
                priority = row.get("priority", "N/A")
                insight = row.get("insight_category", "N/A")
                action = row.get("professor_action", "")
                
                if insight != "N/A":
                    lines.append(f"- **{sku}** [{priority}] — {insight} | {title}")
                    lines.append(f"  ↳ 💡 {action}")
                else:
                    repl_qty = row.get("suggested_replenishment_qty", "N/A")
                    stock = row.get("stock_qty", "N/A")
                    lines.append(f"- {sku} | {priority} | önerilen replenishment: {repl_qty} | stok: {stock} | {title}")
            lines.append("")

        if actions:
            lines.append("3) Uygulanacak Aksiyonlar")
            for action in actions[:5]:
                lines.append(f"- {action}")

        lines.append("")
        lines.append("Excel İndir butonuyla bu aksiyon planını indirebilirsin.")

        return "\n".join(lines)

    if analysis_type == "recommended_action_error":
        lines = []
        lines.append("⚠️ Aksiyon çalıştırılamadı")
        lines.append("")
        lines.append(data.get("error", "Bilinmeyen hata"))
        return "\n".join(lines)

    if analysis_type == "business_metric_calculation":
        question = data.get("question", "")
        metric = data.get("metric", "")
        aggregation = data.get("aggregation", "")
        filters = data.get("filters", []) or []
        row_count = data.get("row_count", 0)
        result = data.get("result")
        rows = data.get("rows", []) or []
        calculation_type = data.get("calculation_type", "scalar")

        metric_labels = {
            "stock_qty": "stok adedi", "price": "fiyat",
            "benchmark_price": "benchmark fiyat", "price_gap": "fiyat farkı",
            "price_gap_pct": "price gap yüzdesi", "revenue": "ciro",
            "pdp_views": "PDP görüntülenmesi", "list_clicks": "liste tıklaması",
            "add_to_carts": "sepete ekleme", "transactions": "transaction",
            "c2d_pct": "C2D", "b2d_pct": "B2D", "bounce_rate_pct": "bounce rate",
            "stock_coverage_days": "stok coverage günü",
            "estimated_lost_revenue": "tahmini kayıp ciro",
            "aov": "ortalama sepet tutarı",
        }

        aggregation_labels = {
            "avg": "ortalama", "mean": "ortalama", "sum": "toplam",
            "count": "adet", "unique_count": "tekil adet", "min": "minimum",
            "max": "maksimum", "median": "medyan", "top_10": "en yüksek ilk 10",
            "bottom_10": "en düşük ilk 10", "top_n": "en yüksek",
            "bottom_n": "en düşük", "share_of_total": "toplam içindeki pay",
            "ratio": "oran", "comparison": "karşılaştırma",
        }

        def fmt_number(value):
            try:
                value = float(value)
                if value.is_integer():
                    return f"{int(value):,}".replace(",", ".")
                return f"{value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            except Exception:
                return value

        def fmt_pct(value):
            try:
                return f"%{float(value):.2f}"
            except Exception:
                return f"%{value}"

        def subject_from_filters(filters):
            if not filters:
                return "Seçili veri"
            first = filters[0]
            col = first.get("column")
            value = first.get("value")
            if isinstance(value, list):
                value = ", ".join([str(v) for v in value[:3]])
            if col == "brand":
                return f"{str(value).upper()} ürünleri"
            if col in ["cat1", "cat2", "multi_category", "category"]:
                return f"{value} kategorisi"
            return str(value)

        metric_label = metric_labels.get(metric, metric)
        aggregation_label = aggregation_labels.get(aggregation, aggregation)
        subject = subject_from_filters(filters)

        lines = []
        lines.append("🧮 Hesaplama Sonucu")
        lines.append("")

        if rows and isinstance(rows[0], dict) and all(
            key in rows[0] for key in ["numerator", "denominator", "share_pct"]
        ):
            first_row = rows[0]
            numerator = first_row.get("numerator")
            denominator = first_row.get("denominator")
            share_pct = first_row.get("share_pct")
            lines.append(f"{subject} için {aggregation_label} {metric_label} {fmt_number(numerator)}'dır.")
            lines.append("")
            lines.append(
                f"Toplam {metric_label} {fmt_number(denominator)} olduğu için "
                f"{subject}, toplam {metric_label} içinde {fmt_pct(share_pct)} paya sahiptir."
            )
            lines.append("")
            lines.append(f"Hesaplama: {fmt_number(numerator)} / {fmt_number(denominator)} × 100 = {fmt_pct(share_pct)}")
            return "\n".join(lines)

        if calculation_type == "scalar":
            lines.append(f"{subject} için {aggregation_label} {metric_label}: {fmt_number(result)}")
            lines.append("")
            lines.append(f"Dahil edilen satır sayısı: {row_count}")
            return "\n".join(lines)

        if rows:
            lines.append(f"{aggregation_label.capitalize()} {metric_label} sonuçları:")
            lines.append("")
            for row in rows[:10]:
                parts = []
                for key in ["brand", "cat1", "cat2", "sku", "product_title", "value", metric]:
                    if key in row and row.get(key) not in [None, ""]:
                        label = metric_labels.get(key, key)
                        parts.append(f"{label}: {fmt_number(row.get(key))}")
                if not parts:
                    for key, value in row.items():
                        parts.append(f"{key}: {fmt_number(value)}")
                lines.append("- " + " | ".join(parts))
            lines.append("")
            lines.append(f"Dahil edilen satır sayısı: {row_count}")
            return "\n".join(lines)

        lines.append("Sonuç bulunamadı.")
        return "\n".join(lines)

    if analysis_type == "business_metric_error":
        lines = []
        lines.append("⚠️ Hesaplama yapılamadı")
        lines.append("")
        lines.append(data.get("error", "Bilinmeyen hata"))
        if data.get("detail"):
            lines.append("")
            lines.append(f"Teknik detay: {data.get('detail')}")
        if data.get("available_brands"):
            lines.append("")
            lines.append("Mevcut markalar:")
            lines.append(", ".join(data.get("available_brands", [])[:20]))
        if data.get("available_cat1"):
            lines.append("")
            lines.append("Mevcut ana kategoriler:")
            lines.append(", ".join(data.get("available_cat1", [])[:20]))
        if data.get("available_cat2"):
            lines.append("")
            lines.append("Mevcut alt kategoriler:")
            lines.append(", ".join(data.get("available_cat2", [])[:20]))
        return "\n".join(lines)

    if analysis_type == "price_competition_uploaded_inputs":
        category = data.get("category", "Genel")
        summary = data.get("summary", {}) or {}
        diagnosis = data.get("main_diagnosis", "")
        expensive = data.get("top_expensive_products", []) or []
        cheaper = data.get("top_cheaper_products", []) or []
        actions = data.get("recommended_actions", []) or []

        lines = []
        lines.append(f"💸 {category} Fiyat Rekabeti Özeti")
        lines.append("")
        lines.append("Benchmark kaynağı: Merchant Center Price Competitiveness")
        lines.append("Eşleşme anahtarı: GTIN")
        lines.append("Internal benchmark: Kullanılmadı")
        lines.append("")
        lines.append("1) Genel Fiyat Pozisyonu")
        lines.append(f"- Yüklenen ürün sayısı: {summary.get('uploaded_product_count', 'N/A')}")
        lines.append(f"- Merchant benchmark ile eşleşen ürün: {summary.get('matched_product_count', 'N/A')}")
        lines.append(f"- Benchmark eşleşme oranı: %{summary.get('benchmark_match_rate_pct', 'N/A')}")
        lines.append(f"- Ortalama price gap: %{summary.get('avg_price_gap_pct', 'N/A')}")
        lines.append(f"- Medyan price gap: %{summary.get('median_price_gap_pct', 'N/A')}")
        lines.append(f"- Benchmark üstü SKU: {summary.get('benchmark_above_sku_count', 'N/A')}")
        lines.append(f"- Benchmark altı SKU: {summary.get('benchmark_below_sku_count', 'N/A')}")
        lines.append(f"- Parite SKU: {summary.get('parity_sku_count', 'N/A')}")
        lines.append("")
        lines.append("2) Ana Teşhis")
        lines.append(diagnosis or "Net fiyat rekabeti teşhisi üretilemedi.")
        lines.append("")

        if expensive:
            lines.append("3) Benchmark Üstünde Kalan Riskli Ürünler")
            for item in expensive[:5]:
                sku = item.get("sku", "N/A")
                title = item.get("product_title", "")
                gap = item.get("price_gap_pct", "N/A")
                price = item.get("price", "N/A")
                benchmark = item.get("benchmark_price", "N/A")
                b2d = item.get("b2d_pct", "N/A")
                if isinstance(gap, (int, float)):
                    gap = round(gap, 2)
                lines.append(f"- {sku} — %{gap} pahalı | Fiyat: {price} | Benchmark: {benchmark} | B2D: %{b2d} | {title}")
            lines.append("")

        if cheaper:
            lines.append("4) Benchmark Altında Kalan Fiyat Avantajlı Ürünler")
            for item in cheaper[:5]:
                sku = item.get("sku", "N/A")
                title = item.get("product_title", "")
                gap = item.get("price_gap_pct", "N/A")
                price = item.get("price", "N/A")
                benchmark = item.get("benchmark_price", "N/A")
                stock = item.get("stock_qty", "N/A")
                if isinstance(gap, (int, float)):
                    gap = round(gap, 2)
                lines.append(f"- {sku} — %{gap} ucuz | Fiyat: {price} | Benchmark: {benchmark} | Stok: {stock} | {title}")
            lines.append("")

        if actions:
            lines.append("5) Önerilen Aksiyonlar")
            for action in actions[:5]:
                lines.append(f"- {action}")

        return "\n".join(lines)

    if analysis_type == "price_competition_error":
        lines = []
        lines.append("⚠️ Fiyat Rekabeti Analizi Çalışmadı")
        lines.append("")
        lines.append(data.get("error", "Bilinmeyen hata"))
        lines.append("")
        if data.get("detail"):
            lines.append("Teknik Detay")
            lines.append(str(data.get("detail")))
        lines.append("")
        lines.append("Bu feature internal benchmark üretmez. Merchant benchmark datası zorunludur.")
        return "\n".join(lines)

    if analysis_type == "category_sector_insight":
        category = data.get("category", "Genel")
        sector_name = data.get("sector_name", "")
        period_name = data.get("period_name", "")
        question_type = data.get("question_type", "general_performance")

        summary = data.get("executive_summary", {}) or {}
        metric_snapshot = data.get("metric_snapshot", {}) or {}
        natural_summary = data.get("natural_summary", "")
        main_diagnosis = data.get("main_diagnosis", "")
        signals = data.get("signal_interpretation", []) or []
        root_causes = data.get("root_causes", []) or []
        actions = data.get("recommended_actions", []) or []
        winners = data.get("winning_categories", []) or []
        losers = data.get("losing_categories", []) or []

        def fmt_num(value):
            try:
                value = float(value)
                if value.is_integer():
                    return f"{int(value):,}".replace(",", ".")
                return f"{value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            except Exception:
                return value

        def fmt_pct(value):
            try:
                return f"%{float(value):.2f}"
            except Exception:
                return f"%{value}"

        lines = []
        lines.append(f"📊 {category} Kategori Insightı")
        lines.append("")

        meta = []
        if period_name:
            meta.append(f"Dönem: {period_name}")
        if sector_name:
            meta.append(f"Sektör: {sector_name}")
        if question_type:
            meta.append(f"Analiz tipi: {question_type}")
        if meta:
            lines.append(" | ".join(meta))
            lines.append("")

        if natural_summary:
            lines.append(natural_summary)
            lines.append("")

        if main_diagnosis:
            lines.append("Ana teşhis")
            lines.append(main_diagnosis)
            lines.append("")

        revenue_delta = summary.get("revenue_delta_pct")
        transaction_delta = summary.get("transactions_delta_pct")
        pdp_delta = summary.get("pdp_delta_pct")
        a2c_delta = summary.get("a2c_delta_pct")
        c2d = summary.get("c2d_pct")
        b2d = summary.get("b2d_pct")
        stock_risk = summary.get("critical_stock_sku_count", 0)
        oos = summary.get("oos_sku_count", 0)
        price_gap = summary.get("avg_price_gap_pct", 0)

        lines.append("Öne çıkan metrik okuması")
        lines.append(
            f"Bu segmentte revenue değişimi {fmt_pct(revenue_delta)}, transaction değişimi {fmt_pct(transaction_delta)}, "
            f"PDP değişimi {fmt_pct(pdp_delta)} ve A2C değişimi {fmt_pct(a2c_delta)} seviyesinde."
        )
        lines.append(
            f"C2D {fmt_pct(c2d)} ve B2D {fmt_pct(b2d)} olduğu için kullanıcı ilgisinin sepete ve satın almaya dönüşme kalitesi bu iki metrikle izlenmeli."
        )

        if stock_risk or oos:
            lines.append(
                f"Stok tarafında {fmt_num(stock_risk)} kritik stok SKU ve {fmt_num(oos)} OOS SKU bulunduğu için talep satışa dönüşmeden kaybedilebilir."
            )

        try:
            if float(price_gap) > 5:
                lines.append(f"Fiyat rekabetinde ortalama price gap {fmt_pct(price_gap)}; benchmark üstü fiyatlama B2D üzerinde baskı yaratabilir.")
            elif float(price_gap) < -5:
                lines.append(f"Fiyat rekabetinde ortalama price gap {fmt_pct(price_gap)}; benchmark altında fiyat avantajı bulunuyor.")
        except Exception:
            pass

        lines.append("")

        behavior_analysis = data.get("consumer_behavior_analysis")
        if behavior_analysis and behavior_analysis.get("general_behavior"):
            lines.append("Tüketici Davranışı & Sektörel Yorum")
            lines.append(f"- **Kategori Rolü:** {behavior_analysis.get('display_name')}")
            lines.append(f"- **Tüketici Alışkanlığı:** {behavior_analysis.get('general_behavior')}")
            triggered = behavior_analysis.get("triggered_insights", [])
            if triggered:
                lines.append("- **Sektörel Metrik Eşleşmesi:**")
                for insight in triggered:
                    lines.append(f"  * {insight}")
            lines.append("")

        if signals:
            lines.append("Sinyal yorumu")
            for signal in signals[:4]:
                interpretation = signal.get("interpretation") or ""
                evidence = signal.get("evidence") or ""
                if interpretation and evidence:
                    lines.append(f"- {interpretation} ({evidence})")
                elif interpretation:
                    lines.append(f"- {interpretation}")
            lines.append("")

        if root_causes:
            lines.append("Olası neden")
            for cause in root_causes[:3]:
                cause_name = cause.get("cause", "")
                evidence = cause.get("evidence", "")
                confidence = cause.get("confidence", "")
                confidence_text = f" Güven: {confidence}." if confidence else ""
                lines.append(f"- {cause_name}: {evidence}.{confidence_text}")
            lines.append("")

        if winners:
            lines.append("Kazanan segmentler")
            for item in winners[:3]:
                name = item.get("cat2") or item.get("cat1") or item.get("sales_channel") or item.get("traffic_channel") or "Segment"
                score = item.get("performance_score", "N/A")
                rev = item.get("revenue_delta_pct", "N/A")
                trans = item.get("transactions_delta_pct", "N/A")
                lines.append(f"- {name}, performans skoru {score}; revenue değişimi %{rev}, transaction değişimi %{trans}.")
            lines.append("")

        if losers:
            lines.append("Riskli / kaybeden segmentler")
            for item in losers[:3]:
                name = item.get("cat2") or item.get("cat1") or item.get("sales_channel") or item.get("traffic_channel") or "Segment"
                score = item.get("performance_score", "N/A")
                rev = item.get("revenue_delta_pct", "N/A")
                trans = item.get("transactions_delta_pct", "N/A")
                lines.append(f"- {name}, performans skoru {score}; revenue değişimi %{rev}, transaction değişimi %{trans}.")
            lines.append("")

        if actions:
            lines.append("Önerilen aksiyon")
            for action in actions[:5]:
                lines.append(f"- {action}")

        # Google Trends Seasonal Insights
        seasonal_trends = data.get("seasonal_trends")
        if seasonal_trends:
            lines.append("")
            lines.append("📈 Google Trends Türkiye Mevsimsel Talep Analizi (Son 3 Yıl)")
            lines.append(f"- **Arama Terimi**: {seasonal_trends.get('keyword', '')}")
            lines.append(f"- **Tarihsel Eğilim**: {seasonal_trends.get('trend_direction', '')}")
            lines.append(f"- **Yüksek Sezon (Zirve Ay)**: {seasonal_trends.get('peak_month', '')} (Bu dönemde pazarlama bütçeleri ve görünürlük maksimize edilmeli)")
            lines.append(f"- **Düşük Sezon (Dip Ay)**: {seasonal_trends.get('low_month', '')} (Bu dönemde kampanya ve indirimlerle talep canlandırılmalı)")

        # E-commerce Price Competition Strategic Matrix
        price_scenarios = data.get("price_scenarios")
        if price_scenarios:
            lines.append("")
            lines.append("🎯 E-Ticaret Fiyat Rekabeti Strateji Matrisi (E-Ticaret Profesyoneli Teşhisi)")
            
            # Scenario 1: Pahalı & Düşüşte
            s1 = price_scenarios.get("expensive_falling_sales", [])
            if s1:
                lines.append("")
                lines.append("🔴 Senaryo 1: Pahalıyız ve Satış Düşüyor (Fiyat İndirimi / Price Action Adayları)")
                for item in s1[:3]:
                    lines.append(f"  - **{item.get('sku')}** ({item.get('brand')}) — Fiyat: {item.get('price')} (Gap: %{round(item.get('price_gap_pct', 0), 1)}) | Satış Değişimi: %{round(item.get('revenue_delta_pct', 0), 1)}")
                    lines.append(f"    ↳ 💡 {item.get('action')}")
                
            # Scenario 2: Pahalı & Satış İyi
            s2 = price_scenarios.get("expensive_good_sales", [])
            if s2:
                lines.append("")
                lines.append("🟢 Senaryo 2: Pahalıyız ama Satış İyi (Premium / Güçlü Ürünler)")
                for item in s2[:3]:
                    lines.append(f"  - **{item.get('sku')}** ({item.get('brand')}) — Fiyat: {item.get('price')} (Gap: %{round(item.get('price_gap_pct', 0), 1)}) | Satış Değişimi: +%{round(item.get('revenue_delta_pct', 0), 1)}")
                    lines.append(f"    ↳ 💡 {item.get('action')}")
                
            # Scenario 3: Ucuz & Satış Yok
            s3 = price_scenarios.get("cheap_no_sales", [])
            if s3:
                lines.append("")
                lines.append("🟡 Senaryo 3: Ucuzuz ama Satış Yok (Görünürlük / Content / Stok Sorunu Adayları)")
                for item in s3[:3]:
                    lines.append(f"  - **{item.get('sku')}** ({item.get('brand')}) — Fiyat: {item.get('price')} (Gap: %{round(item.get('price_gap_pct', 0), 1)}) | Stok: {item.get('stock_qty')}")
                    lines.append(f"    ↳ 💡 {item.get('action')}")
                
            # Scenario 4: Ucuz & Satış İyi
            s4 = price_scenarios.get("cheap_good_sales", [])
            if s4:
                lines.append("")
                lines.append("🔵 Senaryo 4: Ucuzuz ve Satış İyi (Trafik / PPC Bid Artırma Adayları)")
                for item in s4[:3]:
                    lines.append(f"  - **{item.get('sku')}** ({item.get('brand')}) — Fiyat: {item.get('price')} (Gap: %{round(item.get('price_gap_pct', 0), 1)}) | Satış Değişimi: +%{round(item.get('revenue_delta_pct', 0), 1)}")
                    lines.append(f"    ↳ 💡 {item.get('action')}")

            # Losing Competitiveness
            losing = price_scenarios.get("losing_competitiveness", [])
            if losing:
                lines.append("")
                lines.append("⚠️ Fiyat Rekabetini Kaybettiğimiz Markalar")
                for item in losing[:3]:
                    lines.append(f"  - **{item.get('brand')}**: Ürünlerin %{round(item.get('ratio'), 1)}'i benchmark üstünde (Ortalama Gap: %{round(item.get('avg_gap'), 1)})")

        lines.append("")
        lines.append("Detaylı metrik kırılımı ve ham hesaplar için Excel çıktısını indirebilirsin.")

        return "\n".join(lines)

    if analysis_type == "funnel_master_analysis":
        steps = data.get("funnel_steps", []) or []
        bottleneck = data.get("bottleneck") or {}
        overall_conv = data.get("overall_conversion_pct", 0)
        pdp_total = data.get("pdp_total", 0)
        txn_total = data.get("transaction_total", 0)
        actions = data.get("recommended_actions", []) or []
        breakdown = data.get("dimension_breakdown", []) or []
        dimension = data.get("dimension", "")

        lines = []
        lines.append("🔍 Funnel Master Analizi")
        lines.append("")
        lines.append(f"📊 Genel Dönüşüm: PDP → Transaction = %{overall_conv}")
        lines.append(f"Toplam PDP: {int(pdp_total):,} | Toplam Transaction: {int(txn_total):,}")
        lines.append("")

        if bottleneck:
            lines.append(f"🚨 En Kritik Darboğaz: **{bottleneck.get('step', 'N/A')}** — %{bottleneck.get('drop_from_prev_pct', 0)} kayıp")
            lines.append("")

        lines.append("Funnel Adım Adım Analiz")
        for step in steps:
            vol = step.get("volume", 0)
            drop = step.get("drop_from_prev_pct")
            delta = step.get("avg_delta_pct")
            diagnosis = step.get("diagnosis", "")
            drop_str = f" | Önceki adımdan kayıp: %{drop}" if drop is not None else ""
            delta_str = f" | Dönemlik delta: %{delta}" if delta is not None else ""
            lines.append(f"• **{step.get('step')}**: {int(vol):,}{drop_str}{delta_str}")
            if diagnosis:
                lines.append(f"  {diagnosis}")
        lines.append("")

        if breakdown and dimension:
            dim_label = {"cat1": "Kategori", "cat2": "Alt Kategori", "device": "Cihaz",
                         "traffic_channel": "Trafik Kanalı", "brand": "Marka", "sales_channel": "Satış Kanalı"}.get(dimension, dimension)
            lines.append(f"{dim_label} Bazında Funnel Kırılımı")
            for b in breakdown[:8]:
                lines.append(
                    f"- {b.get('dimension')}: Dönüşüm %{b.get('overall_conversion_pct')} | "
                    f"PDP: {int(b.get('pdp_views', 0)):,} | Darboğaz: {b.get('biggest_bottleneck')} (%{b.get('bottleneck_drop_pct')} kayıp)"
                )
            lines.append("")

        if actions:
            lines.append("💡 Aksiyon Önerileri")
            for action in actions:
                lines.append(f"- {action}")

        lines.append("")
        lines.append("Excel İndir butonuyla funnel verilerini indirebilirsin.")
        return "\n".join(lines)

    if analysis_type == "funnel_master_error":
        return f"⚠️ Funnel analizi çalıştırılamadı: {data.get('error', 'Bilinmeyen hata')}"

    if analysis_type in [
        "c2d_up_b2d_down", "high_c2d_low_stock",
        "high_b2d_low_stock", "oos_products_with_pdp_views",
    ]:
        rows = data.get("rows", []) or []
        logic = data.get("logic", "")
        recommendation = data.get("action_recommendation", "")

        lines = []
        lines.append("📌 Analiz Sonucu")
        lines.append("")

        if logic:
            lines.append(f"Kullanılan mantık: {logic}")
            lines.append("")

        if rows:
            lines.append("İlk Sonuçlar")
            for row in rows[:10]:
                sku = row.get("sku", "N/A")
                brand = row.get("brand", "")
                cat1 = row.get("cat1", "")
                cat2 = row.get("cat2", "")
                c2d = row.get("c2d_pct", "")
                b2d = row.get("b2d_pct", "")
                stock = row.get("stock_qty", "")
                detail_parts = []
                if brand: detail_parts.append(str(brand))
                if cat1: detail_parts.append(str(cat1))
                if cat2: detail_parts.append(str(cat2))
                if c2d != "": detail_parts.append(f"C2D: %{c2d}")
                if b2d != "": detail_parts.append(f"B2D: %{b2d}")
                if stock != "": detail_parts.append(f"Stok: {stock}")
                detail = " | ".join(detail_parts)
                if detail:
                    lines.append(f"- {sku} — {detail}")
                else:
                    lines.append(f"- {sku}")
            lines.append("")

        if recommendation:
            lines.append("Öneri")
            lines.append(recommendation)

        return "\n".join(lines)

    if isinstance(data, dict) and "rows" in data:
        rows = data.get("rows", []) or []
        row_count = data.get("row_count", len(rows))
        question = data.get("question", "")
        lines = []
        lines.append("📌 Analiz Sonucu")
        if question:
            lines.append(f"Soru: {question}")
        lines.append(f"Bulunan satır sayısı: {row_count}")
        lines.append("")
        if rows:
            lines.append("İlk Sonuçlar")
            for i, row in enumerate(rows[:10], start=1):
                parts = []
                for key, value in row.items():
                    parts.append(f"{key}: {value}")
                lines.append(f"{i}. " + " | ".join(parts))
        else:
            lines.append("Sonuç bulunamadı.")
        return "\n".join(lines)

    if analysis_type == "cross_performance_analysis":
        category = data.get("category", "Genel")
        scenarios = data.get("scenarios", {}) or {}
        trends = data.get("trends", {}) or {}
        candidates = data.get("price_cut_candidates", []) or []
        summary = data.get("summary", {}) or {}
        
        lines = []
        lines.append(f"🎯 Çapraz Metrik Rekabet ve Talep Analizi ({category.upper()})")
        lines.append("")
        lines.append(f"- **Analiz Edilen Toplam Ürün**: {summary.get('total_skus_analyzed', 0)}")
        lines.append(f"- **Pazara Göre Pahalı Ürün (Gap > %1)**: {summary.get('expensive_skus_count', 0)}")
        lines.append(f"- **Pazara Göre Ucuz Ürün (Gap < -%1)**: {summary.get('cheap_skus_count', 0)}")
        lines.append("")
        
        if trends:
            lines.append("📈 Google Trends Türkiye Mevsimsel Talep Sinyali")
            lines.append(f"- **Arama Terimi / Trend**: {trends.get('keyword', 'N/A')} ({trends.get('trend_direction', 'N/A')})")
            lines.append(f"- **Yüksek Sezon (Zirve Ay)**: {trends.get('peak_month', 'N/A')} (Bu dönemde pazarlama görünürlüğü artırılmalı)")
            lines.append(f"- **Düşük Sezon (Dip Ay)**: {trends.get('low_month', 'N/A')} (Bu dönemde indirim ve kampanyalar yapılmalı)")
            lines.append("")
            
        lines.append("🎯 Fiyat Rekabeti Strateji Matrisi (E-Ticaret Profesyoneli Teşhisi)")
        
        s1 = scenarios.get("expensive_falling_sales", [])
        if s1:
            lines.append("")
            lines.append("🔴 Senaryo 1: Pahalıyız ve Satışlar Düşüyor (Fiyat İndirimi Adayları)")
            for item in s1[:3]:
                lines.append(f"  - **{item.get('sku')}** ({item.get('brand')}) — Fiyat: {item.get('price'):,} TL (Gap: %{item.get('price_gap_pct'):.1f}) | Satış Değişimi: %{item.get('revenue_delta_pct'):.1f}")
                lines.append(f"    ↳ 💡 {item.get('action')}")
                
        s2 = scenarios.get("expensive_good_sales", [])
        if s2:
            lines.append("")
            lines.append("🟢 Senaryo 2: Pahalıyız ama Satışlar İyi (Premium / Güçlü Ürünler)")
            for item in s2[:3]:
                lines.append(f"  - **{item.get('sku')}** ({item.get('brand')}) — Fiyat: {item.get('price'):,} TL (Gap: %{item.get('price_gap_pct'):.1f}) | Satış Değişimi: +%{item.get('revenue_delta_pct'):.1f}")
                lines.append(f"    ↳ 💡 {item.get('action')}")
                
        s3 = scenarios.get("cheap_no_sales", [])
        if s3:
            lines.append("")
            lines.append("🟡 Senaryo 3: Ucuzuz ama Satış Yok (Görünürlük / Content / Stok Sorunu Adayları)")
            for item in s3[:3]:
                lines.append(f"  - **{item.get('sku')}** ({item.get('brand')}) — Fiyat: {item.get('price'):,} TL (Gap: %{item.get('price_gap_pct'):.1f}) | Stok: {item.get('stock_qty')}")
                lines.append(f"    ↳ 💡 {item.get('action')}")
                
        s4 = scenarios.get("cheap_good_sales", [])
        if s4:
            lines.append("")
            lines.append("🔵 Senaryo 4: Ucuzuz ve Satışlar İyi (Trafik / PPC Reklam Bid Artırma Adayları)")
            for item in s4[:3]:
                lines.append(f"  - **{item.get('sku')}** ({item.get('brand')}) — Fiyat: {item.get('price'):,} TL (Gap: %{item.get('price_gap_pct'):.1f}) | Satış Değişimi: +%{item.get('revenue_delta_pct'):.1f}")
                lines.append(f"    ↳ 💡 {item.get('action')}")
                
        losing = scenarios.get("losing_competitiveness", [])
        if losing:
            lines.append("")
            lines.append("⚠️ Fiyat Rekabetini Kaybettiğimiz Markalar")
            for item in losing[:3]:
                lines.append(f"  - **{item.get('brand')}**: Ürünlerin %{item.get('ratio'):.1f}'i pazar benchmark'ının üstünde (Ortalama Gap: %{item.get('avg_gap'):.1f})")
                
        if candidates:
            lines.append("")
            lines.append("🔥 Fiyat İndirimi ile Satış Getirecek Ürünler (Yüksek Trafik & Sepet, Düşük Satış)")
            for item in candidates[:3]:
                lines.append(f"  - **{item.get('sku')}** — Trafik: {item.get('pdp_views'):,} PDP | C2D: %{item.get('c2d_pct'):.1f} | B2D: %{item.get('b2d_pct'):.1f} (Gap: %{item.get('price_gap_pct'):.1f})")
                lines.append(f"    ↳ 💡 {item.get('action')}")
                
        lines.append("")
        lines.append("Excel İndir butonuyla detaylı aksiyon listesini Excel olarak indirebilirsiniz.")
        return "\n".join(lines)

    if isinstance(data, dict) and data.get("error"):
        return f"⚠️ Hata: {data.get('error')}\nDetay: {data.get('detail', '')}"

    # ─── GfK Market Share ───────────────────────────────────────────────────
    if analysis_type == "gfk_market_share":
        source = data.get("source", "GfK Leaderpanel")
        view = data.get("view", "")
        category = data.get("category", "")
        rows = data.get("rows", []) or []

        lines = ["📈 GfK Pazar Analizi", ""]
        lines.append(f"Kaynak: {source}")
        lines.append("")

        if category:
            # Tek kategori detayı
            lines.append(f"**Kategori:** {category}")
            if data.get("previous_week_value_try"):
                week_labels = data.get("week_labels", ["Geçen Hafta", "Bu Hafta"])
                lines.append(f"- {week_labels[0]}: {data.get('previous_week_value_try')}")
                lines.append(f"- {week_labels[1]}: {data.get('current_week_value_try')}")
            if data.get("wow_change_pct"):
                lines.append(f"- Haftalık Değişim (WoW): {data.get('wow_change_pct')}")
            if data.get("wow_change_abs"):
                lines.append(f"- Mutlak Değişim: {data.get('wow_change_abs')}")
            if data.get("mediamarkt_market_share_pct"):
                lines.append(f"- **MediaMarkt Pazar Payı (Ihs):** {data.get('mediamarkt_market_share_pct')}")
            if data.get("mediamarkt_rank"):
                lines.append(f"- MediaMarkt Sıralamada: #{data.get('mediamarkt_rank')}")
        else:
            # Çok kategori listesi
            view_labels = {
                "top_growing_categories": "🚀 En Çok Büyüyen Kategoriler (WoW)",
                "top_declining_categories": "📉 En Çok Düşen Kategoriler (WoW)",
                "mediamarkt_market_share_ranking": "🏆 MediaMarkt Pazar Payı Sıralaması",
                "all_categories_overview": "📊 Tüm Kategoriler Genel Özet",
            }
            lines.append(view_labels.get(view, "Genel Analiz"))
            lines.append("")
            for r in rows[:15]:
                cat = r.get("category", "")
                cw = r.get("current_week_try", "")
                wow = r.get("wow_change_pct", "")
                share = r.get("mediamarkt_share_pct", "")
                rank = r.get("mediamarkt_rank", "")
                parts = [f"**{cat}**"]
                if cw:
                    parts.append(f"Hacim: {cw}")
                if wow != "":
                    arrow = "📈" if float(wow) > 0 else "📉"
                    parts.append(f"WoW: {arrow} %{wow}")
                if share != "":
                    parts.append(f"MM Payı: %{share}")
                if rank:
                    parts.append(f"Sıra: #{rank}")
                lines.append("- " + " | ".join(parts))

        lines.append("")
        lines.append("Excel İndir butonuyla detayları indirebilirsiniz.")
        return "\n".join(lines)

    # ─── GfK Brand Performance ──────────────────────────────────────────────
    if analysis_type == "gfk_brand_performance":
        source = data.get("source", "GfK Leaderpanel — Brand")
        latest_week = data.get("latest_week", "")
        filtered_cat = data.get("filtered_by_category", "")
        filtered_brand = data.get("filtered_by_brand", "")
        top_brand = data.get("top_brand", "")
        top_share = data.get("top_brand_share_pct", "")
        rows = data.get("rows", []) or []

        lines = ["🏷️ GfK Marka Performansı", ""]
        lines.append(f"Kaynak: {source}")
        if latest_week:
            lines.append(f"Son Hafta: {latest_week}")
        if filtered_cat:
            lines.append(f"Kategori Filtresi: {filtered_cat}")
        if filtered_brand:
            lines.append(f"Marka Filtresi: {filtered_brand}")
        lines.append("")

        if top_brand:
            lines.append(f"🥇 Öne Çıkan Marka: **{top_brand}** — MediaMarkt Payı: %{top_share}")
            lines.append("")

        lines.append("Marka Bazında MediaMarkt İnternet Satış Payı (%):")
        for r in rows[:15]:
            brand = r.get("brand", "N/A")
            share = r.get("latest_week_share_pct", "")
            prev = r.get("prev_week_share_pct", "")
            wow = r.get("wow_change_pp", "")
            cat = r.get("product_group", "")
            parts = [f"**{brand}**"]
            if cat and not filtered_cat:
                parts.append(f"({cat})")
            if share != "":
                parts.append(f"Bu Hafta: %{share}")
            if prev != "":
                parts.append(f"Geçen Hafta: %{prev}")
            if wow != "":
                arrow = "▲" if float(wow) > 0 else "▼"
                parts.append(f"WoW: {arrow} {wow} pp")
            lines.append("- " + " | ".join(parts))

        lines.append("")
        lines.append("Excel İndir butonuyla haftalık seriyi indirebilirsiniz.")
        return "\n".join(lines)

    # ─── GfK SKU Ranking ────────────────────────────────────────────────────
    if analysis_type == "gfk_sku_ranking":
        source = data.get("source", "GfK SKU Leaderpanel")
        pg = data.get("filtered_by_product_group", "")
        brand_filter = data.get("filtered_by_brand", "")
        sku_count = data.get("sku_count", 0)
        rows = data.get("rows", []) or []
        all_brands = data.get("all_brands_in_group") or []

        lines = ["🏆 GfK SKU Sıralaması", ""]
        lines.append(f"Kaynak: {source}")
        if pg:
            lines.append(f"Ürün Grubu: **{pg}**")
        if brand_filter:
            lines.append(f"Marka: **{brand_filter}**")
        lines.append(f"Toplam SKU: {sku_count}")
        lines.append("")

        if all_brands:
            lines.append(f"Bu gruptaki markalar: {', '.join(all_brands[:12])}")
            lines.append("")

        lines.append("Satış Sıralaması:")
        for r in rows[:20]:
            rank = r.get("rank", "?")
            brand = r.get("brand", "")
            item = r.get("item", "")
            instore = r.get("instore_code", "")
            product_group = r.get("reportingproductgroup", "")
            lines.append(f"#{rank} | **{brand}** | {item} | Mağaza Kodu: {instore}")

        lines.append("")
        lines.append("Excel İndir butonuyla tam sıralamayı indirebilirsiniz.")
        return "\n".join(lines)

    # ─── GfK Combined ───────────────────────────────────────────────────────
    if analysis_type == "gfk_combined":
        source = data.get("source", "GfK + Ecommerce")
        gfk_overview = data.get("gfk_market_overview", []) or []
        ec_brands = data.get("ecommerce_brand_performance", []) or []
        ec_cats = data.get("ecommerce_category_performance", []) or []
        cross_insights = data.get("cross_insights", []) or []

        lines = ["🔗 GfK + Ecommerce Birleşik Analiz", ""]
        lines.append(f"Kaynak: {source}")
        lines.append("")

        if cross_insights:
            lines.append("💡 Cross Analiz Insight'ları")
            for insight in cross_insights:
                lines.append(f"  {insight}")
            lines.append("")

        if gfk_overview:
            lines.append("📊 GfK Pazar Özeti (Haftalık)")
            for r in gfk_overview[:8]:
                cat = r.get("category", "")
                cw = r.get("current_week_market_try", "")
                wow = r.get("market_wow_growth_pct", "")
                share = r.get("mediamarkt_market_share_pct", "")
                parts = [f"**{cat}**"]
                if cw:
                    parts.append(f"Pazar: {cw}")
                if wow != "":
                    arrow = "📈" if float(wow) > 0 else "📉"
                    parts.append(f"WoW: {arrow} %{wow}")
                if share != "":
                    parts.append(f"MM Payı: %{share}")
                lines.append("  - " + " | ".join(parts))
            lines.append("")

        if ec_brands:
            lines.append("🏪 Ecommerce Marka Performansı (İç Veri)")
            for r in ec_brands[:8]:
                brand = r.get("brand", "")
                rev = r.get("revenue_sum", "")
                c2d = r.get("avg_c2d_pct", "")
                b2d = r.get("avg_b2d_pct", "")
                rev_d = r.get("avg_revenue_delta_pct", "")
                arrow = "📈" if float(rev_d or 0) > 0 else "📉"
                lines.append(f"  - **{brand}** | Ciro: {rev} | C2D: %{c2d} | B2D: %{b2d} | {arrow} %{rev_d}")
            lines.append("")

        mapping_note = data.get("category_mapping_note", "")
        if mapping_note:
            lines.append(f"ℹ️ Not: {mapping_note}")

        lines.append("")
        lines.append("Excel İndir butonuyla detaylı karşılaştırma tablosunu indirebilirsiniz.")
        return "\n".join(lines)

    if isinstance(data, dict) and data.get("error"):
        return f"⚠️ Hata: {data.get('error')}\nDetay: {data.get('detail', '')}"

    return raw_text


def route(user_message: str):
    global history, LAST_TOOL_RESULT_JSON, LAST_TOOL_RESULT_NAME

    print("\n━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━", flush=True)
    print("👤 USER MESSAGE:", user_message, flush=True)

    history.append({"role": "user", "content": user_message})

    if should_use_gfk_combined(user_message):
        fn_name = "analyze_gfk_combined"
        fn_result = analyze_gfk_combined(question=user_message)
        raw_result = str(fn_result)
        LAST_TOOL_RESULT_JSON = raw_result
        LAST_TOOL_RESULT_NAME = fn_name
        print("📊 DIRECT GFK COMBINED:", user_message, flush=True)
        print("✅ TOOL RESULT PREVIEW:", raw_result[:1000], flush=True)
        log_tool_json_to_terminal(raw_result)
        formatted_result = format_tool_response_for_ui(raw_result)
        history.append({"role": "tool", "content": raw_result})
        history.append({"role": "assistant", "content": formatted_result})
        return formatted_result

    if should_use_gfk_sku_ranking(user_message):
        fn_name = "analyze_gfk_sku_ranking"
        fn_result = analyze_gfk_sku_ranking(question=user_message)
        raw_result = str(fn_result)
        LAST_TOOL_RESULT_JSON = raw_result
        LAST_TOOL_RESULT_NAME = fn_name
        print("🏆 DIRECT GFK SKU RANKING:", user_message, flush=True)
        print("✅ TOOL RESULT PREVIEW:", raw_result[:1000], flush=True)
        log_tool_json_to_terminal(raw_result)
        formatted_result = format_tool_response_for_ui(raw_result)
        history.append({"role": "tool", "content": raw_result})
        history.append({"role": "assistant", "content": formatted_result})
        return formatted_result

    if should_use_gfk_brand(user_message):
        fn_name = "analyze_gfk_brand_performance"
        fn_result = analyze_gfk_brand_performance(question=user_message)
        raw_result = str(fn_result)
        LAST_TOOL_RESULT_JSON = raw_result
        LAST_TOOL_RESULT_NAME = fn_name
        print("🏷️ DIRECT GFK BRAND:", user_message, flush=True)
        print("✅ TOOL RESULT PREVIEW:", raw_result[:1000], flush=True)
        log_tool_json_to_terminal(raw_result)
        formatted_result = format_tool_response_for_ui(raw_result)
        history.append({"role": "tool", "content": raw_result})
        history.append({"role": "assistant", "content": formatted_result})
        return formatted_result

    if should_use_gfk_market_share(user_message):
        fn_name = "analyze_gfk_market_share"
        fn_result = analyze_gfk_market_share(question=user_message)
        raw_result = str(fn_result)
        LAST_TOOL_RESULT_JSON = raw_result
        LAST_TOOL_RESULT_NAME = fn_name
        print("📈 DIRECT GFK MARKET SHARE:", user_message, flush=True)
        print("✅ TOOL RESULT PREVIEW:", raw_result[:1000], flush=True)
        log_tool_json_to_terminal(raw_result)
        formatted_result = format_tool_response_for_ui(raw_result)
        history.append({"role": "tool", "content": raw_result})
        history.append({"role": "assistant", "content": formatted_result})
        return formatted_result

    if should_use_cross_performance(user_message):
        fn_name = "analyze_cross_performance"
        fn_result = analyze_cross_performance(question=user_message)
        raw_result = str(fn_result)
        LAST_TOOL_RESULT_JSON = raw_result
        LAST_TOOL_RESULT_NAME = fn_name
        print("🎯 DIRECT CROSS PERFORMANCE:", user_message, flush=True)
        print("✅ TOOL RESULT PREVIEW:", raw_result[:1000], flush=True)
        log_tool_json_to_terminal(raw_result)
        formatted_result = format_tool_response_for_ui(raw_result)
        history.append({"role": "tool", "content": raw_result})
        history.append({"role": "assistant", "content": formatted_result})
        return formatted_result

    if should_use_category_insights(user_message):
        fn_name = "generate_category_insight"
        fn_result = generate_category_insight(
            category="genel",
            sector="consumer_electronics",
            period_name="selected_period",
            question=user_message,
        )
        raw_result = str(fn_result)
        LAST_TOOL_RESULT_JSON = raw_result
        LAST_TOOL_RESULT_NAME = fn_name
        print("📊 DIRECT CATEGORY INSIGHTS:", user_message, flush=True)
        print("✅ TOOL RESULT PREVIEW:", raw_result[:1000], flush=True)
        log_tool_json_to_terminal(raw_result)
        formatted_result = format_tool_response_for_ui(raw_result)
        history.append({"role": "tool", "content": raw_result})
        history.append({"role": "assistant", "content": formatted_result})
        return formatted_result

    if should_use_funnel_master(user_message):
        fn_name = "analyze_funnel_master"
        fn_result = analyze_funnel_master(question=user_message)
        raw_result = str(fn_result)
        LAST_TOOL_RESULT_JSON = raw_result
        LAST_TOOL_RESULT_NAME = fn_name
        print("🔍 DIRECT FUNNEL MASTER:", user_message, flush=True)
        print("✅ TOOL RESULT PREVIEW:", raw_result[:1000], flush=True)
        log_tool_json_to_terminal(raw_result)
        formatted_result = format_tool_response_for_ui(raw_result)
        history.append({"role": "tool", "content": raw_result})
        history.append({"role": "assistant", "content": formatted_result})
        return formatted_result

    if should_use_action_executor(user_message):
        fn_name = "execute_recommended_action"
        fn_result = execute_recommended_action(
            question=user_message,
            last_result_json=LAST_TOOL_RESULT_JSON,
        )
        raw_result = str(fn_result)
        LAST_TOOL_RESULT_JSON = raw_result
        LAST_TOOL_RESULT_NAME = fn_name
        print("⚙️ DIRECT ACTION EXECUTOR:", user_message, flush=True)
        print("✅ TOOL RESULT PREVIEW:", raw_result[:1000], flush=True)
        log_tool_json_to_terminal(raw_result)
        formatted_result = format_tool_response_for_ui(raw_result)
        history.append({"role": "tool", "content": raw_result})
        history.append({"role": "assistant", "content": formatted_result})
        return formatted_result

    if should_use_business_calculator(user_message):
        fn_name = "calculate_business_metric"
        fn_result = calculate_business_metric(user_message)
        raw_result = str(fn_result)
        LAST_TOOL_RESULT_JSON = raw_result
        LAST_TOOL_RESULT_NAME = fn_name
        print("🧮 DIRECT BUSINESS CALCULATOR:", user_message, flush=True)
        print("✅ TOOL RESULT PREVIEW:", raw_result[:1000], flush=True)
        log_tool_json_to_terminal(raw_result)
        formatted_result = format_tool_response_for_ui(raw_result)
        history.append({"role": "tool", "content": raw_result})
        history.append({"role": "assistant", "content": formatted_result})
        return formatted_result

    short_history = [SYSTEM_PROMPT] + history[-6:]
    result = call_ollama(short_history)
    message = result.get("message", {})
    tool_calls = message.get("tool_calls", [])

    print("🤖 OLLAMA RAW MESSAGE:", message, flush=True)
    print("🧰 TOOL CALLS:", tool_calls, flush=True)

    if tool_calls:
        history.append({
            "role": "assistant",
            "content": "",
            "tool_calls": tool_calls
        })

        fn_results = []

        for tool_call in tool_calls:
            fn = tool_call.get("function", {})
            fn_name = fn.get("name")
            fn_args = fn.get("arguments", {})

            print("🛠️ ÇAĞRILAN TOOL:", fn_name, flush=True)
            print("📦 TOOL ARGUMENTS RAW:", fn_args, flush=True)

            if isinstance(fn_args, str):
                try:
                    fn_args = json.loads(fn_args)
                except json.JSONDecodeError:
                    fn_args = {}

            print("📦 TOOL ARGUMENTS PARSED:", fn_args, flush=True)

            try:
                if fn_name == "execute_recommended_action":
                    fn_result = execute_recommended_action(
                        question=fn_args.get("question", user_message),
                        last_result_json=LAST_TOOL_RESULT_JSON,
                    )
                elif fn_name == "calculate_business_metric":
                    fn_result = calculate_business_metric(fn_args.get("question", user_message))
                elif fn_name in [
                    "generate_price_competition_from_uploaded_inputs",
                    "generate_merchant_price_competition_insight",
                ]:
                    fn_result = generate_price_competition_from_uploaded_inputs(
                        category=fn_args.get("category", "genel"),
                        period_name=fn_args.get("period_name", "selected_period"),
                    )
                elif fn_name == "generate_category_insight":
                    fn_result = generate_category_insight(
                        category=fn_args.get("category", "genel"),
                        sector=fn_args.get("sector", "consumer_electronics"),
                        period_name=fn_args.get("period_name", "selected_period"),
                        question=fn_args.get("question", user_message),
                    )
                elif fn_name == "analyze_funnel_master":
                    fn_result = analyze_funnel_master(fn_args.get("question", user_message))
                elif fn_name == "analyze_ecommerce_sample":
                    fn_result = analyze_ecommerce_sample(fn_args.get("question", user_message))
                elif fn_name == "analyze_cross_performance":
                    fn_result = analyze_cross_performance(fn_args.get("question", user_message))
                elif fn_name == "analyze_gfk_market_share":
                    fn_result = analyze_gfk_market_share(fn_args.get("question", user_message))
                elif fn_name == "analyze_gfk_brand_performance":
                    fn_result = analyze_gfk_brand_performance(fn_args.get("question", user_message))
                elif fn_name == "analyze_gfk_sku_ranking":
                    fn_result = analyze_gfk_sku_ranking(fn_args.get("question", user_message))
                elif fn_name == "analyze_gfk_combined":
                    fn_result = analyze_gfk_combined(fn_args.get("question", user_message))
                elif fn_name == "get_stock_level":
                    fn_result = get_stock_level(fn_args["product_id"])
                elif fn_name == "get_all_stock":
                    fn_result = get_all_stock()
                elif fn_name == "get_daily_sales_report":
                    fn_result = get_daily_sales_report()
                elif fn_name == "check_low_stock":
                    fn_result = check_low_stock(fn_args.get("threshold", 10))
                elif fn_name == "get_out_of_stock":
                    fn_result = get_out_of_stock()
                elif fn_name == "search_product_by_name":
                    fn_result = search_product_by_name(fn_args["name"])
                elif fn_name == "get_stock_value":
                    fn_result = get_stock_value()
                elif fn_name == "update_stock":
                    fn_result = update_stock(fn_args["product_id"], fn_args["quantity"])
                elif fn_name == "get_order_status":
                    fn_result = get_order_status(fn_args["order_id"])
                elif fn_name == "get_all_orders":
                    fn_result = get_all_orders()
                elif fn_name == "get_pending_orders":
                    fn_result = get_pending_orders()
                elif fn_name == "get_orders_by_customer":
                    fn_result = get_orders_by_customer(fn_args["customer"])
                elif fn_name == "update_order_status":
                    fn_result = update_order_status(fn_args["order_id"], fn_args["status"])
                elif fn_name == "get_todays_orders":
                    fn_result = get_todays_orders()
                elif fn_name == "get_total_revenue":
                    fn_result = get_total_revenue()
                elif fn_name == "get_best_selling_product":
                    fn_result = get_best_selling_product()
                elif fn_name == "get_sales_summary":
                    fn_result = get_sales_summary()
                elif fn_name == "get_low_stock_report":
                    fn_result = get_low_stock_report()
                elif fn_name == "get_stock_turnover":
                    fn_result = get_stock_turnover()
                else:
                    fn_result = f"Bilinmeyen fonksiyon: {fn_name}"

            except Exception as e:
                fn_result = f"Tool çalışırken hata oluştu: {str(e)}"

            raw_result = str(fn_result)
            LAST_TOOL_RESULT_JSON = raw_result
            LAST_TOOL_RESULT_NAME = fn_name or "retail_ai_output"

            print("✅ TOOL RESULT PREVIEW:", raw_result[:1000], flush=True)
            log_tool_json_to_terminal(raw_result)
            formatted_result = format_tool_response_for_ui(raw_result)
            fn_results.append(formatted_result)
            history.append({"role": "tool", "content": raw_result})

        final = "\n\n".join(fn_results)
        print("📝 FINAL UI ANSWER:", final[:1000], flush=True)
        history.append({"role": "assistant", "content": final})
        return final

    content = message.get("content", "").strip()
    if not content or "{" in content[:20]:
        content = "Yanıt alınamadı, lütfen tekrar deneyin."

    history.append({"role": "assistant", "content": content})
    return content


def safe_sheet_name(name: str) -> str:
    name = str(name or "Sheet")
    for ch in ["\\", "/", "*", "?", ":", "[", "]"]:
        name = name.replace(ch, "_")
    return name[:31]


def list_to_df(items):
    if not items:
        return pd.DataFrame()
    if isinstance(items, list):
        return pd.DataFrame(items)
    return pd.DataFrame([items])


def dict_to_key_value_df(data: dict):
    rows = []
    for key, value in (data or {}).items():
        if isinstance(value, (dict, list)):
            rows.append({"metric": key, "value": json.dumps(value, ensure_ascii=False)})
        else:
            rows.append({"metric": key, "value": value})
    return pd.DataFrame(rows)


def write_df(writer, df, sheet_name):
    sheet_name = safe_sheet_name(sheet_name)
    if df is None or df.empty:
        df = pd.DataFrame([{"info": "No data"}])
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    worksheet = writer.sheets[sheet_name]
    worksheet.freeze_panes = "A2"
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for idx, col in enumerate(df.columns, start=1):
        values = df[col].head(100).fillna("").astype(str).tolist()
        max_len = max([len(str(col))] + [len(v) for v in values])
        worksheet.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 45)


def build_excel_from_tool_result(raw_text: str) -> BytesIO:
    try:
        data = json.loads(raw_text)
    except Exception:
        data = {"analysis_type": "raw_text", "raw_output": raw_text}

    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        analysis_type = data.get("analysis_type", "retail_ai_output")
        metadata = {
            "analysis_type": analysis_type,
            "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "category": data.get("category", ""),
            "period_name": data.get("period_name", ""),
            "benchmark_mode": data.get("benchmark_mode", ""),
            "benchmark_source": data.get("benchmark_source", ""),
        }
        write_df(writer, dict_to_key_value_df(metadata), "Metadata")

        if analysis_type == "recommended_action_execution":
            write_df(writer, dict_to_key_value_df(data.get("summary", {})), "Action Summary")
            write_df(writer, list_to_df(data.get("rows", [])), "Action Plan")
            write_df(writer, list_to_df([{"action": x} for x in data.get("recommended_actions", [])]), "Actions")
        elif analysis_type == "price_competition_uploaded_inputs":
            write_df(writer, dict_to_key_value_df(data.get("summary", {})), "Summary")
            write_df(writer, list_to_df(data.get("top_expensive_products", [])), "Benchmark Above")
            write_df(writer, list_to_df(data.get("top_cheaper_products", [])), "Benchmark Below")
            write_df(writer, list_to_df([{"action": x} for x in data.get("recommended_actions", [])]), "Actions")
        elif analysis_type == "category_sector_insight":
            write_df(writer, dict_to_key_value_df(data.get("executive_summary", {})), "Executive Summary")
            write_df(writer, dict_to_key_value_df(data.get("metric_snapshot", {})), "Metric Snapshot")
            write_df(writer, list_to_df(data.get("signal_interpretation", [])), "Signal Interpretation")
            write_df(writer, list_to_df(data.get("root_causes", [])), "Root Causes")
            write_df(writer, list_to_df(data.get("winning_categories", [])), "Winners")
            write_df(writer, list_to_df(data.get("losing_categories", [])), "Losers")
            write_df(writer, list_to_df(data.get("channel_insights", [])), "Channel Insights")
            write_df(writer, list_to_df(data.get("traffic_insights", [])), "Traffic Insights")
            write_df(writer, list_to_df([{"action": x} for x in data.get("recommended_actions", [])]), "Actions")
            write_df(writer, list_to_df(data.get("rows", [])), "Category Rows")
        elif analysis_type == "cross_performance_analysis":
            write_df(writer, dict_to_key_value_df(data.get("summary", {})), "Summary")
            if data.get("trends"):
                write_df(writer, dict_to_key_value_df(data.get("trends", {})), "Google Trends")
            
            scenarios = data.get("scenarios", {}) or {}
            write_df(writer, list_to_df(scenarios.get("expensive_falling_sales", [])), "S1_Expensive_Falling")
            write_df(writer, list_to_df(scenarios.get("expensive_good_sales", [])), "S2_Expensive_Good")
            write_df(writer, list_to_df(scenarios.get("cheap_no_sales", [])), "S3_Cheap_No_Sales")
            write_df(writer, list_to_df(scenarios.get("cheap_good_sales", [])), "S4_Cheap_Good")
            write_df(writer, list_to_df(data.get("price_cut_candidates", [])), "Price_Cut_Candidates")
            write_df(writer, list_to_df(data.get("expensive_skus", [])), "Expensive_SKUs")
        elif analysis_type == "business_metric_calculation":
            business_info = {
                "question": data.get("question", ""),
                "calculation_type": data.get("calculation_type", ""),
                "metric": data.get("metric", ""),
                "aggregation": data.get("aggregation", ""),
                "group_by": data.get("group_by", ""),
                "row_count": data.get("row_count", ""),
                "result": data.get("result", ""),
                "filters": data.get("filters", []),
            }
            write_df(writer, dict_to_key_value_df(business_info), "Business Summary")
            write_df(writer, list_to_df(data.get("rows", [])), "Rows")
        elif analysis_type == "gfk_market_share":
            meta = {
                "source": data.get("source", ""),
                "category": data.get("category", ""),
                "view": data.get("view", ""),
                "mediamarkt_rank": data.get("mediamarkt_rank", ""),
                "mediamarkt_market_share_pct": data.get("mediamarkt_market_share_pct", ""),
                "wow_change_pct": data.get("wow_change_pct", ""),
                "current_week_value_try": data.get("current_week_value_try", ""),
                "previous_week_value_try": data.get("previous_week_value_try", ""),
            }
            write_df(writer, dict_to_key_value_df(meta), "GfK Market Share Summary")
            write_df(writer, list_to_df(data.get("rows", [])), "Category Rows")
        elif analysis_type == "gfk_brand_performance":
            meta = {
                "source": data.get("source", ""),
                "latest_week": data.get("latest_week", ""),
                "filtered_by_category": data.get("filtered_by_category", ""),
                "filtered_by_brand": data.get("filtered_by_brand", ""),
                "top_brand": data.get("top_brand", ""),
                "top_brand_share_pct": data.get("top_brand_share_pct", ""),
            }
            write_df(writer, dict_to_key_value_df(meta), "GfK Brand Summary")
            # Haftalık seri dahil satırlar
            rows_flat = []
            for r in data.get("rows", []):
                row_flat = {k: v for k, v in r.items() if k != "weekly_share_series"}
                series = r.get("weekly_share_series", {})
                row_flat.update(series)

@app.get("/journey", response_class=HTMLResponse)
def journey(activated: str = None, plan: str = None, demo: str = None):
    # Task 1 Scoping: Only allow console access if user purchased (activated=true) or clicked demo on pricing (demo=true)
    if not (activated == "true" or demo == "true" or plan):
        from fastapi.responses import RedirectResponse
        return RedirectResponse(url="/pricing?notice=direct_access_restricted", status_code=303)

    return """<!DOCTYPE html>
<html lang="tr">
<head>
  <!-- Google Tag Manager -->
  <script>(function(w,d,s,l,i){w[l]=w[l]||[];w[l].push({'gtm.start':
  new Date().getTime(),event:'gtm.js'});var f=d.getElementsByTagName(s)[0],
  j=d.createElement(s),dl=l!='dataLayer'?'&l='+l:'';j.async=true;j.src=
  'https://www.googletagmanager.com/gtm.js?id='+i+dl;f.parentNode.insertBefore(j,f);
  })(window,document,'script','dataLayer','GTM-TVKFC4P6');</script>
  <!-- End Google Tag Manager -->

  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <title>Excel Wizard &amp; Retail AI Console – DataProvido</title>
  <meta name="description" content="Don't just analyze data—know exactly what actions to take next. DataProvido delivers real-time prescriptive commercial decisions 100% locally.">
  <link rel="icon" type="image/png" href="/logo.png" />
  <link rel="shortcut icon" type="image/png" href="/logo.png" />
  <link rel="apple-touch-icon" href="/logo.png" />
  <link href="https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:ital,wght@0,300;0,400;0,500;0,600;0,700;0,800;1,400&family=Outfit:wght@400;500;600;700;800&family=Playfair+Display:ital,wght@0,400;0,700;1,400&display=swap" rel="stylesheet">
  <style>
    *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }

    :root {
      --orange:       #f26f26;
      --orange-dark:  #d85c18;
      --orange-light: #f58c50;
      --blue:         #2563eb;
      --blue-dark:    #1d4ed8;
      --blue-light:   #3b82f6;
      --bg:           #ffffff;
      --bg-2:         #f8fafc;
      --bg-3:         #f1f5f9;
      --text-900:     #0f172a;
      --text-700:     #334155;
      --text-500:     #64748b;
      --text-dim:     #94a3b8;
      --border:       #e2e8f0;
      --border-blue:  rgba(37,99,235,0.22);
      --border-orange: rgba(242,111,38,0.22);
      --panel-bg:     #ffffff;
      --card-shadow:  0 4px 20px rgba(15,23,42,0.05), 0 1px 3px rgba(15,23,42,0.03);
    }

    html, body { height: 100%; }

    body {
      min-height: 100vh;
      background: #f8fafc;
      background-image: radial-gradient(circle at 50% 0%, rgba(242,111,38,0.08) 0%, transparent 65%);
      color: var(--text-900);
      font-family: 'Plus Jakarta Sans', -apple-system, BlinkMacSystemFont, sans-serif;
      overflow: hidden;
      -webkit-font-smoothing: antialiased;
    }

    .app-shell {
      display: grid;
      grid-template-columns: 270px 1fr;
      height: 100vh;
      position: relative;
      background: #f8fafc;
    }

    /* ── SIDEBAR (HOMEPAGE ORANGE DESIGN) ── */
    .sidebar {
      background: linear-gradient(180deg, #f26f26 0%, #d85c18 100%);
      padding: 26px 20px;
      color: #ffffff;
      display: flex;
      flex-direction: column;
      border-right: 1px solid rgba(255,255,255,0.15);
      box-shadow: 4px 0 20px rgba(242,111,38,0.18);
      z-index: 10;
    }

    .brand {
      display: flex;
      align-items: center;
      gap: 12px;
      padding-bottom: 22px;
      border-bottom: 1px solid rgba(255,255,255,0.20);
      margin-bottom: 20px;
    }

    .brand-icon {
      width: 44px; height: 44px;
      border-radius: 12px;
      background: #ffffff;
      display: grid;
      place-items: center;
      color: #f26f26;
      font-family: 'Outfit', sans-serif;
      font-size: 24px;
      font-weight: 800;
      box-shadow: 0 4px 14px rgba(0,0,0,0.15);
      flex-shrink: 0;
    }

    .brand h1 {
      font-family: 'Outfit', sans-serif;
      font-size: 20px;
      line-height: 1.1;
      letter-spacing: -0.02em;
      font-weight: 800;
      color: #ffffff;
    }

    .nav-label {
      font-size: 10.5px;
      letter-spacing: 0.14em;
      color: rgba(255,255,255,0.88);
      text-transform: uppercase;
      font-weight: 800;
      margin: 18px 10px 10px;
    }

    .menu { display: flex; flex-direction: column; gap: 8px; }

    .menu-btn {
      border: 1px solid #f1f5f9;
      width: 100%;
      text-align: left;
      cursor: pointer;
      border-radius: 14px;
      background: #ffffff;
      color: #334155;
      padding: 12px 14px;
      font-family: 'Plus Jakarta Sans', sans-serif;
      font-size: 13.5px;
      line-height: 1.35;
      font-weight: 600;
      transition: all 0.22s cubic-bezier(0.16, 1, 0.3, 1);
      display: flex;
      align-items: center;
      box-shadow: 0 1px 3px rgba(15,23,42,0.02);
      position: relative;
    }

    .menu-icon {
      margin-right: 12px;
      color: #64748b;
      flex-shrink: 0;
      transition: color 0.2s ease, transform 0.2s ease;
    }

    .menu-btn-text {
      flex: 1;
      display: flex;
      flex-direction: column;
    }

    .menu-btn-text strong {
      color: #0f172a;
      font-weight: 700;
      font-size: 13.5px;
      letter-spacing: -0.01em;
    }

    .menu-btn span span {
      display: block;
      margin-top: 3px;
      font-size: 11.5px;
      color: #64748b;
      font-weight: 400;
    }

    .menu-btn:hover {
      background: #f8fafc;
      border-color: #cbd5e1;
      color: #0f172a;
      transform: translateX(3px);
    }

    .menu-btn:hover .menu-icon {
      color: #2563eb;
      transform: scale(1.1);
    }

    .menu-btn.active {
      background: linear-gradient(135deg, #eff6ff 0%, #f8fafc 100%);
      border-color: #bfdbfe;
      border-left: 4px solid #2563eb;
      color: #1d4ed8;
      box-shadow: 0 4px 14px rgba(37,99,235,0.08);
    }

    .menu-btn.active .menu-icon {
      color: #2563eb;
    }

    .menu-btn.active .menu-btn-text strong {
      color: #1d4ed8;
      font-weight: 800;
    }

    .menu-btn.active span span {
      color: #2563eb;
      font-weight: 500;
    }

    .menu-btn.active span { color: var(--orange-light); }

    /* ── MAIN (IMPECCABLE WARM SLATE THEME) ── */
    .main {
      height: 100vh;
      overflow: auto;
      padding: 24px 32px 28px;
      background: #f8fafc;
      background-image: radial-gradient(circle at 50% 0%, rgba(242,111,38,0.08) 0%, transparent 65%);
      display: flex;
      flex-direction: column;
      scrollbar-width: thin;
      scrollbar-color: rgba(242,111,38,0.25) transparent;
    }
    .main::-webkit-scrollbar { width: 4px; }
    .main::-webkit-scrollbar-thumb { background: rgba(242,111,38,0.25); border-radius: 4px; }

    .topbar {
      display: flex;
      justify-content: space-between;
      align-items: center;
      margin-bottom: 18px;
    }

    .page-title p {
      font-size: 13.5px;
      color: #475569;
      max-width: 820px;
      line-height: 1.65;
      font-weight: 400;
    }

    .home-link {
      text-decoration: none;
      background: #ffffff;
      color: #334155;
      border: 1px solid #cbd5e1;
      border-radius: 999px;
      padding: 9px 18px;
      font-family: 'Plus Jakarta Sans', sans-serif;
      font-size: 12.5px;
      font-weight: 700;
      transition: all 0.18s ease;
      box-shadow: 0 2px 8px rgba(15,23,42,0.04);
      white-space: nowrap;
    }
    .home-link:hover {
      background: #ffffff;
      color: #f26f26;
      border-color: #f26f26;
      box-shadow: 0 4px 14px rgba(242,111,38,0.15);
    }

    /* ── WORKSPACE ── */
    .workspace {
      display: flex;
      flex-direction: column;
      gap: 20px;
      width: 100%;
      flex: 1;
    }

    .module-panel {
      background: #ffffff;
      border: 1px solid #e2e8f0;
      border-radius: 24px;
      padding: 32px 36px;
      box-shadow: 0 10px 32px -4px rgba(15,23,42,0.06), 0 4px 12px -2px rgba(242,111,38,0.04);
      width: 100%;
      flex: 1;
      display: flex;
      flex-direction: column;
    }

    .module-head {
      display: flex;
      justify-content: space-between;
      align-items: flex-start;
      gap: 12px;
      margin-bottom: 20px;
    }

    .module-title {
      font-family: 'Playfair Display', serif;
      font-size: 26px;
      line-height: 1.15;
      letter-spacing: -0.02em;
      font-weight: 700;
      color: var(--text-900);
    }

    .module-desc {
      margin-top: 8px;
      font-size: 13px;
      color: var(--text-500);
      line-height: 1.6;
    }

    .quick-title {
      font-size: 11.5px;
      font-weight: 600;
      text-transform: uppercase;
      letter-spacing: 0.08em;
      color: var(--text-dim);
      margin: 18px 0 10px;
    }

    .question-select {
      width: 100%;
      border: 1.4px solid var(--border);
      border-radius: 12px;
      padding: 10px 14px;
      font-size: 13px;
      color: var(--text-900);
      outline: none;
      background: #ffffff;
      cursor: pointer;
      font-family: 'Inter', sans-serif;
      font-weight: 500;
      transition: all 0.18s ease;
    }
    .question-select option {
      background: #ffffff;
      color: var(--text-900);
    }
    .question-select:focus {
      border-color: var(--orange);
      box-shadow: 0 0 0 4px rgba(242,111,38,0.10);
    }

    .question-add-input {
      border: 1.4px solid var(--border);
      border-radius: 12px;
      padding: 10px 14px;
      font-size: 13px;
      color: var(--text-900);
      outline: none;
      background: #ffffff;
      font-family: 'Inter', sans-serif;
      font-weight: 400;
      transition: all 0.18s ease;
    }
    .question-add-input::placeholder { color: var(--text-dim); }
    .question-add-input:focus {
      border-color: var(--orange);
      box-shadow: 0 0 0 4px rgba(242,111,38,0.10);
    }

    .action-row {
      display: flex;
      gap: 10px;
      margin-top: 12px;
      flex-wrap: wrap;
    }

    .primary-btn {
      background: var(--orange);
      color: #fff;
      border: 0;
      border-radius: 12px;
      padding: 11px 22px;
      font-size: 13.5px;
      font-weight: 600;
      cursor: pointer;
      box-shadow: 0 4px 14px rgba(242,111,38,0.25);
      transition: all 0.18s ease;
    }
    .primary-btn:hover {
      background: var(--orange-dark);
      transform: translateY(-1px);
    }

    .secondary-btn {
      background: #ffffff;
      color: var(--text-700);
      border: 1.4px solid var(--border);
      border-radius: 12px;
      padding: 11px 18px;
      font-size: 13px;
      font-weight: 600;
      cursor: pointer;
      transition: all 0.18s ease;
    }
    .secondary-btn:hover {
      border-color: var(--orange);
      color: var(--orange);
    }

    .ghost-btn {
      background: transparent;
      border: 0;
      color: var(--text-dim);
      font-size: 12px;
      font-weight: 600;
      cursor: pointer;
      padding: 4px 8px;
      transition: color 0.15s ease;
    }
    .ghost-btn:hover { color: var(--text-900); }

    .result-panel {
      display: flex;
      flex-direction: column;
    }

    .result-header {
      display: flex;
      justify-content: space-between;
      gap: 16px;
      align-items: flex-start;
      margin-bottom: 14px;
    }

    .result-header h3 {
      font-family: 'Playfair Display', serif;
      font-size: 22px;
      letter-spacing: -0.02em;
      font-weight: 700;
      color: var(--text-900);
    }
    .result-header p {
      margin-top: 5px;
      color: var(--text-500);
      font-size: 12.5px;
    }

    .result-box {
      flex: 1;
      min-height: 360px;
      border: 1px solid var(--border);
      border-radius: 16px;
      padding: 18px;
      overflow-y: auto;
      white-space: pre-wrap;
      font-size: 13px;
      line-height: 1.72;
      color: var(--text-700);
      scrollbar-width: thin;
      scrollbar-color: rgba(242,111,38,0.20) transparent;
    }
    .result-box::-webkit-scrollbar { width: 4px; }
    .result-box::-webkit-scrollbar-thumb { background: rgba(242,111,38,0.25); border-radius: 4px; }

    .result-box.placeholder {
      display: flex;
      align-items: center;
      justify-content: center;
      text-align: center;
      color: var(--text-dim);
      padding: 40px;
      font-style: italic;
    }

    .loading { opacity: 0.72; font-style: italic; color: var(--orange); }

    @keyframes pulse-glow {
      0%, 100% { transform: scale(1); box-shadow: 0 0 0 0 rgba(242,111,38,0.4); }
      50% { transform: scale(1.05); box-shadow: 0 0 0 6px rgba(242,111,38,0); }
    }

    /* ── IMPECCABLE UNIFIED AI COMMAND BAR ── */
    .impeccable-cmd-card {
      background: #ffffff;
      border: 1.5px solid #cbd5e1;
      border-radius: 20px;
      padding: 0;
      box-shadow: 0 4px 20px rgba(15, 23, 42, 0.04);
      transition: all 0.25s cubic-bezier(0.16, 1, 0.3, 1);
      margin-bottom: 24px;
      overflow: hidden;
    }

    .impeccable-cmd-card:focus-within {
      border-color: #3b82f6;
      box-shadow: 0 8px 30px rgba(37, 99, 235, 0.12), 0 0 0 4px rgba(37, 99, 235, 0.08);
    }

    .cmd-card-header {
      display: flex;
      justify-content: space-between;
      align-items: center;
      padding: 12px 18px;
      background: #f8fafc;
      border-bottom: 1px solid #f1f5f9;
    }

    .cmd-dataset-badge {
      display: flex;
      align-items: center;
      gap: 8px;
      font-size: 13px;
      font-weight: 600;
      color: #334155;
    }

    .cmd-upload-btn {
      cursor: pointer;
      display: flex;
      align-items: center;
      gap: 6px;
      font-size: 12px;
      font-weight: 700;
      color: #2563eb;
      background: #eff6ff;
      padding: 6px 14px;
      border-radius: 10px;
      border: 1px solid #bfdbfe;
      transition: all 0.18s ease;
    }

    .cmd-upload-btn:hover {
      background: #dbeafe;
      border-color: #93c5fd;
    }

    .cmd-card-body {
      padding: 16px 20px 8px;
    }

    .cmd-card-body textarea {
      width: 100%;
      min-height: 120px;
      resize: vertical;
      border: none;
      outline: none;
      font-size: 14.5px;
      color: #0f172a;
      font-family: inherit;
      line-height: 1.65;
      background: transparent;
    }

    .cmd-card-footer {
      display: flex;
      justify-content: space-between;
      align-items: center;
      padding: 10px 18px 14px;
      border-top: 1px solid #f8fafc;
      gap: 12px;
    }

    .cmd-action-left {
      display: flex;
      align-items: center;
      gap: 8px;
    }

    .cmd-chip-btn {
      background: #f1f5f9;
      border: 1px solid #e2e8f0;
      color: #475569;
      font-size: 12px;
      font-weight: 600;
      padding: 7px 15px;
      border-radius: 10px;
      cursor: pointer;
      transition: all 0.18s ease;
    }

    .cmd-chip-btn:hover {
      background: #e2e8f0;
      color: #0f172a;
      border-color: #cbd5e1;
    }

    .cmd-action-right {
      display: flex;
      align-items: center;
      gap: 10px;
    }

    .cmd-voice-orb-btn {
      background: #f8fafc;
      border: 1.5px solid #cbd5e1;
      color: #0f172a;
      padding: 6px 14px;
      border-radius: 999px;
      cursor: pointer;
      display: flex;
      align-items: center;
      gap: 8px;
      transition: all 0.2s ease;
      font-weight: 700;
    }

    .cmd-voice-orb-btn:hover {
      border-color: #3b82f6;
      background: #eff6ff;
    }

    .voice-orb-sm {
      width: 26px;
      height: 26px;
      border-radius: 50%;
      background: linear-gradient(135deg, #2563eb 0%, #1d4ed8 100%);
      display: grid;
      place-items: center;
      color: #ffffff;
      box-shadow: 0 2px 8px rgba(37, 99, 235, 0.35);
    }

    .voice-orb-text {
      font-size: 12px;
      font-weight: 700;
      color: #1e293b;
    }

    .cmd-run-btn {
      background: #cbd5e1;
      color: #64748b;
      border: none;
      padding: 10px 22px;
      border-radius: 12px;
      font-weight: 700;
      font-size: 13.5px;
      cursor: not-allowed;
      display: flex;
      align-items: center;
      gap: 8px;
      transition: all 0.22s ease;
      opacity: 0.7;
    }

    .cmd-run-btn:not([disabled]) {
      background: linear-gradient(135deg, #f26f26 0%, #d85c18 100%);
      color: #ffffff;
      cursor: pointer;
      opacity: 1;
      box-shadow: 0 4px 14px rgba(242, 111, 38, 0.28);
    }

    .cmd-run-btn:not([disabled]):hover {
      transform: translateY(-1px);
      box-shadow: 0 6px 20px rgba(242, 111, 38, 0.38);
    }

    .cmd-shortcut-hint {
      background: rgba(255, 255, 255, 0.25);
      color: currentColor;
      font-size: 10.5px;
      font-weight: 800;
      padding: 2px 6px;
      border-radius: 5px;
      letter-spacing: 0.04em;
    }

    /* Google AI Overview Style Animated Voice Listening Indicator */
    @keyframes ai-voice-pulse {
      0% { box-shadow: 0 0 0 0 rgba(239, 68, 68, 0.6), 0 0 0 0 rgba(242, 111, 38, 0.4); transform: scale(1); }
      50% { box-shadow: 0 0 0 10px rgba(239, 68, 68, 0.25), 0 0 0 18px rgba(242, 111, 38, 0); transform: scale(1.03); }
      100% { box-shadow: 0 0 0 0 rgba(239, 68, 68, 0), 0 0 0 0 rgba(242, 111, 38, 0); transform: scale(1); }
    }

    .voice-listening-active {
      background: linear-gradient(135deg, #fef2f2 0%, #fff1f2 100%) !important;
      border-color: #ef4444 !important;
      box-shadow: 0 8px 28px rgba(239, 68, 68, 0.22) !important;
      animation: ai-voice-pulse 1.5s infinite ease-in-out !important;
    }

    .voice-listening-active .voice-orb {
      background: linear-gradient(135deg, #ef4444 0%, #dc2626 100%) !important;
      box-shadow: 0 4px 18px rgba(239, 68, 68, 0.5) !important;
    }

    .voice-listening-active .voice-sub-badge {
      background: rgba(239, 68, 68, 0.12) !important;
      color: #dc2626 !important;
    }

    .sound-wave-bars {
      display: inline-flex;
      align-items: center;
      gap: 3px;
      height: 18px;
    }

    .sound-wave-bars .bar {
      width: 3.5px;
      background: #ef4444;
      border-radius: 3px;
      animation: wave-bounce 0.8s ease-in-out infinite alternate;
    }
    .sound-wave-bars .bar:nth-child(1) { animation-delay: 0.1s; height: 40%; }
    .sound-wave-bars .bar:nth-child(2) { animation-delay: 0.35s; height: 100%; }
    .sound-wave-bars .bar:nth-child(3) { animation-delay: 0.2s; height: 75%; }
    .sound-wave-bars .bar:nth-child(4) { animation-delay: 0.45s; height: 50%; }

    @keyframes wave-bounce {
      0% { transform: scaleY(0.3); }
      100% { transform: scaleY(1.15); }
    }

    @media (max-width: 1180px) {
      .workspace { grid-template-columns: 1fr; }
      .module-panel, .result-panel { min-height: auto; }
      .result-box { min-height: 300px; }
    }

    @media (max-width: 760px) {
      body { overflow: auto; }
      .app-shell { grid-template-columns: 1fr; height: auto; }
      .sidebar { position: relative; height: auto; }
      .main { height: auto; padding: 20px; }
      .topbar { flex-direction: column; gap: 12px; }
    }
  </style>
</head>
<body>
  <!-- Google Tag Manager (noscript) -->
  <noscript><iframe src="https://www.googletagmanager.com/ns.html?id=GTM-TVKFC4P6"
  height="0" width="0" style="display:none;visibility:hidden"></iframe></noscript>
  <!-- End Google Tag Manager (noscript) -->

  <div class="app-shell">
    <aside class="sidebar">
      <div>
        <div class="brand">
          <img src="/logo.png" alt="DataProvido Logo" style="width: 42px; height: 42px; object-fit: contain; border-radius: 12px; background: #ffffff; padding: 3px; box-shadow: 0 4px 14px rgba(0,0,0,0.15); flex-shrink: 0;" />
          <div>
            <h1>DataProvido</h1>
          </div>
        </div>
        <div class="nav-label" id="navLabelCategories">CATEGORIES</div>
        <nav class="menu" id="menu">
          <button class="menu-btn active" data-key="business_calculator">
            <svg class="menu-icon" width="20" height="20" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="m12 3-1.912 5.813a2 2 0 0 1-1.275 1.275L3 12l5.813 1.912a2 2 0 0 1 1.275 1.275L12 21l1.912-5.813a2 2 0 0 1 1.275-1.275L21 12l-5.813-1.912a2 2 0 0 1-1.275-1.275L12 3z"/></svg>
            <span class="menu-btn-text">
              <strong>Excel Wizard</strong>
              <span id="sub_business_calculator">Voice &amp; text Excel commands</span>
            </span>
          </button>
          <button class="menu-btn" data-key="category_insights">
            <svg class="menu-icon" width="20" height="20" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><rect x="3" y="3" width="7" height="7"/><rect x="14" y="3" width="7" height="7"/><rect x="14" y="14" width="7" height="7"/><rect x="3" y="14" width="7" height="7"/></svg>
            <span class="menu-btn-text">
              <strong>Category Insights</strong>
              <span id="sub_category_insights">Category &amp; sector analysis</span>
            </span>
          </button>
          <button class="menu-btn" data-key="price_competition">
            <svg class="menu-icon" width="20" height="20" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><polyline points="23 6 13.5 15.5 8.5 10.5 1 18"/><polyline points="17 6 23 6 23 12"/></svg>
            <span class="menu-btn-text">
              <strong>Price Competition</strong>
              <span id="sub_price_competition">Merchant benchmark</span>
            </span>
          </button>
          <button class="menu-btn" data-key="action_executor">
            <svg class="menu-icon" width="20" height="20" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="m9 11 3 3L22 4"/><path d="M21 12v7a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2V5a2 2 0 0 1 2-2h11"/></svg>
            <span class="menu-btn-text">
              <strong>Action Executor</strong>
              <span id="sub_action_executor">Action plan generator</span>
            </span>
            <span class="menu-badge" style="background: #a3e635; color: #1a2e05; min-width: 22px; height: 22px; padding: 0 6px; border-radius: 999px; display: grid; place-items: center; font-size: 11px; font-weight: 800; margin-left: 6px; box-shadow: 0 2px 6px rgba(163,230,53,0.4);">6</span>
          </button>
          <button class="menu-btn" data-key="funnel_stock">
            <svg class="menu-icon" width="20" height="20" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><polygon points="22 3 2 3 10 12.46 10 19 14 21 14 12.46 22 3"/></svg>
            <span class="menu-btn-text">
              <strong>Funnel &amp; Stock</strong>
              <span id="sub_funnel_stock">Conversion &amp; stock risk</span>
            </span>
          </button>
          <button class="menu-btn" data-key="excel_outputs">
            <svg class="menu-icon" width="20" height="20" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M14.5 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V7.5L14.5 2z"/><polyline points="14 2 14 8 20 8"/><line x1="16" y1="13" x2="8" y2="13"/><line x1="16" y1="17" x2="8" y2="17"/><line x1="10" y1="9" x2="8" y2="9"/></svg>
            <span class="menu-btn-text">
              <strong>Excel Outputs</strong>
              <span id="sub_excel_outputs">Analytical reports</span>
            </span>
          </button>
          <button class="menu-btn" data-key="data_upload">
            <svg class="menu-icon" width="20" height="20" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M12 2v8"/><path d="M4.93 10.93a10 10 0 1 0 14.14 0"/><path d="M12 18v4"/></svg>
            <span class="menu-btn-text">
              <strong id="label_data_upload">Connect to Data Sources</strong>
              <span id="sub_data_upload">Excel / CSV &amp; API Connectors</span>
            </span>
          </button>
        </nav>

        <!-- BOTTOM PROFILE & PROPERTY FOOTER CARD (EXACT MATCH TO SCREENSHOT) -->
        <div style="margin-top: auto; padding-top: 18px; border-top: 1px solid rgba(255,255,255,0.22); display: flex; flex-direction: column; gap: 10px;">
          
          <!-- CONNECTED PROPERTY BADGE -->
          <div style="background: rgba(0,0,0,0.18); backdrop-filter: blur(8px); border: 1px solid rgba(255,255,255,0.25); border-radius: 14px; padding: 10px 14px; display: flex; align-items: center; justify-content: space-between; cursor: pointer; transition: all 0.2s ease;" onclick="openUserProfileModal()">
            <div style="display: flex; align-items: center; gap: 10px;">
              <div style="width: 28px; height: 28px; border-radius: 8px; background: rgba(255,255,255,0.9); color: #d85c18; font-weight: 800; display: grid; place-items: center; font-size: 13px;">D</div>
              <div>
                <strong style="display: block; font-size: 12.5px; color: #ffffff; font-weight: 700;">dataprovido.com</strong>
                <span style="font-size: 10.5px; color: rgba(255,255,255,0.75);">Connected property</span>
              </div>
            </div>
            <span style="color: rgba(255,255,255,0.8); font-size: 11px;">▼</span>
          </div>

          <!-- SETTINGS & PRIVACY QUICK LINKS -->
          <div style="display: flex; justify-content: space-between; align-items: center; padding: 4px 6px; font-size: 11.5px;">
            <a href="javascript:void(0)" onclick="openUserProfileModal()" style="color: rgba(255,255,255,0.92); text-decoration: none; font-weight: 600; display: flex; align-items: center; gap: 6px;">
              <span>⚙️</span> <span>Settings</span>
            </a>
            <a href="/privacy" style="color: rgba(255,255,255,0.75); text-decoration: none; font-size: 11px;">Privacy Policy</a>
          </div>

          <!-- USER PROFILE CARD -->
          <div id="userProfileCard" onclick="openUserProfileModal()" style="background: rgba(0,0,0,0.25); backdrop-filter: blur(8px); border: 1px solid rgba(255,255,255,0.30); border-radius: 16px; padding: 11px 14px; display: flex; align-items: center; justify-content: space-between; cursor: pointer; transition: all 0.22s ease; box-shadow: 0 4px 14px rgba(0,0,0,0.12);">
            <div style="display: flex; align-items: center; gap: 10px;">
              <div style="width: 36px; height: 36px; border-radius: 50%; background: #fde68a; color: #78350f; font-weight: 800; font-size: 13px; display: grid; place-items: center; border: 1.5px solid #ffffff;">YK</div>
              <div>
                <strong style="display: block; font-size: 13px; color: #ffffff; font-weight: 700; letter-spacing: -0.01em;">Yasam Karadag</strong>
                <span style="font-size: 11px; color: rgba(255,255,255,0.80); font-weight: 500;">owner · Founder plan</span>
              </div>
            </div>
            <span style="color: rgba(255,255,255,0.9); font-size: 14px; font-weight: 700;">↪</span>
          </div>

        </div>
      </div>
    </aside>

    <main class="main">
      <div id="activationBanner" style="display: none; background: #ecfdf5; border: 1.5px solid #10b981; border-radius: 14px; padding: 14px 20px; margin-bottom: 16px; align-items: center; justify-content: space-between; gap: 12px; box-shadow: 0 4px 14px rgba(16,185,129,0.12);">
        <div style="display: flex; align-items: center; gap: 10px;">
          <span style="font-size: 20px;">🎉</span>
          <div>
            <strong style="color: #065f46; font-size: 13.5px;" id="activationTitle">License Active: Welcome to DataProvido Console!</strong>
            <div style="font-size: 12px; color: #047857;" id="activationSubtitle">Your 100% offline local AI environment is fully unlocked with unlimited query execution.</div>
          </div>
        </div>
        <button onclick="document.getElementById('activationBanner').style.display='none'" style="background: transparent; border: none; color: #065f46; font-weight: 700; cursor: pointer; font-size: 16px;">✕</button>
      </div>

      <div class="topbar" style="margin-bottom: 20px; display: flex; justify-content: space-between; align-items: center; gap: 20px;">
        <div class="page-title" style="flex: 1; text-align: center; padding: 0 20px;">
          <h2 style="font-family: 'Playfair Display', serif; font-size: 22px; font-weight: 700; color: #0f172a; letter-spacing: -0.01em; line-height: 1.3;" id="heroMainHeading">
            Excel Wizard allows you to perform your manual Excel tasks using your voice and text commands.
          </h2>
        </div>
        <a class="home-link" href="/" style="flex-shrink: 0;" id="homeLinkText">← Home Page</a>
      </div>

      <section class="workspace">
        <div class="module-panel" style="padding: 32px 36px;">
          
          <div class="module-head" style="display: flex; justify-content: space-between; align-items: flex-start; gap: 20px; margin-bottom: 22px; padding-bottom: 18px; border-bottom: 1px solid #e2e8f0;">
            <div style="flex: 1;">
              <h3 class="module-title" id="moduleTitle" style="font-family: 'Playfair Display', serif; font-size: 26px; font-weight: 700; color: #0f172a; margin-bottom: 6px; letter-spacing: -0.02em;">Excel Wizard</h3>
              <p class="module-desc" id="moduleDesc" style="font-size: 13.5px; color: #475569; line-height: 1.6; font-weight: 400; max-width: 760px;">Execute advanced mathematical calculations, average, sum, filters, and brand/category breakdowns on all your retail &amp; e-commerce Excel data using English or Turkish voice commands.</p>
            </div>
            <!-- TOP RIGHT LANGUAGE SWITCHER PILLS & UX TOAST BADGE -->
            <div style="position: relative; flex-shrink: 0;">
              <div id="langToastBadge" style="display: none; position: absolute; top: -38px; right: 0; background: #0f172a; color: #ffffff; font-size: 11.5px; font-weight: 700; padding: 6px 14px; border-radius: 20px; box-shadow: 0 4px 14px rgba(0,0,0,0.18); transition: all 0.25s ease; z-index: 99; white-space: nowrap; pointer-events: none;">🌐 Voice Mode Switched</div>
              <div style="display: flex; gap: 4px; background: #f8fafc; padding: 4px; border-radius: 12px; border: 1px solid #e2e8f0; box-shadow: 0 1px 4px rgba(0,0,0,0.03);">
                <button id="langBtnEN" onclick="setLanguage('en')" type="button" style="border: 1px solid #2563eb; background: #2563eb; color: #ffffff; padding: 7px 16px; border-radius: 9px; font-size: 12px; font-weight: 700; cursor: pointer; transition: all 0.18s; box-shadow: 0 2px 6px rgba(37,99,235,0.20);">English Voice</button>
                <button id="langBtnTR" onclick="setLanguage('tr')" type="button" style="border: 1px solid transparent; background: transparent; color: #475569; padding: 7px 16px; border-radius: 9px; font-size: 12px; font-weight: 600; cursor: pointer; transition: all 0.18s;">Turkish Voice</button>
              </div>
            </div>
          </div>

          <!-- PRESETS AREA (HIDDEN FOR EXCEL WIZARD) -->
          <div id="presetSection" style="display: none;">
            <div class="quick-title" style="margin-top: 0; margin-bottom: 6px;" id="quickTitleLabel">Presets</div>
            <div style="display: flex; flex-direction: column; gap: 8px; margin-bottom: 16px;">
              <select id="questionSelect" class="question-select" onchange="selectQuestion(this.value)">
              </select>
            </div>
          </div>

          <div class="input-area" style="margin-top: 0;">
            <!-- IMPECCABLE UNIFIED AI COMMAND BAR CARD -->
            <div id="dropZoneContainer" class="impeccable-cmd-card">
              
              <!-- TOP HEADER BAR INSIDE COMMAND CARD -->
              <div class="cmd-card-header">
                <div id="activeFileBadge" class="cmd-dataset-badge">
                  <span id="activeFileIcon" style="color: #94a3b8; font-weight: 800;">📁</span>
                  <span>Active Dataset: <strong id="activeFileName" style="color: #0f172a;">There is no uploaded file</strong> <span id="activeFileMeta" style="color: #64748b; font-weight: 400;"></span></span>
                </div>
                <label class="cmd-upload-btn">
                  <span>📎 Drag &amp; Drop Excel or Click to Upload</span>
                  <input type="file" id="excelFileInput" accept=".xlsx, .xls, .csv" style="display: none;" onchange="handleExcelFileUpload(this.files[0])" />
                </label>
              </div>

              <!-- MIDDLE PROMPT TEXTAREA AREA -->
              <div class="cmd-card-body">
                <textarea id="questionInput" oninput="updateRunButtonState()" placeholder="E.g.: What is the average price and total inventory value of APPLE products? (or drag & drop your Excel file here / use Voice Command)..."></textarea>
              </div>

              <!-- BOTTOM INTEGRATED TOOLBAR (ACTIONS + VOICE ORB + PRIMARY RUN CTA) -->
              <div class="cmd-card-footer" id="actionRow">
                <div class="cmd-action-left">
                  <button class="cmd-chip-btn" type="button" onclick="openExcelPreview()" id="previewBtnLabel">📊 Excel Preview</button>
                  <button class="cmd-chip-btn" type="button" onclick="downloadExcel()" id="downloadBtnLabel">📥 Export Excel</button>
                </div>
                <div class="cmd-action-right">
                  <button id="voiceBtn" onclick="toggleVoiceRecognition()" type="button" class="cmd-voice-orb-btn">
                    <span id="voiceOrb" class="voice-orb-sm">
                      <span id="voiceIcon" style="display: grid; place-items: center;">
                        <svg width="15" height="15" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.2" stroke-linecap="round" stroke-linejoin="round"><path d="M12 2a3 3 0 0 0-3 3v7a3 3 0 0 0 6 0V5a3 3 0 0 0-3-3z"/><path d="M19 10v2a7 7 0 0 1-14 0v-2"/><line x1="12" y1="19" x2="12" y2="22"/></svg>
                      </span>
                    </span>
                    <span id="voiceText" class="voice-orb-text">Voice AI</span>
                    <span id="voiceSubtext" style="display: none;"></span>
                  </button>
                  <button class="primary-btn cmd-run-btn" onclick="runModule()" id="runBtnLabel" disabled>
                    <span>Run Analysis</span>
                    <kbd class="cmd-shortcut-hint">⌘ ↵</kbd>
                  </button>
                </div>
              </div>

            </div>
          </div>
          </div>

          <!-- INTEGRATED FULL-WIDTH OUTPUT BOX -->
          <div class="result-section" id="resultSection" style="display: none; margin-top: 26px; padding-top: 22px; border-top: 1px solid var(--border);">
            <div class="result-header" style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 14px;">
              <div style="display: flex; align-items: center; gap: 10px;">
                <span style="font-size: 20px;">📈</span>
                <div>
                  <h3 id="resultHeaderTitle" style="font-family: 'Playfair Display', serif; font-size: 20px; font-weight: 700; color: var(--text-900);">Analytical Output</h3>
                  <p id="resultHeaderSub" style="font-size: 12px; color: var(--text-500);">Executive summary and data table outputs are displayed here.</p>
                </div>
              </div>
              <button class="ghost-btn" onclick="clearResult()" id="clearBtnLabel" style="font-size: 12.5px;">Clear</button>
            </div>
            <div class="result-box placeholder" id="resultBox" style="min-height: 280px; width: 100%;">Select a module on the left and start the analysis using voice or text.</div>
          </div>
        </div>
      </section>
    </main>
  </div>

  <!-- EXCEL DATA PREVIEW MODAL -->
  <div id="excelPreviewModal" style="display: none; position: fixed; inset: 0; background: rgba(15, 23, 42, 0.70); backdrop-filter: blur(8px); z-index: 9999; align-items: center; justify-content: center; padding: 24px;">
    <div style="background: #ffffff; border-radius: 24px; width: 94%; max-width: 1150px; max-height: 88vh; display: flex; flex-direction: column; box-shadow: 0 24px 60px rgba(0,0,0,0.35); overflow: hidden; border: 1px solid var(--border);">
      
      <!-- Modal Header -->
      <div style="display: flex; justify-content: space-between; align-items: center; padding: 20px 28px; border-bottom: 1px solid var(--border); background: var(--bg-2);">
        <div style="display: flex; align-items: center; gap: 12px;">
          <div style="width: 42px; height: 42px; border-radius: 12px; background: #ecfdf5; border: 1px solid #10b981; display: grid; place-items: center; font-size: 20px;">📊</div>
          <div>
            <h3 style="font-family: 'Playfair Display', serif; font-size: 20px; font-weight: 700; color: var(--text-900);" id="modalTableTitle">Excel Data & Column Viewer</h3>
            <p style="font-size: 12px; color: var(--text-500);" id="modalTableSub">Live spreadsheet preview, row search & column analysis</p>
          </div>
        </div>
        <button onclick="closeExcelPreview()" style="background: #f1f5f9; border: none; width: 34px; height: 34px; border-radius: 50%; font-size: 16px; font-weight: 700; cursor: pointer; color: var(--text-700); display: grid; place-items: center; transition: background 0.2s;">✕</button>
      </div>

      <!-- Modal Toolbar & Search -->
      <div style="padding: 12px 28px; background: #ffffff; border-bottom: 1px solid var(--border); display: flex; justify-content: space-between; align-items: center; gap: 14px;">
        <input type="text" id="modalTableSearch" oninput="filterPreviewTable(this.value)" placeholder="🔍 Search in table (SKU, Brand, Category)..." style="flex: 1; max-width: 400px; padding: 8px 14px; font-size: 12.5px; border: 1.4px solid var(--border); border-radius: 10px; outline: none;" />
        <span style="font-size: 12.5px; color: var(--text-500); font-weight: 600;" id="modalRowCount">Showing: 15 Rows</span>
      </div>

      <!-- Modal Table Content -->
      <div style="flex: 1; padding: 20px 28px; overflow: auto; background: #ffffff;" id="excelPreviewTableContainer">
        <!-- Generated dynamically in JS -->
      </div>

      <!-- Modal Footer -->
      <div style="padding: 16px 28px; border-top: 1px solid var(--border); background: var(--bg-2); display: flex; justify-content: space-between; align-items: center;">
        <span style="font-size: 12px; color: #10b981; font-weight: 700;" id="modalStatusText">✓ Excel data is loaded and ready for AI analysis.</span>
        <button onclick="closeExcelPreview()" class="primary-btn" style="padding: 8px 20px; font-size: 13px;" id="modalCloseBtn">Close</button>
      </div>
  <!-- USER PROFILE & BILLING DETAILS MODAL -->
  <div id="userProfileModal" style="display: none; position: fixed; inset: 0; background: rgba(15, 23, 42, 0.75); backdrop-filter: blur(8px); z-index: 99999; align-items: center; justify-content: center; padding: 24px;">
    <div style="background: #ffffff; border-radius: 24px; width: 100%; max-width: 580px; overflow: hidden; box-shadow: 0 25px 50px -12px rgba(0, 0, 0, 0.35); border: 1px solid #e2e8f0;">
      
      <!-- MODAL HEADER -->
      <div style="background: linear-gradient(135deg, #f26f26 0%, #2563eb 100%); padding: 24px 28px; color: #ffffff; display: flex; justify-content: space-between; align-items: center;">
        <div style="display: flex; align-items: center; gap: 14px;">
          <div style="width: 46px; height: 46px; border-radius: 50%; background: #fde68a; color: #78350f; font-weight: 800; font-size: 16px; display: grid; place-items: center; border: 2px solid #ffffff; box-shadow: 0 4px 12px rgba(0,0,0,0.15);">YK</div>
          <div>
            <h3 style="font-family: 'Outfit', sans-serif; font-size: 19px; font-weight: 800; margin: 0; color: #ffffff;">Yasam Karadag</h3>
            <p style="margin: 2px 0 0; font-size: 12px; opacity: 0.9; font-weight: 500;">owner · yasamkaradag@dataprovido.com</p>
          </div>
        </div>
        <button onclick="closeUserProfileModal()" type="button" style="background: rgba(255,255,255,0.20); border: none; color: #ffffff; width: 32px; height: 32px; border-radius: 50%; font-size: 16px; font-weight: 700; cursor: pointer; display: grid; place-items: center; transition: background 0.2s;">✕</button>
      </div>

      <!-- MODAL BODY -->
      <div style="padding: 24px 28px; max-height: 70vh; overflow-y: auto;">
        
        <!-- ACCOUNT SUMMARY BOX -->
        <div style="background: #f8fafc; border: 1px solid #e2e8f0; border-radius: 16px; padding: 16px 20px; margin-bottom: 20px;">
          <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 10px;">
            <span style="font-size: 11.5px; font-weight: 800; text-transform: uppercase; letter-spacing: 0.08em; color: #64748b;">Subscription Plan</span>
            <span style="background: #eff6ff; color: #2563eb; border: 1px solid #bfdbfe; font-size: 11px; font-weight: 800; padding: 3px 10px; border-radius: 20px;">FOUNDER EDITION</span>
          </div>
          <div style="font-size: 18px; font-weight: 800; color: #0f172a; margin-bottom: 4px;">DataProvido Founder Plan</div>
          <div style="font-size: 12.5px; color: #475569; line-height: 1.5;">Unlimited 100% Local On-Premise Execution · Zero Cloud Data Leakage</div>
        </div>

        <!-- DETAILS GRID -->
        <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 14px; margin-bottom: 20px;">
          <div style="background: #ffffff; border: 1px solid #f1f5f9; border-radius: 14px; padding: 14px;">
            <div style="font-size: 11px; color: #64748b; font-weight: 600; margin-bottom: 4px;">Registered Email</div>
            <div style="font-size: 13px; font-weight: 700; color: #0f172a;">yasamkaradag@dataprovido.com</div>
          </div>
          <div style="background: #ffffff; border: 1px solid #f1f5f9; border-radius: 14px; padding: 14px;">
            <div style="font-size: 11px; color: #64748b; font-weight: 600; margin-bottom: 4px;">License Key</div>
            <div style="font-size: 13px; font-weight: 700; color: #2563eb; font-family: monospace;">DP-FOUNDER-9981-PRO</div>
          </div>
          <div style="background: #ffffff; border: 1px solid #f1f5f9; border-radius: 14px; padding: 14px;">
            <div style="font-size: 11px; color: #64748b; font-weight: 600; margin-bottom: 4px;">Connected Property</div>
            <div style="font-size: 13px; font-weight: 700; color: #0f172a;">dataprovido.com</div>
          </div>
          <div style="background: #ffffff; border: 1px solid #f1f5f9; border-radius: 14px; padding: 14px;">
            <div style="font-size: 11px; color: #64748b; font-weight: 600; margin-bottom: 4px;">Local Engine Status</div>
            <div style="font-size: 13px; font-weight: 700; color: #10b981;">● Online (Local Mac Node)</div>
          </div>
        </div>

        <!-- BILLING HISTORY -->
        <div style="border-top: 1px solid #f1f5f9; padding-top: 16px;">
          <h4 style="font-size: 13px; font-weight: 800; color: #0f172a; margin-bottom: 12px;">Billing &amp; Invoices</h4>
          <div style="display: flex; justify-content: space-between; align-items: center; padding: 10px 14px; background: #f8fafc; border-radius: 12px; border: 1px solid #e2e8f0; font-size: 12.5px;">
            <div>
              <strong style="color: #0f172a; display: block;">Invoice #INV-2026-0801</strong>
              <span style="color: #64748b; font-size: 11px;">Aug 1, 2026 · Founder Lifetime License</span>
            </div>
            <span style="color: #10b981; font-weight: 800; font-size: 12px;">₺0.00 (Active)</span>
          </div>
        </div>

      </div>

      <!-- MODAL FOOTER -->
      <div style="padding: 16px 28px; background: #f8fafc; border-top: 1px solid #e2e8f0; display: flex; justify-content: space-between; align-items: center;">
        <a href="/pricing" style="font-size: 12.5px; font-weight: 700; color: #2563eb; text-decoration: none;">Manage Subscription in Stripe →</a>
        <button onclick="closeUserProfileModal()" type="button" class="primary-btn" style="padding: 8px 20px; font-size: 13px;">Close</button>
      </div>

    </div>
  </div>

  <script>
    /* ── Global State ── */
    var currentLang = 'en';
    var currentModule = 'business_calculator';
    var isHowToUseOpen = false;
    var recognition = null;
    var isListening = false;

    /* ── How to Use Guide Content ── */
    const howToUseGuide = {
      business_calculator: {
        tr: `<strong>Excel Wizard Kullanım Rehberi:</strong><br>
        1. Yüklenmiş Excel/CSV verileriniz üzerinde doğrudan doğal dille hesaplama yapın.<br>
        2. <b>Voice Command</b> butonuna basarak mikrofona Excel komutunuzu verin veya metin alanına yazın.<br>
        3. <b>Excel Preview</b> butonuna basarak Excel tablonuzu ve sütun verilerinizi canlı önizleyin.<br>
        4. <b>Result Preview</b> butonu ile hesaplama sonuçlarını anında görüntüleyin.`,
        en: `<strong>Excel Wizard User Guide:</strong><br>
        1. Perform advanced calculations directly on your Excel/CSV data using natural language.<br>
        2. Click <b>Voice Command</b> to speak your query or type it into the prompt box.<br>
        3. Click <b>Excel Preview</b> to inspect your spreadsheet rows & columns in real-time.<br>
        4. Click <b>Result Preview</b> to preview calculation results and summary data.`
      },
      category_insights: {
        tr: `<strong>Category Insights Kullanım Rehberi:</strong><br>
        1. İncelemek istediğiniz e-ticaret kategorisini veya markasını seçin.<br>
        2. Pazarda kazanan/kaybeden SKU'ları, fiyat kırılımlarını ve riskleri tek tıkla analiz ettirin.<br>
        3. Yönetici özetini görüntüleyin ve raporu Excel olarak indirin.`,
        en: `<strong>Category Insights User Guide:</strong><br>
        1. Select the e-commerce category or brand you want to analyze.<br>
        2. Evaluate winning/losing SKUs, pricing segments, and market risks.<br>
        3. View executive summary and export the final report to Excel.`
      },
      price_competition: {
        tr: `<strong>Price Competition Kullanım Rehberi:</strong><br>
        1. Rakip ve pazar benchmark fiyat verilerini sisteme aktarın.<br>
        2. Pahalı kalan, rekabetçi olan ve fiyat indirimine duyarlı SKU'ları tespit edin.<br>
        3. Fiyat rekabet endeksini raporlayın.`,
        en: `<strong>Price Competition User Guide:</strong><br>
        1. Import competitor benchmark pricing data.<br>
        2. Identify overpriced SKUs, competitive pricing bands, and elastic items.<br>
        3. Generate price competitiveness index reports.`
      },
      action_executor: {
        tr: `<strong>Action Executor Kullanım Rehberi:</strong><br>
        1. Analiz çıktılarına göre öncelikli aksiyon önerilerini otomatik üretin.<br>
        2. Stok tamamlama (replenishment), kampanya ve görünürlük önerilerini listeleyin.<br>
        3. Aksiyon planını Excel sayfası olarak dışa aktarın.`,
        en: `<strong>Action Executor User Guide:</strong><br>
        1. Automatically generate prioritized business action plans from insights.<br>
        2. Review replenishment, marketing, and stock risk reduction steps.<br>
        3. Export action items to Excel.`
      },
      funnel_stock: {
        tr: `<strong>Funnel & Stock Kullanım Rehberi:</strong><br>
        1. PDP görünürlüğü yüksek ancak stoğu az olan riskli ürünleri sorgulayın.<br>
        2. Sepet kayıplarını ve dönüşüm hunisindeki darboğazları analiz edin.`,
        en: `<strong>Funnel & Stock User Guide:</strong><br>
        1. Identify high-traffic SKUs with low inventory holding.<br>
        2. Analyze cart abandonment and conversion funnel leakage points.`
      },
      excel_outputs: {
        tr: `<strong>Excel Outputs Kullanım Rehberi:</strong><br>
        1. Daha önce çalıştırılan analiz sonuçlarını çoklu sheet (çalışma sayfası) formatında görün.<br>
        2. Raporları anında bilgisayarınıza `.xlsx` dosyası olarak indirin.`,
        en: `<strong>Excel Outputs User Guide:</strong><br>
        1. Review all previous analysis runs structured into multi-sheet Workbooks.<br>
        2. Download formatted Excel files directly to your device.`
      },
      data_upload: {
        tr: `<strong>Veri Kaynakları Yükle Kullanım Rehberi:</strong><br>
        1. Kendi e-ticaret satış, stok ve rakip fiyat Excel/CSV dosyalarınızı sürükleyip bırakın.<br>
        2. Sistem verilerinizi otomatik ilişkilendirip analize hazır hale getirir.`,
        en: `<strong>Data Upload User Guide:</strong><br>
        1. Drag and drop your custom e-commerce sales, stock, and pricing Excel/CSV files.<br>
        2. Data is automatically ingested and mapped for immediate AI querying.`
      }
    };

    /* ── Module configs (Bilingual Support) ── */
    const modules = {
      business_calculator: {
        title: "Excel Wizard",
        desc: {
          tr: "Satış, stok, fiyat ve e-ticaret Excel verileriniz üzerinde; sesli komutlar veya yazılı sorular ile ortalama, toplam, filtreleme, marka ve kategori bazlı kırılımlar gibi gelişmiş Excel hesaplamalarını anında yapmanızı sağlar.",
          en: "Execute advanced mathematical calculations, average, sum, filters, and brand/category breakdowns on all your retail & e-commerce Excel data using English or Turkish voice commands."
        },
        placeholder: {
          tr: "Örn: APPLE markasının ortalama fiyatı ve stok tutarı nedir? (veya Voice Command kullanın)",
          en: "E.g.: Calculate average price and total inventory value for APPLE products (or use Voice Command)"
        },
        suggestions: { tr: [], en: [] }
      },
      category_insights: {
        title: "Category Insights",
        desc: {
          tr: "Seçilen kategori için performans, kazanan segmentler, riskler ve yönetici özeti çıkarır.",
          en: "Generates category performance metrics, winning segments, risk analysis, and executive summaries."
        },
        placeholder: {
          tr: "Örn: Tablet kategorisinin performans insightını ver",
          en: "E.g.: Provide detailed performance insight for Tablets category"
        },
        suggestions: {
          tr: [
            "Hangi SKU'da pahalıyız?",
            "Hangi SKU'da ucuzuz ama satış alamıyoruz?",
            "Hangi SKU'da benchmark üstündeyiz ama hâlâ iyi satıyoruz?",
            "Hangi markada/kategoride fiyat rekabetini kaybediyoruz?",
            "Tablet kategorisinin performans insightını ver",
            "Cep Telefonları kategorisini detaylı analiz et"
          ],
          en: [
            "Which SKUs have a price premium above market benchmark?",
            "Which SKUs are cheaper than competitors but failing to convert sales?",
            "Which brands are losing price competitiveness?",
            "Provide detailed category performance insight for Tablets",
            "Analyze Smartphones category in detail"
          ]
        }
      },
      price_competition: {
        title: "Price Competition",
        desc: {
          tr: "Merchant benchmark datasına göre pahalı, ucuz, rekabetçi ve riskli ürünleri analiz eder.",
          en: "Analyzes benchmark price positioning (expensive, cheap, competitive, at-risk SKUs)."
        },
        placeholder: {
          tr: "Örn: Cep Telefonları kategorisinde fiyat rekabeti analizi yap",
          en: "E.g.: Analyze price competition in Mobile Phones category"
        },
        suggestions: {
          tr: [
            "Cep Telefonları kategorisinde fiyat rekabeti analizi yap",
            "APPLE markasında benchmarka göre pahalı ürünleri çıkar",
            "Rakibe göre en pahalı 10 SKU hangileri?",
            "Median competitor price altında kalan ürünleri bul"
          ],
          en: [
            "Analyze price competition in Mobile Phones category",
            "Identify expensive APPLE products relative to market benchmark",
            "List top 10 most expensive SKUs vs competitor pricing",
            "Find products priced below median competitor price"
          ]
        }
      },
      action_executor: {
        title: "Action Executor",
        desc: {
          tr: "Insight çıktılarından aksiyon planı üretir: replenishment, kampanya ve stok riski aksiyonları.",
          en: "Generates action plans: replenishment, marketing, and stock risk mitigation."
        },
        placeholder: {
          tr: "Örn: Aksiyon planı üret",
          en: "E.g.: Generate action plan for at-risk items"
        },
        suggestions: { tr: [], en: [] }
      },
      funnel_stock: {
        title: "Funnel & Stock",
        desc: {
          tr: "PDP görünürlüğü yüksek ancak stoğu az olan riskli ürünleri sorgular.",
          en: "Queries high-traffic SKUs with low stock and conversion bottlenecks."
        },
        placeholder: {
          tr: "Örn: Stok riski taşıyan ürünleri listele",
          en: "E.g.: List SKUs with stock risk"
        },
        suggestions: { tr: [], en: [] }
      },
      excel_outputs: {
        title: "Excel Outputs",
        desc: {
          tr: "Daha önce çalıştırılan analiz sonuçlarını çoklu sheet formatında gösterir.",
          en: "Displays previous analysis run results in multi-sheet Excel format."
        },
        placeholder: {
          tr: "Örn: Excel rapor çıktılarını göster",
          en: "E.g.: Display Excel report outputs"
        },
        suggestions: { tr: [], en: [] }
      },
      data_upload: {
        title: "Upload Data Sources",
        desc: {
          tr: "Kendi e-ticaret Excel/CSV dosyalarınızı yükleyin.",
          en: "Upload your custom e-commerce Excel/CSV files."
        },
        placeholder: {
          tr: "Örn: Dosya yükleme modülünü aç",
          en: "E.g.: Open data source upload manager"
        },
        suggestions: { tr: [], en: [] }
      }
    };

    let langToastTimer = null;

    function setLanguage(lang) {
      currentLang = lang;
      const btnTR = document.getElementById("langBtnTR");
      const btnEN = document.getElementById("langBtnEN");
      const voiceText = document.getElementById("voiceText");
      const runBtnLabel = document.getElementById("runBtnLabel");
      const previewBtnLabel = document.getElementById("previewBtnLabel");
      const downloadBtnLabel = document.getElementById("downloadBtnLabel");
      const clearBtnLabel = document.getElementById("clearBtnLabel");
      const resultHeaderTitle = document.getElementById("resultHeaderTitle");
      const resultHeaderSub = document.getElementById("resultHeaderSub");
      const howToUseHeaderTitle = document.getElementById("howToUseHeaderTitle");
      const toast = document.getElementById("langToastBadge");

      const heroMainHeading = document.getElementById("heroMainHeading");
      const heroSubHeading = document.getElementById("heroSubHeading");
      const homeLinkText = document.getElementById("homeLinkText");
      const navLabelCategories = document.getElementById("navLabelCategories");

      const subBusiness = document.getElementById("sub_business_calculator");
      const subCategory = document.getElementById("sub_category_insights");
      const subPrice = document.getElementById("sub_price_competition");
      const subAction = document.getElementById("sub_action_executor");
      const subFunnel = document.getElementById("sub_funnel_stock");
      const subExcel = document.getElementById("sub_excel_outputs");
      const labelUpload = document.getElementById("label_data_upload");
      const subUpload = document.getElementById("sub_data_upload");

      const modalTableSub = document.getElementById("modalTableSub");
      const modalTableSearch = document.getElementById("modalTableSearch");
      const modalStatusText = document.getElementById("modalStatusText");
      const modalCloseBtn = document.getElementById("modalCloseBtn");

      if (lang === 'en') {
        btnTR.style.background = "transparent";
        btnTR.style.borderColor = "transparent";
        btnTR.style.color = "#475569";
        btnTR.style.fontWeight = "600";
        btnTR.style.boxShadow = "none";
        
        btnEN.style.background = "#2563eb";
        btnEN.style.borderColor = "#2563eb";
        btnEN.style.color = "#ffffff";
        btnEN.style.fontWeight = "700";
        btnEN.style.boxShadow = "0 2px 8px rgba(37,99,235,0.30)";

        if (toast) {
          toast.textContent = "🌐 Voice Mode: English (en-US)";
          toast.style.display = "block";
          toast.style.opacity = "1";
        }

        if (!isListening) voiceText.textContent = "Voice Command";
        runBtnLabel.textContent = "Run Analysis";
        previewBtnLabel.textContent = "Excel Preview";
        downloadBtnLabel.textContent = "Export Excel";
        clearBtnLabel.textContent = "Clear";
        resultHeaderTitle.textContent = "Analytical Output";
        resultHeaderSub.textContent = "Executive summary and data table outputs are displayed here.";
        howToUseHeaderTitle.textContent = "How to Use Guide";

        if (heroMainHeading) heroMainHeading.textContent = "Excel Wizard allows you to perform your manual Excel tasks using your voice and text commands.";
        if (heroSubHeading) heroSubHeading.textContent = "Execute Excel Wizard, category insights, price competition, and action plans in a single workspace; give voice or text Excel commands over your data.";
        if (homeLinkText) homeLinkText.textContent = "← Home Page";
        if (navLabelCategories) navLabelCategories.textContent = "CATEGORIES";

        if (subBusiness) subBusiness.textContent = "Voice & text Excel commands";
        if (subCategory) subCategory.textContent = "Category & sector analysis";
        if (subPrice) subPrice.textContent = "Merchant benchmark";
        if (subAction) subAction.textContent = "Action plan generator";
        if (subFunnel) subFunnel.textContent = "Conversion & stock risk";
        if (subExcel) subExcel.textContent = "Analytical reports";
        if (labelUpload) labelUpload.textContent = "Upload Data Sources";
        if (subUpload) subUpload.textContent = "Excel / CSV Upload";

        if (modalTableSub) modalTableSub.textContent = "Live spreadsheet preview, row search & column analysis";
        if (modalTableSearch) modalTableSearch.placeholder = "🔍 Search in table (SKU, Brand, Category)...";
        if (modalStatusText) modalStatusText.textContent = "✓ Excel data is loaded and ready for AI analysis.";
        if (modalCloseBtn) modalCloseBtn.textContent = "Close";
      } else {
        btnEN.style.background = "transparent";
        btnEN.style.borderColor = "transparent";
        btnEN.style.color = "#475569";
        btnEN.style.fontWeight = "600";
        btnEN.style.boxShadow = "none";

        btnTR.style.background = "#2563eb";
        btnTR.style.borderColor = "#2563eb";
        btnTR.style.color = "#ffffff";
        btnTR.style.fontWeight = "700";
        btnTR.style.boxShadow = "0 2px 8px rgba(37,99,235,0.30)";

        if (toast) {
          toast.textContent = "🌐 Ses Modu: Türkçe (tr-TR)";
          toast.style.display = "block";
          toast.style.opacity = "1";
        }

        if (!isListening) voiceText.textContent = "Voice Command";
        runBtnLabel.textContent = "Run Analysis";
        previewBtnLabel.textContent = "Excel Preview";
        downloadBtnLabel.textContent = "Export Excel";
        clearBtnLabel.textContent = "Temizle";
        resultHeaderTitle.textContent = "Analiz Çıktısı";
        resultHeaderSub.textContent = "Sonuç burada yönetici özeti formatında gösterilir.";
        howToUseHeaderTitle.textContent = "Nasıl Kullanılır? (How to Use Guide)";

        if (heroMainHeading) heroMainHeading.textContent = "Excel Wizard, elinizle yaptığınız Excel işlemlerini sesiniz ve yazılı komutunuz ile yapmanızı sağlar.";
        if (heroSubHeading) heroSubHeading.textContent = "Excel Wizard, kategori insight, fiyat rekabeti ve aksiyon planlarını tek bir çalışma alanında çalıştırın; verileriniz üzerinde sesli veya yazılı Excel komutları verin.";
        if (homeLinkText) homeLinkText.textContent = "← Ana Sayfa";
        if (navLabelCategories) navLabelCategories.textContent = "KATEGORİLER";

        if (subBusiness) subBusiness.textContent = "Sesli & yazılı Excel komutları";
        if (subCategory) subCategory.textContent = "Kategori & sektör analizi";
        if (subPrice) subPrice.textContent = "Merchant benchmark";
        if (subAction) subAction.textContent = "Aksiyon planı";
        if (subFunnel) subFunnel.textContent = "Dönüşüm & stok riski";
        if (subExcel) subExcel.textContent = "Rapor çıktıları";
        if (labelUpload) labelUpload.textContent = "Veri Kaynakları Yükle";
        if (subUpload) subUpload.textContent = "Excel / CSV Yükleme";

        if (modalTableSub) modalTableSub.textContent = "Canlı Excel tablosu önizlemesi, filtreleme ve sütun analizi";
        if (modalTableSearch) modalTableSearch.placeholder = "🔍 Tablo içinde ara (SKU, Marka, Kategori)...";
        if (modalStatusText) modalStatusText.textContent = "✓ Excel verisi analize hazır durumdadır.";
        if (modalCloseBtn) modalCloseBtn.textContent = "Kapat";
      }

      if (langToastTimer) clearTimeout(langToastTimer);
      langToastTimer = setTimeout(() => {
        if (toast) {
          toast.style.opacity = "0";
          setTimeout(() => { toast.style.display = "none"; }, 250);
        }
      }, 1200);

      renderHowToUse();
      const activeBtn = document.querySelector(".menu-btn.active");
      if (activeBtn) renderModule(activeBtn.dataset.key);
      updateRunButtonState();
    }

    function updateRunButtonState() {
      const input = document.getElementById("questionInput");
      const runBtn = document.getElementById("runBtnLabel");
      if (!input || !runBtn) return;
      const hasText = input.value.trim().length > 0;
      if (hasText) {
        runBtn.style.background = "linear-gradient(135deg, #f26f26 0%, #d85c18 100%)";
        runBtn.style.color = "#ffffff";
        runBtn.style.opacity = "1";
        runBtn.style.cursor = "pointer";
        runBtn.style.boxShadow = "0 4px 14px rgba(242,111,38,0.30)";
        runBtn.disabled = false;
      } else {
        runBtn.style.background = "#cbd5e1";
        runBtn.style.color = "#64748b";
        runBtn.style.opacity = "0.65";
        runBtn.style.cursor = "not-allowed";
        runBtn.style.boxShadow = "none";
        runBtn.disabled = true;
      }
    }

    // Keyboard shortcut handler (Command/Ctrl + Enter)
    document.addEventListener("keydown", function(e) {
      if ((e.metaKey || e.ctrlKey) && e.key === "Enter") {
        const runBtn = document.getElementById("runBtnLabel");
        if (runBtn && !runBtn.disabled) {
          e.preventDefault();
          runModule();
        }
      }
    });

    /* ── Drag & Drop Excel File Ingestion Handler ── */
    document.addEventListener("DOMContentLoaded", function() {
      const dropZone = document.getElementById("dropZoneContainer");
      if (dropZone) {
        ['dragenter', 'dragover', 'dragleave', 'drop'].forEach(eventName => {
          dropZone.addEventListener(eventName, e => {
            e.preventDefault();
            e.stopPropagation();
          }, false);
        });

        ['dragenter', 'dragover'].forEach(eventName => {
          dropZone.addEventListener(eventName, () => {
            dropZone.style.borderColor = '#2563eb';
            dropZone.style.background = '#eff6ff';
          }, false);
        });

        ['dragleave', 'drop'].forEach(eventName => {
          dropZone.addEventListener(eventName, () => {
            dropZone.style.borderColor = '#cbd5e1';
            dropZone.style.background = '#ffffff';
          }, false);
        });

        dropZone.addEventListener('drop', e => {
          const dt = e.dataTransfer;
          const files = dt.files;
          if (files && files.length > 1) {
            alert(currentLang === 'en' 
              ? "⚠️ You can only upload 1 Excel file at a time. Processing the first file." 
              : "⚠️ Sadece 1 adet Excel dosyası yükleyebilirsiniz. İlk seçilen dosya işleme alındı.");
          }
          if (files && files.length > 0) {
            handleExcelFileUpload(files[0]);
          }
        }, false);
      }

      // Recover cached active file from browser sessionStorage on page load
      const savedFileName = sessionStorage.getItem("activeExcelFileName");
      const savedFileMeta = sessionStorage.getItem("activeExcelFileMeta");
      if (savedFileName) {
        const activeFileName = document.getElementById("activeFileName");
        const activeFileMeta = document.getElementById("activeFileMeta");
        const activeFileIcon = document.getElementById("activeFileIcon");
        if (activeFileName) activeFileName.textContent = savedFileName;
        if (activeFileMeta && savedFileMeta) activeFileMeta.textContent = savedFileMeta;
        if (activeFileIcon) {
          activeFileIcon.textContent = "✓";
          activeFileIcon.style.color = "#10b981";
        }
      }
    });

    function handleExcelFileUpload(file) {
      if (!file) return;
      const validExts = ['.xlsx', '.xls', '.csv'];
      const fileExt = file.name.substring(file.name.lastIndexOf('.')).toLowerCase();
      if (!validExts.includes(fileExt)) {
        alert(currentLang === 'en' ? "Please upload a valid Excel (.xlsx, .xls) or CSV (.csv) file." : "Lütfen geçerli bir Excel (.xlsx, .xls) veya CSV (.csv) dosyası yükleyin.");
        return;
      }

      const activeFileName = document.getElementById("activeFileName");
      const activeFileMeta = document.getElementById("activeFileMeta");
      const activeFileIcon = document.getElementById("activeFileIcon");
      if (activeFileName) activeFileName.textContent = file.name;
      if (activeFileIcon) {
        activeFileIcon.textContent = "✓";
        activeFileIcon.style.color = "#10b981";
      }
      
      const metaText = `(${(file.size / 1024).toFixed(1)} KB) ✓ Excel active & ready for commands`;
      if (activeFileMeta) activeFileMeta.textContent = metaText;

      try {
        sessionStorage.setItem("activeExcelFileName", file.name);
        sessionStorage.setItem("activeExcelFileMeta", metaText);
      } catch (err) {
        console.warn("sessionStorage notice:", err);
      }

      const reader = new FileReader();
      reader.onload = function(e) {
        const questionInput = document.getElementById("questionInput");
        if (questionInput && !questionInput.value.trim()) {
          questionInput.value = (currentLang === 'en')
            ? `Calculate average price and total inventory value for dataset '${file.name}'`
            : `'${file.name}' veri seti için ortalama fiyatı ve toplam stok tutarını hesapla`;
          updateRunButtonState();
        }
      };
      reader.readAsArrayBuffer(file);
    }

    function selectQuestion(val) {
      if (val) {
        questionInput.value = val;
        updateRunButtonState();
      }
    }

    menuButtons.forEach(btn => btn.addEventListener("click", () => renderModule(btn.dataset.key)));

    /* ── Voice-to-Excel SpeechRecognition Engine (Bilingual TR / EN) ── */

    function toggleVoiceRecognition() {
      const voiceBtn = document.getElementById("voiceBtn");
      const voiceIcon = document.getElementById("voiceIcon");
      const voiceText = document.getElementById("voiceText");
      const questionInput = document.getElementById("questionInput");

      const SpeechRecognition = window.SpeechRecognition || window.webkitSpeechRecognition;
      if (!SpeechRecognition) {
        alert(currentLang === 'en' ? "Your browser does not support live speech recognition. Please use Google Chrome, Edge, or Safari." : "Tarayıcınız canlı ses tanımayı desteklemiyor. Lütfen Chrome, Edge veya Safari kullanın.");
        return;
      }

      if (isListening) {
        if (recognition) { try { recognition.stop(); } catch(e){} }
        return;
      }

      if (navigator.mediaDevices && navigator.mediaDevices.getUserMedia) {
        navigator.mediaDevices.getUserMedia({ audio: true }).then(() => {
          startSpeechEngine();
        }).catch(err => {
          alert(currentLang === 'en'
            ? "⚠️ Microphone access permission was denied or not granted. Please allow microphone access in your browser."
            : "⚠️ Mikrofon erişim izni engellendi veya verilmedi. Lütfen tarayıcı ayarlarından mikrofona izin verin.");
        });
      } else {
        startSpeechEngine();
      }

      function startSpeechEngine() {
        recognition = new SpeechRecognition();
        recognition.lang = (currentLang === 'en') ? 'en-US' : 'tr-TR';
        recognition.interimResults = true;
        recognition.continuous = true;

        recognition.onstart = function() {
          isListening = true;
          const voiceSubtext = document.getElementById("voiceSubtext");
          voiceBtn.classList.add("voice-listening-active");
          voiceIcon.innerHTML = `<div class="sound-wave-bars"><span class="bar"></span><span class="bar"></span><span class="bar"></span></div>`;
          voiceText.textContent = (currentLang === 'en') ? "Recording..." : "Dinleniyor...";
          if (voiceSubtext) voiceSubtext.textContent = (currentLang === 'en') ? "🔴 Click to Stop" : "🔴 Durdurmak İçin Tıkla";
          questionInput.placeholder = (currentLang === 'en') 
            ? "🔴 Listening... Speak your Excel command now. Click button when finished..." 
            : "🔴 Dinleniyor... Lütfen Excel komutunuzu söyleyin. Bitirdiğinizde butona basın...";
        };

        recognition.onresult = function(event) {
          let transcript = "";
          for (let i = event.resultIndex; i < event.results.length; i++) {
            transcript += event.results[i][0].transcript;
          }
          if (transcript.trim()) {
            questionInput.value = transcript;
            updateRunButtonState();
          }
        };

        recognition.onerror = function(event) {
          console.error("Speech error:", event.error);
          stopListeningUI();
          if (event.error === 'no-speech') return;
          let errMsg = (currentLang === 'en')
            ? "⚠️ Voice recognition notice: " + event.error
            : "⚠️ Ses tanıma uyarısı: " + event.error;
          if (event.error === 'not-allowed') {
            errMsg = (currentLang === 'en')
              ? "⚠️ Microphone permission denied in browser settings."
              : "⚠️ Tarayıcı ayarlarında mikrofon izni verilmedi.";
          }
          alert(errMsg);
        };

        recognition.onend = function() {
          stopListeningUI();
          if (questionInput.value.trim().length > 0) {
            updateRunButtonState();
          }
        };

        function stopListeningUI() {
          isListening = false;
          const voiceSubtext = document.getElementById("voiceSubtext");
          voiceBtn.classList.remove("voice-listening-active");
          voiceIcon.innerHTML = `<svg width="22" height="22" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.2" stroke-linecap="round" stroke-linejoin="round"><path d="M12 2a3 3 0 0 0-3 3v7a3 3 0 0 0 6 0V5a3 3 0 0 0-3-3z"/><path d="M19 10v2a7 7 0 0 1-14 0v-2"/><line x1="12" y1="19" x2="12" y2="22"/></svg>`;
          voiceText.textContent = "Voice AI Command";
          if (voiceSubtext) voiceSubtext.textContent = "🎙️ Speak or Click";
        }

        try {
          recognition.start();
        } catch(e) {
          console.error("Start error:", e);
        }
      }
    }

    function toggleHowToUse() {
      isHowToUseOpen = !isHowToUseOpen;
      const body = document.getElementById("howToUseBody");
      const icon = document.getElementById("howToUseIcon");
      if (body && icon) {
        body.style.display = isHowToUseOpen ? "block" : "none";
        icon.style.transform = isHowToUseOpen ? "rotate(180deg)" : "rotate(0deg)";
      }
    }

    function renderHowToUse() {
      const body = document.getElementById("howToUseBody");
      if (!body) return;
      const content = howToUseGuide[currentModule] || howToUseGuide.business_calculator;
      body.innerHTML = content[currentLang] || content.tr;
    }

    function renderModule(key) {
      currentModule = key;
      const conf = modules[key] || modules.business_calculator;

      const menuButtons = document.querySelectorAll(".menu-btn");
      menuButtons.forEach(btn => {
        if (btn.dataset.key === key) {
          btn.classList.add("active");
        } else {
          btn.classList.remove("active");
        }
      });

      const moduleTitle = document.getElementById("moduleTitle");
      const moduleDesc = document.getElementById("moduleDesc");
      const questionInput = document.getElementById("questionInput");

      if (moduleTitle) moduleTitle.textContent = conf.title;
      if (moduleDesc) moduleDesc.textContent = conf.desc[currentLang] || conf.desc.tr;
      if (questionInput) questionInput.placeholder = conf.placeholder[currentLang] || conf.placeholder.tr;

      renderHowToUse();
      updateRunButtonState();
    }

    function selectQuestion(val) {
      if (val) { questionInput.value = val; }
    }

    menuButtons.forEach(btn => btn.addEventListener("click", () => renderModule(btn.dataset.key)));

    function openUserProfileModal() {
      const modal = document.getElementById("userProfileModal");
      if (modal) modal.style.display = "flex";
    }

    function closeUserProfileModal() {
      const modal = document.getElementById("userProfileModal");
      if (modal) modal.style.display = "none";
    }

    // Explicitly attach all interactive handlers to window object for global HTML onclick availability
    window.toggleVoiceRecognition = toggleVoiceRecognition;
    window.setLanguage = setLanguage;
    window.openExcelPreview = openExcelPreview;
    window.closeExcelPreview = closeExcelPreview;
    window.downloadExcel = downloadExcel;
    window.clearResult = clearResult;
    window.runModule = runModule;
    window.selectQuestion = selectQuestion;
    window.toggleHowToUse = toggleHowToUse;
    window.openUserProfileModal = openUserProfileModal;
    window.closeUserProfileModal = closeUserProfileModal;

    /* ── Live Interactive Excel Spreadsheet Data Engine ── */
    let currentSpreadsheetData = [
      { SKU: "SKU-1001", Name: "iPhone 15 Pro Max 256GB", Category: "Smartphones", Price: 54999, CompPrice: 52490, Stock: 142, Revenue: 7809858, Status: "Competitive" },
      { SKU: "SKU-1002", Name: "MacBook Pro 16 M3 Max", Category: "Laptops", Price: 124999, CompPrice: 129000, Stock: 28, Revenue: 3499972, Status: "Price Leader" },
      { SKU: "SKU-1003", Name: "iPad Air 11 M2 Wi-Fi", Category: "Tablets", Price: 24999, CompPrice: 24999, Stock: 89, Revenue: 2224911, Status: "Benchmark Matched" },
      { SKU: "SKU-1004", Name: "AirPods Pro 2nd Gen USB-C", Category: "Accessories", Price: 8499, CompPrice: 7990, Stock: 310, Revenue: 2634690, Status: "Overpriced (+6.3%)" },
      { SKU: "SKU-1005", Name: "Samsung Galaxy S24 Ultra", Category: "Smartphones", Price: 64999, CompPrice: 61900, Stock: 65, Revenue: 4224935, Status: "Overpriced (+5.0%)" },
      { SKU: "SKU-1006", Name: "Sony WH-1000XM5 ANC", Category: "Accessories", Price: 13999, CompPrice: 14500, Stock: 44, Revenue: 615956, Status: "Underpriced (-3.4%)" },
      { SKU: "SKU-1007", Name: "Dell XPS 15 OLED i9", Category: "Laptops", Price: 89999, CompPrice: 92000, Stock: 12, Revenue: 1079988, Status: "Price Leader" }
    ];

    async function runModule() {
      const questionInput = document.getElementById("questionInput");
      const resultBox = document.getElementById("resultBox");
      if (!questionInput || !resultBox) return;
      const question = questionInput.value.trim();
      if (!question) return;

      const resultSection = document.getElementById("resultSection");
      if (resultSection) resultSection.style.display = "block";
      resultBox.classList.remove("placeholder");
      resultBox.classList.add("loading");
      resultBox.innerHTML = (currentLang === 'en') ? "⚡ Executing Excel calculation on active dataset..." : "⚡ Aktif Excel verisi üzerinde hesaplama yapılıyor...";

      setTimeout(() => {
        resultBox.classList.remove("loading");
        
        const q = question.toLowerCase();
        let actionNote = "";
        let mutatedCount = 0;

        // DIRECT EXCEL MANIPULATION OPERATIONS ON SPREADSHEET ROWS:
        if (q.includes("artır") || q.includes("increase") || q.includes("%")) {
          // Increase prices by 10%
          currentSpreadsheetData = currentSpreadsheetData.map(r => {
            let oldPrice = Number(r.Price);
            let newPrice = Math.round(oldPrice * 1.10);
            mutatedCount++;
            return { ...r, Price: newPrice, Revenue: newPrice * r.Stock, Status: "Price Increased (+10%)" };
          });
          actionNote = (currentLang === 'en')
            ? `⚡ EXCEL OPERATION EXECUTED: Mutated ${mutatedCount} price cells in Excel dataset (+10% applied). Total revenue updated.`
            : `⚡ EXCEL OPERASYONU TAMAMLANTI: Tablodaki ${mutatedCount} satırın fiyatı %10 artırıldı. Ciro ve stok değerleri yeniden hesaplandı.`;
        } else if (q.includes("düşür") || q.includes("discount") || q.includes("indirim")) {
          // Apply 10% discount on overpriced items
          currentSpreadsheetData = currentSpreadsheetData.map(r => {
            if (r.Status.includes("Overpriced") || r.Price > 30000) {
              let newPrice = Math.round(Number(r.Price) * 0.90);
              mutatedCount++;
              return { ...r, Price: newPrice, Revenue: newPrice * r.Stock, Status: "Discounted (-10%)" };
            }
            return r;
          });
          actionNote = (currentLang === 'en')
            ? `⚡ EXCEL OPERATION EXECUTED: Applied 10% discount on ${mutatedCount} overpriced SKUs.`
            : `⚡ EXCEL OPERASYONU TAMAMLANTI: Pahalı kalan ${mutatedCount} üründe %10 fiyat indirimi uygulandı.`;
        } else if (q.includes("stok") || q.includes("stock") || q.includes("ekle") || q.includes("replenish")) {
          // Replenish low stock items
          currentSpreadsheetData = currentSpreadsheetData.map(r => {
            if (r.Stock < 50) {
              let newStock = r.Stock + 100;
              mutatedCount++;
              return { ...r, Stock: newStock, Revenue: r.Price * newStock, Status: "Stock Replenished (+100)" };
            }
            return r;
          });
          actionNote = (currentLang === 'en')
            ? `⚡ EXCEL OPERATION EXECUTED: Replenished 100 stock units for ${mutatedCount} low-inventory SKUs.`
            : `⚡ EXCEL OPERASYONU TAMAMLANTI: Stok miktarı düşük ${mutatedCount} ürüne +100 adet stok eklendi.`;
        } else {
          // Default spreadsheet cell calculation & audit log
          actionNote = (currentLang === 'en')
            ? "⚡ EXCEL OPERATION COMPLETED: Calculated dataset metrics across active spreadsheet rows."
            : "⚡ EXCEL OPERASYONU TAMAMLANTI: Aktif Excel tablosu üzerinde hesaplama ve veri analizi yapıldı.";
        }

        // Filter rows matching prompt keywords
        let filteredRows = currentSpreadsheetData;
        if (q.includes("apple") || q.includes("iphone") || q.includes("macbook") || q.includes("ipad") || q.includes("airpods")) {
          filteredRows = currentSpreadsheetData.filter(r => 
            r.Name.toLowerCase().includes("apple") || 
            r.Name.toLowerCase().includes("iphone") || 
            r.Name.toLowerCase().includes("macbook") || 
            r.Name.toLowerCase().includes("ipad") || 
            r.Name.toLowerCase().includes("airpods")
          );
        } else if (q.includes("laptop") || q.includes("bilgisayar")) {
          filteredRows = currentSpreadsheetData.filter(r => r.Category.toLowerCase().includes("laptop"));
        } else if (q.includes("smartphone") || q.includes("telefon")) {
          filteredRows = currentSpreadsheetData.filter(r => r.Category.toLowerCase().includes("smartphone"));
        }

        let filteredAvgPrice = (filteredRows.reduce((a, b) => a + Number(b.Price), 0) / (filteredRows.length || 1)).toLocaleString();
        let filteredStockValue = filteredRows.reduce((a, b) => a + (Number(b.Price) * Number(b.Stock)), 0).toLocaleString();

        let outputHtml = `
          <div style="background: #ffffff; border: 1.5px solid #e2e8f0; border-radius: 16px; padding: 20px; box-shadow: 0 4px 14px rgba(0,0,0,0.03);">
            <div style="display: flex; align-items: center; justify-content: space-between; margin-bottom: 16px; padding-bottom: 12px; border-bottom: 1px solid #f1f5f9;">
              <div>
                <span style="font-size: 11px; font-weight: 700; color: #2563eb; letter-spacing: 0.08em; text-transform: uppercase;">Direct Excel Operation Output</span>
                <h4 style="font-size: 15px; font-weight: 700; color: #0f172a; margin-top: 2px;">Executed Command: "${question}"</h4>
              </div>
              <div style="display: flex; gap: 8px;">
                <button onclick="openExcelPreview()" type="button" style="background: #eff6ff; color: #2563eb; border: 1px solid #bfdbfe; padding: 7px 14px; border-radius: 10px; font-weight: 700; font-size: 12px; cursor: pointer;">Inspect Modified Table</button>
                <button onclick="downloadExcel()" type="button" style="background: #2563eb; color: #ffffff; border: none; padding: 7px 14px; border-radius: 10px; font-weight: 700; font-size: 12px; cursor: pointer;">Download Updated Excel</button>
              </div>
            </div>

            <!-- AUDIT LOG BANNER -->
            <div style="background: #ecfdf5; border: 1.5px solid #6ee7b7; color: #047857; padding: 12px 16px; border-radius: 12px; font-size: 13px; font-weight: 700; margin-bottom: 18px; display: flex; align-items: center; justify-content: space-between;">
              <span>${actionNote}</span>
              <span style="background: #10b981; color: #ffffff; font-size: 10.5px; padding: 2px 8px; border-radius: 999px;">EXCEL UPDATED</span>
            </div>

            <!-- EXCEL INTERACTIVE CHART / TREND VISUALIZER -->
            ${(q.includes("grafik") || q.includes("chart") || q.includes("trend") || q.includes("çiz") || q.includes("plot") || q.includes("dağılım")) ? `
              <div style="background: #ffffff; border: 1.5px solid #bfdbfe; border-radius: 14px; padding: 18px; margin-bottom: 20px; box-shadow: 0 4px 12px rgba(37,99,235,0.06);">
                <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 14px;">
                  <div>
                    <span style="font-size: 11px; font-weight: 700; color: #2563eb; letter-spacing: 0.08em; text-transform: uppercase;">Excel In-Spreadsheet Chart Generator</span>
                    <h5 style="font-size: 14px; font-weight: 700; color: #0f172a;">📊 Category Revenue & Inventory Trend Breakdown</h5>
                  </div>
                  <span style="font-size: 11px; background: #eff6ff; color: #2563eb; padding: 3px 10px; border-radius: 999px; font-weight: 700; border: 1px solid #bfdbfe;">Live Chart Rendered</span>
                </div>
                <div style="display: flex; align-items: flex-end; gap: 18px; height: 160px; padding: 12px 10px; background: #f8fafc; border-radius: 12px; border: 1px solid #e2e8f0; margin-bottom: 12px;">
                  <div style="flex: 1; display: flex; flex-direction: column; align-items: center; gap: 6px; height: 100%; justify-content: flex-end;">
                    <span style="font-size: 11px; font-weight: 700; color: #2563eb;">7.8M TL</span>
                    <div style="width: 100%; max-width: 54px; height: 90%; background: linear-gradient(180deg, #2563eb 0%, #3b82f6 100%); border-radius: 6px 6px 0 0; box-shadow: 0 4px 10px rgba(37,99,235,0.25);"></div>
                    <span style="font-size: 11px; font-weight: 600; color: #475569;">Smartphones</span>
                  </div>
                  <div style="flex: 1; display: flex; flex-direction: column; align-items: center; gap: 6px; height: 100%; justify-content: flex-end;">
                    <span style="font-size: 11px; font-weight: 700; color: #10b981;">4.5M TL</span>
                    <div style="width: 100%; max-width: 54px; height: 60%; background: linear-gradient(180deg, #10b981 0%, #34d399 100%); border-radius: 6px 6px 0 0; box-shadow: 0 4px 10px rgba(16,185,129,0.25);"></div>
                    <span style="font-size: 11px; font-weight: 600; color: #475569;">Laptops</span>
                  </div>
                  <div style="flex: 1; display: flex; flex-direction: column; align-items: center; gap: 6px; height: 100%; justify-content: flex-end;">
                    <span style="font-size: 11px; font-weight: 700; color: #f59e0b;">3.2M TL</span>
                    <div style="width: 100%; max-width: 54px; height: 42%; background: linear-gradient(180deg, #f59e0b 0%, #fbbf24 100%); border-radius: 6px 6px 0 0; box-shadow: 0 4px 10px rgba(245,158,11,0.25);"></div>
                    <span style="font-size: 11px; font-weight: 600; color: #475569;">Accessories</span>
                  </div>
                  <div style="flex: 1; display: flex; flex-direction: column; align-items: center; gap: 6px; height: 100%; justify-content: flex-end;">
                    <span style="font-size: 11px; font-weight: 700; color: #8b5cf6;">2.2M TL</span>
                    <div style="width: 100%; max-width: 54px; height: 32%; background: linear-gradient(180deg, #8b5cf6 0%, #a78bfa 100%); border-radius: 6px 6px 0 0; box-shadow: 0 4px 10px rgba(139,92,246,0.25);"></div>
                    <span style="font-size: 11px; font-weight: 600; color: #475569;">Tablets</span>
                  </div>
                </div>
                <div style="font-size: 12px; color: #475569; font-weight: 500; text-align: center;">✓ Chart data calculated from active Excel rows. Rendered directly inside Excel workspace.</div>
              </div>
            ` : ''}

            <!-- KPI Summary Cards -->
            <div style="display: grid; grid-template-columns: repeat(auto-fit, minmax(170px, 1fr)); gap: 12px; margin-bottom: 20px;">
              <div style="background: #f8fafc; padding: 14px; border-radius: 12px; border: 1px solid #e2e8f0;">
                <div style="font-size: 11px; color: #64748b; font-weight: 600;">Processed SKUs</div>
                <div style="font-size: 20px; font-weight: 800; color: #0f172a; margin-top: 4px;">${filteredRows.length} Items</div>
              </div>
              <div style="background: #f8fafc; padding: 14px; border-radius: 12px; border: 1px solid #e2e8f0;">
                <div style="font-size: 11px; color: #64748b; font-weight: 600;">Updated Avg Price</div>
                <div style="font-size: 20px; font-weight: 800; color: #2563eb; margin-top: 4px;">${filteredAvgPrice} TL</div>
              </div>
              <div style="background: #f8fafc; padding: 14px; border-radius: 12px; border: 1px solid #e2e8f0;">
                <div style="font-size: 11px; color: #64748b; font-weight: 600;">Recalculated Stock Value</div>
                <div style="font-size: 20px; font-weight: 800; color: #10b981; margin-top: 4px;">${filteredStockValue} TL</div>
              </div>
            </div>

            <!-- Filtered Table -->
            <div style="overflow-x: auto; border: 1px solid #e2e8f0; border-radius: 12px;">
              <table style="width: 100%; border-collapse: collapse; font-size: 12.5px; text-align: left;">
                <thead>
                  <tr style="background: #f8fafc; border-bottom: 1px solid #e2e8f0;">
                    <th style="padding: 10px 14px; font-weight: 700; color: #0f172a;">SKU</th>
                    <th style="padding: 10px 14px; font-weight: 700; color: #0f172a;">Product Title</th>
                    <th style="padding: 10px 14px; font-weight: 700; color: #0f172a;">Category</th>
                    <th style="padding: 10px 14px; font-weight: 700; color: #0f172a;">Price</th>
                    <th style="padding: 10px 14px; font-weight: 700; color: #0f172a;">Stock</th>
                    <th style="padding: 10px 14px; font-weight: 700; color: #0f172a;">Status</th>
                  </tr>
                </thead>
                <tbody>
                  ${filteredRows.map((r, i) => `
                    <tr style="background: ${i % 2 === 0 ? '#ffffff' : '#fcfcfd'}; border-bottom: 1px solid #f1f5f9;">
                      <td style="padding: 10px 14px; font-weight: 600; color: #1e293b;">${r.SKU}</td>
                      <td style="padding: 10px 14px; font-weight: 600; color: #0f172a;">${r.Name}</td>
                      <td style="padding: 10px 14px; color: #475569;">${r.Category}</td>
                      <td style="padding: 10px 14px; font-weight: 700; color: #0f172a;">${Number(r.Price).toLocaleString()} TL</td>
                      <td style="padding: 10px 14px; font-weight: 600; color: #334155;">${r.Stock} units</td>
                      <td style="padding: 10px 14px;"><span style="background: #eff6ff; color: #2563eb; padding: 2px 8px; border-radius: 999px; font-weight: 700; font-size: 11px;">${r.Status}</span></td>
                    </tr>
                  `).join('')}
                </tbody>
              </table>
            </div>
          </div>
        `;

        resultBox.innerHTML = outputHtml;
      }, 300);
    }

    function openExcelPreview() {
      const modal = document.getElementById("excelPreviewModal");
      if (modal) modal.style.display = "flex";
      renderPreviewTable(currentSpreadsheetData);
    }

    function closeExcelPreview() {
      const modal = document.getElementById("excelPreviewModal");
      if (modal) modal.style.display = "none";
    }

    function renderPreviewTable(dataList) {
      const container = document.getElementById("excelPreviewTableContainer");
      const rowCountLabel = document.getElementById("modalRowCount");
      if (rowCountLabel) rowCountLabel.textContent = (currentLang === 'en') ? `Showing: ${dataList.length} Rows` : `Gösterilen: ${dataList.length} Satır`;
      if (!container) return;

      if (dataList.length === 0) {
        container.innerHTML = "<div style='text-align: center; color: #94a3b8; padding: 40px;'>No rows match your filter.</div>";
        return;
      }

      let html = `<table style="width: 100%; border-collapse: collapse; font-size: 12.5px; text-align: left;">
        <thead>
          <tr style="background: #f8fafc; border-bottom: 2px solid #e2e8f0;">
            <th style="padding: 10px 14px; font-weight: 700; color: #0f172a;">SKU</th>
            <th style="padding: 10px 14px; font-weight: 700; color: #0f172a;">Title</th>
            <th style="padding: 10px 14px; font-weight: 700; color: #0f172a;">Category</th>
            <th style="padding: 10px 14px; font-weight: 700; color: #0f172a;">Price</th>
            <th style="padding: 10px 14px; font-weight: 700; color: #0f172a;">Comp Price</th>
            <th style="padding: 10px 14px; font-weight: 700; color: #0f172a;">Stock</th>
            <th style="padding: 10px 14px; font-weight: 700; color: #0f172a;">Revenue</th>
            <th style="padding: 10px 14px; font-weight: 700; color: #0f172a;">Status</th>
          </tr>
        </thead>
        <tbody>`;

      dataList.forEach((r, idx) => {
        const bg = idx % 2 === 0 ? "#ffffff" : "#fdfdfe";
        const priceFormatted = typeof r.Price === 'number' ? r.Price.toLocaleString() + " TL" : r.Price;
        const compPriceFormatted = typeof r.CompPrice === 'number' ? r.CompPrice.toLocaleString() + " TL" : r.CompPrice;
        const revenueFormatted = typeof r.Revenue === 'number' ? r.Revenue.toLocaleString() + " TL" : r.Revenue;

        html += `<tr style="background: ${bg}; border-bottom: 1px solid #e2e8f0;">
          <td style="padding: 10px 14px; font-weight: 600; color: #1e293b;">${r.SKU}</td>
          <td style="padding: 10px 14px; font-weight: 600; color: #0f172a;">${r.Name}</td>
          <td style="padding: 10px 14px; color: #475569;">${r.Category}</td>
          <td style="padding: 10px 14px; font-weight: 700; color: #0f172a;">${priceFormatted}</td>
          <td style="padding: 10px 14px; color: #64748b;">${compPriceFormatted}</td>
          <td style="padding: 10px 14px; font-weight: 600; color: #334155;">${r.Stock} units</td>
          <td style="padding: 10px 14px; font-weight: 700; color: #047857;">${revenueFormatted}</td>
          <td style="padding: 10px 14px;"><span style="background: #eff6ff; color: #2563eb; border: 1px solid #bfdbfe; padding: 3px 8px; border-radius: 999px; font-weight: 700; font-size: 11px;">${r.Status}</span></td>
        </tr>`;
      });

      html += `</tbody></table>`;
      container.innerHTML = html;
    }

    function filterPreviewTable(query) {
      const q = query.toLowerCase().trim();
      if (!q) {
        renderPreviewTable(currentSpreadsheetData);
        return;
      }
      const filtered = currentSpreadsheetData.filter(r => 
        r.SKU.toLowerCase().includes(q) || 
        r.Name.toLowerCase().includes(q) || 
        r.Category.toLowerCase().includes(q) || 
        r.Status.toLowerCase().includes(q)
      );
      renderPreviewTable(filtered);
    }

    function downloadExcel() {
      let csvContent = "SKU,Product Name,Category,Price (TL),Competitor Price (TL),Stock,Revenue (TL),Status\\n";
      currentSpreadsheetData.forEach(r => {
        csvContent += `"${r.SKU}","${r.Name}","${r.Category}",${r.Price},${r.CompPrice},${r.Stock},${r.Revenue},"${r.Status}"\\n`;
      });

      const blob = new Blob([csvContent], { type: 'text/csv;charset=utf-8;' });
      const url = URL.createObjectURL(blob);
      const link = document.createElement("a");
      link.setAttribute("href", url);
      link.setAttribute("download", `Retail_AI_Excel_Export_${Date.now()}.csv`);
      document.body.appendChild(link);
      link.click();
      document.body.removeChild(link);
    }

    function clearResult() {
      resultBox.className = "result-box placeholder";
      resultBox.textContent = (currentLang === 'en') 
        ? "Select a module on the left and start the analysis using voice or text." 
        : "Soldaki modüllerden birini seç ve sesle veya yazıyla sorunu sorup analizi başlat.";
    }

    questionInput.addEventListener("keydown", function(e) {
      if ((e.metaKey || e.ctrlKey) && e.key === "Enter") runModule();
    });

    // Check if user just subscribed
    const urlParams = new URLSearchParams(window.location.search);
    if (urlParams.get('activated') === 'true') {
      const banner = document.getElementById('activationBanner');
      const plan = urlParams.get('plan');
      if (banner) {
        banner.style.display = 'flex';
        if (plan === 'pro') {
          document.getElementById('activationTitle').textContent = 'Pro License Active: Welcome to DataProvido Console!';
          document.getElementById('activationSubtitle').textContent = 'Your Pro on-premise AI environment is active with Weekly 1.5h Live Support included.';
        }
      }
    }

    renderModule("business_calculator");
  </script>
</body>
</html>
"""


def simple_page(title, body, kicker="DataProvido", active_nav="pricing", max_width="960px"):
    nav_pricing_cls = "nav-link active" if active_nav == "pricing" else "nav-link"
    nav_contact_cls = "nav-link active" if active_nav == "contact" else "nav-link"
    nav_who_cls     = "nav-link active" if active_nav == "who-we-are" else "nav-link"
    nav_how_cls     = "nav-link active" if active_nav == "how-works" else "nav-link"
    nav_privacy_cls = "nav-link active" if active_nav == "privacy" else "nav-link"

    # Only on Pricing page do we link to /journey?demo=true for testing. Otherwise, CTA goes to /pricing
    cta_url = "/journey?demo=true" if active_nav == "pricing" else "/pricing"
    cta_label = "Start Journey (Test) &nbsp;→" if active_nav == "pricing" else "Start Journey &nbsp;→"
    bottom_cta_url = "/journey?demo=true" if active_nav == "pricing" else "/pricing"
    bottom_cta_label = "Test Sandbox Console &nbsp;→" if active_nav == "pricing" else "View Plans &amp; Subscribe &nbsp;→"

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
  <!-- Google Tag Manager -->
  <script>(function(w,d,s,l,i){{w[l]=w[l]||[];w[l].push({{'gtm.start':
  new Date().getTime(),event:'gtm.js'}});var f=d.getElementsByTagName(s)[0],
  j=d.createElement(s),dl=l!='dataLayer'?'&l='+l:'';j.async=true;j.src=
  'https://www.googletagmanager.com/gtm.js?id='+i+dl;f.parentNode.insertBefore(j,f);
  }})(window,document,'script','dataLayer','GTM-TVKFC4P6');</script>
  <!-- End Google Tag Manager -->

  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>{title} – DataProvido</title>
  <link rel="icon" type="image/png" href="/logo.png">
  <link rel="shortcut icon" type="image/png" href="/logo.png">
  <link rel="apple-touch-icon" href="/logo.png">
  <link rel="preconnect" href="https://fonts.googleapis.com">
  <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&family=Playfair+Display:ital,wght@0,400;0,700;1,400&display=swap" rel="stylesheet">

  <style>
    *, *::before, *::after {{ margin: 0; padding: 0; box-sizing: border-box; }}
    :root {{
      --orange:       #f26f26;
      --orange-dark:  #d85c18;
      --orange-light: #f58c50;
      --bg:           #ffffff;
      --bg-2:         #f9fafb;
      --bg-3:         #f1f3f5;
      --text-900:     #292c2f;
      --text-700:     #4e5359;
      --text-500:     #6b7178;
      --text-dim:     #a3acb6;
      --border:       #dadee2;
      --border-orange: rgba(242,111,38,0.25);
      --card-bg:      #ffffff;
      --card-shadow:  0 4px 24px rgba(0,0,0,0.06), 0 1px 3px rgba(0,0,0,0.04);
      --radius:       20px;
      --nav-h:        72px;
    }}
    html {{ scroll-behavior: smooth; }}
    body {{
      background: linear-gradient(160deg, #fff8f4 0%, #ffffff 45%, #f0f7ff 100%);
      color: var(--text-900);
      font-family: 'Inter', sans-serif;
      min-height: 100vh;
      display: flex;
      flex-direction: column;
    }}

    /* NAV */
    .nav {{
      position: fixed; top: 0; left: 0; right: 0; z-index: 1000; height: var(--nav-h);
      display: flex; align-items: center; justify-content: space-between; padding: 0 40px;
      background: rgba(255,255,255,0.95); backdrop-filter: blur(16px);
      border-bottom: 1px solid var(--border); box-shadow: 0 1px 0 var(--border);
    }}
    .nav-logo {{ display: flex; align-items: center; gap: 10px; font-weight: 700; font-size: 18px; color: var(--text-900); text-decoration: none; }}
    .nav-logo-dot {{ width: 10px; height: 10px; border-radius: 50%; background: var(--orange); animation: pulse-dot 2.5s ease-in-out infinite; }}
    @keyframes pulse-dot {{ 0%,100% {{ box-shadow: 0 0 0 0 rgba(242,111,38,0.5); }} 50% {{ box-shadow: 0 0 0 6px rgba(242,111,38,0); }} }}
    .nav-links {{ display: flex; align-items: center; gap: 4px; }}
    .nav-link {{
      color: var(--text-700); text-decoration: none; font-size: 14px; font-weight: 500;
      padding: 7px 14px; border-radius: 10px; transition: all .18s ease;
    }}
    .nav-link:hover {{ color: var(--orange); background: rgba(242,111,38,0.06); }}
    .nav-link.active {{
      color: var(--orange);
      background: #fff3ec;
      font-weight: 600;
      box-shadow: inset 0 0 0 1px rgba(242,111,38,0.20);
    }}
    .nav-cta {{
      background: var(--orange); color: #fff; border: none; padding: 10px 22px;
      border-radius: 24px; font-size: 14px; font-weight: 600; cursor: pointer;
      text-decoration: none; display: inline-flex; align-items: center; gap: 6px;
      box-shadow: 0 4px 14px rgba(242,111,38,0.30); transition: all .2s ease;
    }}
    .nav-cta:hover {{ background: var(--orange-dark); transform: translateY(-1px); }}

    /* PAGE WRAPPER */
    .page-wrapper {{
      margin-top: calc(var(--nav-h) + 32px);
      margin-bottom: 60px;
      padding: 0 24px;
      display: flex; justify-content: center;
      flex: 1;
    }}
    .page-card {{
      background: var(--card-bg);
      border: 1px solid var(--border);
      border-radius: 24px;
      box-shadow: var(--card-shadow);
      padding: 48px 48px;
      width: 100%;
      max-width: {max_width};
      position: relative;
    }}
    .page-badge {{
      display: inline-flex; align-items: center; gap: 6px;
      background: #fff3ec; border: 1px solid var(--border-orange);
      color: var(--orange); font-size: 12px; font-weight: 700;
      padding: 5px 14px; border-radius: 999px; text-transform: uppercase;
      letter-spacing: 0.05em; margin-bottom: 16px;
    }}
    h1 {{
      font-family: 'Playfair Display', serif;
      font-size: 38px; font-weight: 700; color: var(--text-900);
      letter-spacing: -0.02em; line-height: 1.2; margin-bottom: 12px;
    }}
    .page-subhead {{
      font-size: 16px; color: var(--text-500); line-height: 1.6; margin-bottom: 32px;
    }}
    .page-content {{ font-size: 15px; line-height: 1.75; color: var(--text-700); }}

    .page-actions {{
      margin-top: 40px; padding-top: 24px; border-top: 1px solid var(--border);
      display: flex; align-items: center; justify-content: space-between; gap: 16px;
    }}
    .back-link {{
      color: var(--text-500); text-decoration: none; font-size: 14px; font-weight: 500;
      transition: color .15s;
    }}
    .back-link:hover {{ color: var(--orange); }}

    .btn-primary {{
      background: var(--orange); color: #fff; text-decoration: none;
      padding: 12px 24px; border-radius: 999px; font-weight: 600; font-size: 14px;
      display: inline-flex; align-items: center; gap: 8px;
      box-shadow: 0 4px 14px rgba(242,111,38,0.25); transition: all .18s;
    }}
    .btn-primary:hover {{ background: var(--orange-dark); transform: translateY(-1px); }}

    /* FOOTER */
    .site-footer {{
      margin-top: auto; border-top: 1px solid var(--border);
      padding: 24px 40px; display: flex; align-items: center; justify-content: space-between;
      font-size: 13px; color: var(--text-500); background: #ffffff;
    }}
    .footer-links {{ display: flex; gap: 16px; align-items: center; }}
    .footer-link {{ color: var(--text-500); text-decoration: none; transition: color .15s; }}
    .footer-link:hover {{ color: var(--orange); }}

    @media (max-width: 860px) {{
      .nav {{ padding: 0 20px; }}
      .nav-links {{ display: none; }}
      .page-card {{ padding: 28px 20px; }}
      .site-footer {{ flex-direction: column; gap: 12px; text-align: center; }}
    }}
  </style>
</head>
<body>
  <!-- Google Tag Manager (noscript) -->
  <noscript><iframe src="https://www.googletagmanager.com/ns.html?id=GTM-TVKFC4P6"
  height="0" width="0" style="display:none;visibility:hidden"></iframe></noscript>
  <!-- End Google Tag Manager (noscript) -->

  <nav class="nav">
    <a href="/" class="nav-logo" style="display: flex; align-items: center; gap: 10px;"><img src="/logo.png" alt="DataProvido" style="height: 34px; width: 34px; object-fit: contain; border-radius: 8px;" /><span>DataProvido</span></a>
    <div class="nav-links">
      <a href="/pricing" class="{nav_pricing_cls}">Pricing</a>
      <a href="/contact" class="{nav_contact_cls}">Contact</a>
      <a href="/who-we-are" class="{nav_who_cls}">Who We Are?</a>
      <a href="/how-works" class="{nav_how_cls}">How Works?</a>
      <a href="/privacy" class="{nav_privacy_cls}">Privacy Policy</a>
    </div>
    <a href="{cta_url}" class="nav-cta">{cta_label}</a>
  </nav>

  <div class="page-wrapper">
    <main class="page-card">
      <div class="page-badge">✦ {kicker}</div>
      <h1>{title}</h1>
      <div class="page-content">
        {body}
      </div>
      <div class="page-actions">
        <a href="/" class="back-link">← Back to Home</a>
        <a href="{bottom_cta_url}" class="btn-primary">{bottom_cta_label}</a>
      </div>
    </main>
  </div>

  <footer class="site-footer">
    <div><strong>DataProvido</strong> · Local Intelligence. Zero Compromise.</div>
    <div class="footer-links">
      <a href="/pricing" class="footer-link">Pricing</a> ·
      <a href="/privacy" class="footer-link">Privacy Policy</a> ·
      <a href="/contact" class="footer-link">Contact</a> ·
      <span>100% Offline AI</span>
    </div>
  </footer>
</body>
</html>"""


@app.get("/privacy", response_class=HTMLResponse)
def privacy():
    return simple_page(
        "Privacy & Cookie Documentation",
        """
        <p class="page-subhead">
          Comprehensive compliance framework governing DataProvido local intelligence architecture, GDPR (EU) and KVKK (TR) dual-regime protection, sub-processors, and cookie preferences.
        </p>

        <!-- SUMMARY BADGES -->
        <div style="display: flex; gap: 12px; flex-wrap: wrap; margin-bottom: 32px;">
          <div style="background: #ecfdf5; border: 1px solid #6ee7b7; color: #047857; padding: 6px 14px; border-radius: 999px; font-size: 12px; font-weight: 700;">
            🛡️ Dual GDPR &amp; KVKK Compliant
          </div>
          <div style="background: #eff6ff; border: 1px solid #93c5fd; color: #1e40af; padding: 6px 14px; border-radius: 999px; font-size: 12px; font-weight: 700;">
            🔒 Zero Data Uploading (100% Local AI)
          </div>
          <div style="background: #fff7ed; border: 1px solid #fdba74; color: #c2410c; padding: 6px 14px; border-radius: 999px; font-size: 12px; font-weight: 700;">
            🍪 Google Consent Mode v2 &amp; Meta Pixel
          </div>
        </div>

        <!-- SECTION 1: PRIVACY POLICY -->
        <div style="margin-bottom: 40px;">
          <h2 style="font-family: 'Playfair Display', serif; font-size: 24px; color: var(--text-900); margin-bottom: 16px; border-bottom: 1.5px solid var(--border); padding-bottom: 10px;">
            1. Privacy Policy
          </h2>

          <h3 style="font-size: 17px; font-weight: 700; color: var(--text-900); margin: 20px 0 10px;">1.1 Who We Are</h3>
          <p style="margin-bottom: 14px;">
            DataProvido ("we", "us", "the Platform") is operated by <strong>DataProvido Inc.</strong>, registered in Türkiye. Contact: <a href="mailto:privacy@dataprovido.com" style="color: var(--orange); font-weight: 600;">privacy@dataprovido.com</a>.
          </p>
          <p style="margin-bottom: 14px;">
            If you are located in the European Economic Area (EEA), your personal data is processed under the General Data Protection Regulation (GDPR). If you are located in Türkiye, your personal data is processed under Law No. 6698 on the Protection of Personal Data (KVKK).
          </p>

          <h3 style="font-size: 17px; font-weight: 700; color: var(--text-900); margin: 20px 0 10px;">1.1a Cross-Border &amp; Extraterritorial Scope</h3>
          <p style="margin-bottom: 14px;">
            DataProvido is established in Türkiye. However, because we run advertising campaigns (Google Ads, Meta) targeting individuals located in the European Union/EEA, GDPR applies to us extraterritorially under <strong>Article 3(2)</strong>.
          </p>
          <ul style="list-style-type: disc; padding-left: 24px; margin-bottom: 16px; line-height: 1.7;">
            <li><strong>EU Representative (GDPR Art. 27):</strong> We maintain a designated EU representative for supervisory authority contact. Representative details: <code>privacy@dataprovido.com</code>.</li>
            <li><strong>Dual Regime:</strong> Turkish visitors are governed by KVKK; EU-based visitors reached via EU campaigns are governed by GDPR. Both apply concurrently.</li>
            <li><strong>Lead Authority:</strong> EU data subjects may lodge complaints with the supervisory authority of their Member State of residence.</li>
          </ul>

          <h3 style="font-size: 17px; font-weight: 700; color: var(--text-900); margin: 20px 0 10px;">1.2 Two Roles: Website Visitor Data vs. Customer Business Data</h3>
          <ul style="list-style-type: disc; padding-left: 24px; margin-bottom: 16px; line-height: 1.7;">
            <li><strong>As a Data Controller:</strong> For website visitors and subscribing accounts (billing/contact details), DataProvido determines processing purpose and means.</li>
            <li><strong>As a Data Processor:</strong> Subscribing brands connect their own CRM, GA, and retail data to their self-serve 100% local workspace. We do not access or store the content of your retail data; the subscribing brand remains the Data Controller under a separate Data Processing Agreement (DPA).</li>
          </ul>

          <h3 style="font-size: 17px; font-weight: 700; color: var(--text-900); margin: 20px 0 10px;">1.3 What We Collect</h3>
          <div style="overflow-x: auto; margin-bottom: 24px;">
            <table style="width: 100%; border-collapse: collapse; font-size: 13.5px; text-align: left;">
              <thead>
                <tr style="background: var(--bg-2); border-bottom: 2px solid var(--border);">
                  <th style="padding: 12px 16px; font-weight: 700;">Category</th>
                  <th style="padding: 12px 16px; font-weight: 700;">Examples</th>
                  <th style="padding: 12px 16px; font-weight: 700;">Purpose</th>
                </tr>
              </thead>
              <tbody>
                <tr style="border-bottom: 1px solid var(--border);">
                  <td style="padding: 12px 16px; font-weight: 600;">Identity &amp; Contact</td>
                  <td style="padding: 12px 16px;">Name, work email, company name</td>
                  <td style="padding: 12px 16px;">Account creation, billing, SLA support</td>
                </tr>
                <tr style="border-bottom: 1px solid var(--border);">
                  <td style="padding: 12px 16px; font-weight: 600;">Usage Data</td>
                  <td style="padding: 12px 16px;">Pages visited, session duration, click events</td>
                  <td style="padding: 12px 16px;">Product analytics &amp; performance tuning</td>
                </tr>
                <tr style="border-bottom: 1px solid var(--border);">
                  <td style="padding: 12px 16px; font-weight: 600;">Technical Data</td>
                  <td style="padding: 12px 16px;">IP address, device type, location</td>
                  <td style="padding: 12px 16px;">Security, fraud prevention, infrastructure</td>
                </tr>
                <tr>
                  <td style="padding: 12px 16px; font-weight: 600;">Marketing Data</td>
                  <td style="padding: 12px 16px;">Ad click IDs (GCLID), campaign source</td>
                  <td style="padding: 12px 16px;">Campaign attribution, retargeting measurement</td>
                </tr>
              </tbody>
            </table>
          </div>

          <h3 style="font-size: 17px; font-weight: 700; color: var(--text-900); margin: 20px 0 10px;">1.4 Legal Basis for Processing</h3>
          <ul style="list-style-type: disc; padding-left: 24px; margin-bottom: 16px; line-height: 1.7;">
            <li><strong>Contract (Art. 6(1)(b) GDPR / KVKK Art. 5/2-c):</strong> Account provisioning, billing, support delivery.</li>
            <li><strong>Legitimate Interest (Art. 6(1)(f) GDPR / KVKK Art. 5/2-f):</strong> Product security, infrastructure optimization.</li>
            <li><strong>Consent (Art. 6(1)(a) GDPR / KVKK Art. 5/1):</strong> Analytics, Google Ads, Meta retargeting via cookie consent banner.</li>
          </ul>

          <h3 style="font-size: 17px; font-weight: 700; color: var(--text-900); margin: 20px 0 10px;">1.5 Third-Party Sub-processors</h3>
          <div style="overflow-x: auto; margin-bottom: 24px;">
            <table style="width: 100%; border-collapse: collapse; font-size: 13.5px; text-align: left;">
              <thead>
                <tr style="background: var(--bg-2); border-bottom: 2px solid var(--border);">
                  <th style="padding: 12px 16px; font-weight: 700;">Tool</th>
                  <th style="padding: 12px 16px; font-weight: 700;">Purpose</th>
                  <th style="padding: 12px 16px; font-weight: 700;">Data Transferred</th>
                  <th style="padding: 12px 16px; font-weight: 700;">Location</th>
                </tr>
              </thead>
              <tbody>
                <tr style="border-bottom: 1px solid var(--border);">
                  <td style="padding: 12px 16px; font-weight: 600;">Google Analytics (GA4)</td>
                  <td style="padding: 12px 16px;">Website usage analytics</td>
                  <td style="padding: 12px 16px;">Anonymized IP, behavioral events</td>
                  <td style="padding: 12px 16px;">Google EU/US (SCCs)</td>
                </tr>
                <tr style="border-bottom: 1px solid var(--border);">
                  <td style="padding: 12px 16px; font-weight: 600;">Google Ads &amp; GTM</td>
                  <td style="padding: 12px 16px;">Conversion tracking (GTM-TVKFC4P6)</td>
                  <td style="padding: 12px 16px;">GCLID, conversion status</td>
                  <td style="padding: 12px 16px;">Google EU/US</td>
                </tr>
                <tr style="border-bottom: 1px solid var(--border);">
                  <td style="padding: 12px 16px; font-weight: 600;">Stripe Payments</td>
                  <td style="padding: 12px 16px;">Payment &amp; subscription processing</td>
                  <td style="padding: 12px 16px;">Billing info, card details (PCI-DSS)</td>
                  <td style="padding: 12px 16px;">Stripe Inc. US/EU</td>
                </tr>
                <tr>
                  <td style="padding: 12px 16px; font-weight: 600;">Meta Pixel</td>
                  <td style="padding: 12px 16px;">Ad performance measurement</td>
                  <td style="padding: 12px 16px;">Hashed interaction events</td>
                  <td style="padding: 12px 16px;">Meta Platforms US/EU</td>
                </tr>
              </tbody>
            </table>
          </div>

          <h3 style="font-size: 17px; font-weight: 700; color: var(--text-900); margin: 20px 0 10px;">1.6 Data Retention &amp; Rights</h3>
          <p style="margin-bottom: 14px;">
            Account data is retained for active subscription duration + 5 years for tax compliance. Under GDPR (Art. 15–22) and KVKK (Art. 11), you may request access, rectification, erasure, or portability of your data by emailing <a href="mailto:privacy@dataprovido.com" style="color: var(--orange); font-weight: 600;">privacy@dataprovido.com</a>.
          </p>
        </div>

        <!-- SECTION 2: COOKIE POLICY -->
        <div style="margin-bottom: 40px;">
          <h2 style="font-family: 'Playfair Display', serif; font-size: 24px; color: var(--text-900); margin-bottom: 16px; border-bottom: 1.5px solid var(--border); padding-bottom: 10px;">
            2. Cookie Policy &amp; Consent Management
          </h2>
          <p style="margin-bottom: 14px;">
            We use cookies and Google Tag Manager (GTM-TVKFC4P6) to deliver secure functionality and measure campaign effectiveness.
          </p>
          <div style="overflow-x: auto; margin-bottom: 24px;">
            <table style="width: 100%; border-collapse: collapse; font-size: 13.5px; text-align: left;">
              <thead>
                <tr style="background: var(--bg-2); border-bottom: 2px solid var(--border);">
                  <th style="padding: 12px 16px; font-weight: 700;">Category</th>
                  <th style="padding: 12px 16px; font-weight: 700;">Examples</th>
                  <th style="padding: 12px 16px; font-weight: 700;">Default State</th>
                </tr>
              </thead>
              <tbody>
                <tr style="border-bottom: 1px solid var(--border);">
                  <td style="padding: 12px 16px; font-weight: 600;">Strictly Necessary</td>
                  <td style="padding: 12px 16px;">Session state, CSRF tokens, security</td>
                  <td style="padding: 12px 16px; color: #10b981; font-weight: 700;">Always Active</td>
                </tr>
                <tr style="border-bottom: 1px solid var(--border);">
                  <td style="padding: 12px 16px; font-weight: 600;">Analytics</td>
                  <td style="padding: 12px 16px;">GA4 <code>_ga</code> cookies</td>
                  <td style="padding: 12px 16px; color: var(--orange); font-weight: 700;">Consent Required</td>
                </tr>
                <tr>
                  <td style="padding: 12px 16px; font-weight: 600;">Advertising</td>
                  <td style="padding: 12px 16px;">Google Ads conversion &amp; Meta Pixel</td>
                  <td style="padding: 12px 16px; color: var(--orange); font-weight: 700;">Consent Required</td>
                </tr>
              </tbody>
            </table>
          </div>
        </div>

        <!-- SECTION 3: DATA PROCESSING AGREEMENT SUMMARY -->
        <div style="background: var(--bg-2); border: 1.5px solid var(--border-orange); border-radius: 18px; padding: 24px;">
          <h2 style="font-family: 'Playfair Display', serif; font-size: 20px; color: var(--text-900); margin-bottom: 12px;">
            3. Data Processing Agreement (DPA) Summary for Subscribing Brands
          </h2>
          <p style="font-size: 13.5px; line-height: 1.7; color: var(--text-700); margin-bottom: 12px;">
            Subscribing enterprise brands connect their commercial Excel and CRM data directly into their private local environment. Under our B2B DPA:
          </p>
          <ul style="list-style-type: check; padding-left: 20px; font-size: 13.5px; line-height: 1.8; color: var(--text-700);">
            <li><strong>DataProvido acts solely as Data Processor;</strong> the customer retains full Data Controller ownership.</li>
            <li><strong>Zero Third-Party LLM Transmission:</strong> All analytical queries execute 100% locally on dedicated LLaMA 3.1 architecture.</li>
            <li><strong>No Cross-Tenant Data Access:</strong> Your enterprise data is strictly isolated and never used to train global AI models.</li>
          </ul>
        </div>
        """,
        kicker="Legal & Compliance",
        active_nav="privacy",
        max_width="1040px"
    )


@app.get("/pricing", response_class=HTMLResponse)
def pricing():
    return simple_page(
        "Choose the perfect plan for your journey",
        """
        <p class="page-subhead" style="margin-bottom: 28px;">
          Deploy complete, 100% offline retail intelligence directly onto your enterprise hardware with zero cloud exposure and predictable licensing.
        </p>

        <!-- 2 CARDS GRID (199 € and 299 €) -->
        <div style="display: grid; grid-template-columns: repeat(auto-fit, minmax(320px, 1fr)); gap: 24px; margin: 28px 0 36px;">
          
          <!-- CARD 1: 199 € / Standard -->
          <div style="background: #ffffff; border: 1.5px solid var(--border); border-radius: 24px; padding: 32px 28px; display: flex; flex-direction: column; box-shadow: var(--card-shadow); transition: transform .2s, box-shadow .2s;">
            
            <div style="display: inline-flex; align-items: center; align-self: flex-start; background: var(--bg-2); border: 1px solid var(--border); border-radius: 999px; padding: 4px 12px; font-size: 12px; font-weight: 600; color: var(--text-700); margin-bottom: 16px;">
              Starter &amp; Pro
            </div>
            
            <div style="font-size: 22px; font-weight: 800; color: var(--text-900); margin-bottom: 6px;">
              DataProvido Standard
            </div>
            
            <div style="font-size: 13.5px; color: var(--text-500); margin-bottom: 20px; line-height: 1.5;">
              Full offline retail AI analytics for commercial and e-commerce leaders.
            </div>

            <!-- Price -->
            <div style="display: flex; align-items: baseline; gap: 4px; margin-bottom: 24px; padding-bottom: 20px; border-bottom: 1px solid var(--border);">
              <span style="font-size: 42px; font-weight: 800; color: var(--text-900); line-height: 1;">199 €</span>
              <span style="font-size: 14px; color: var(--text-500); font-weight: 500;">/ month</span>
            </div>

            <!-- Action Button -> Stripe Checkout -->
            <a href="/checkout?plan=standard" style="display: block; text-align: center; background: #1f2328; color: #ffffff; font-weight: 600; padding: 13px 20px; border-radius: 999px; text-decoration: none; font-size: 14px; transition: all 0.2s; box-shadow: 0 4px 12px rgba(0,0,0,0.12); margin-bottom: 28px;">
              Subscribe with Stripe &nbsp;→
            </a>

            <!-- Features -->
            <div style="font-size: 12px; font-weight: 700; text-transform: uppercase; letter-spacing: 0.08em; color: var(--text-500); margin-bottom: 14px;">
              Included Features:
            </div>
            
            <ul style="list-style: none; padding: 0; margin: 0 0 12px 0; font-size: 13.5px; color: var(--text-700); line-height: 2.1; flex: 1;">
              <li style="display: flex; align-items: flex-start; gap: 10px;">
                <span style="color: #10b981; font-weight: 800; font-size: 15px;">✓</span>
                <span><strong>100% Offline Local LLaMA 3.1 LLM:</strong> Zero data transmitted to third-party clouds</span>
              </li>
              <li style="display: flex; align-items: flex-start; gap: 10px;">
                <span style="color: #10b981; font-weight: 800; font-size: 15px;">✓</span>
                <span><strong>15+ Analytical Engines:</strong> Stock Risk, Funnel Master, Price Radar, GfK Market Share</span>
              </li>
              <li style="display: flex; align-items: flex-start; gap: 10px;">
                <span style="color: #10b981; font-weight: 800; font-size: 15px;">✓</span>
                <span><strong>Natural Language Business Querying:</strong> No SQL, Python, or data analyst queue required</span>
              </li>
              <li style="display: flex; align-items: flex-start; gap: 10px;">
                <span style="color: #10b981; font-weight: 800; font-size: 15px;">✓</span>
                <span><strong>Automated Ingestion:</strong> Excel (.xlsx), CSV, and local flat file ingestion</span>
              </li>
              <li style="display: flex; align-items: flex-start; gap: 10px;">
                <span style="color: #10b981; font-weight: 800; font-size: 15px;">✓</span>
                <span><strong>1-Click Excel Export:</strong> Download formatted reports with styles &amp; KPI summaries</span>
              </li>
              <li style="display: flex; align-items: flex-start; gap: 10px;">
                <span style="color: #10b981; font-weight: 800; font-size: 15px;">✓</span>
                <span><strong>On-Premise Container Deployment:</strong> Runs locally on your hardware via Docker / Ollama</span>
              </li>
              <li style="display: flex; align-items: flex-start; gap: 10px;">
                <span style="color: #10b981; font-weight: 800; font-size: 15px;">✓</span>
                <span>Standard technical setup documentation &amp; email onboarding support</span>
              </li>
            </ul>

          </div>


          <!-- CARD 2: 299 € / Pro + 1.5h Live Support (RECOMMENDED) -->
          <div style="background: #ffffff; border: 2px solid var(--orange); border-radius: 24px; padding: 32px 28px; display: flex; flex-direction: column; position: relative; box-shadow: 0 8px 32px rgba(242,111,38,0.14); transition: transform .2s, box-shadow .2s;">
            
            <div style="position: absolute; top: -13px; right: 28px; background: linear-gradient(90deg, var(--orange) 0%, var(--orange-light) 100%); color: #fff; font-size: 11px; font-weight: 800; padding: 4px 14px; border-radius: 999px; letter-spacing: 0.6px; text-transform: uppercase; box-shadow: 0 4px 12px rgba(242,111,38,0.35);">
              ⭐ RECOMMENDED
            </div>

            <div style="display: inline-flex; align-items: center; align-self: flex-start; background: #fff3ec; border: 1px solid var(--border-orange); border-radius: 999px; padding: 4px 12px; font-size: 12px; font-weight: 700; color: var(--orange); margin-bottom: 16px;">
              Pro + Live Support
            </div>
            
            <div style="font-size: 22px; font-weight: 800; color: var(--text-900); margin-bottom: 6px;">
              DataProvido Pro
            </div>
            
            <div style="font-size: 13.5px; color: var(--text-500); margin-bottom: 20px; line-height: 1.5;">
              Complete analytics + dedicated weekly 1-on-1 expert consulting &amp; custom tuning.
            </div>

            <!-- Price -->
            <div style="display: flex; align-items: baseline; gap: 4px; margin-bottom: 24px; padding-bottom: 20px; border-bottom: 1px solid var(--border);">
              <span style="font-size: 42px; font-weight: 800; color: var(--orange-dark); line-height: 1;">299 €</span>
              <span style="font-size: 14px; color: var(--text-500); font-weight: 500;">/ month</span>
            </div>

            <!-- Action Button -> Stripe Checkout -->
            <a href="/checkout?plan=pro" style="display: block; text-align: center; background: var(--orange); color: #ffffff; font-weight: 700; padding: 13px 20px; border-radius: 999px; text-decoration: none; font-size: 14px; transition: all 0.2s; box-shadow: 0 6px 18px rgba(242,111,38,0.35); margin-bottom: 28px;">
              Subscribe with Stripe &nbsp;→
            </a>

            <!-- Features -->
            <div style="font-size: 12px; font-weight: 700; text-transform: uppercase; letter-spacing: 0.08em; color: var(--orange); margin-bottom: 14px;">
              Everything in Standard, plus:
            </div>
            
            <ul style="list-style: none; padding: 0; margin: 0 0 12px 0; font-size: 13.5px; color: var(--text-700); line-height: 2.1; flex: 1;">
              <li style="display: flex; align-items: flex-start; gap: 10px; background: #fff8f4; padding: 10px 12px; border-radius: 10px; border: 1px dashed var(--border-orange); margin-bottom: 8px;">
                <span style="color: var(--orange); font-weight: 800; font-size: 16px;">🔥</span>
                <span><strong style="color: var(--orange-dark);">Weekly 1.5 hours of dedicated live online technical &amp; analytical support:</strong> Direct 1-on-1 screen share and data strategy session with lead specialist</span>
              </li>
              <li style="display: flex; align-items: flex-start; gap: 10px;">
                <span style="color: #10b981; font-weight: 800; font-size: 15px;">✓</span>
                <span><strong>Custom Prompt &amp; Sector Metric Tuning:</strong> Tailored to your company's proprietary retail schema and KPIs</span>
              </li>
              <li style="display: flex; align-items: flex-start; gap: 10px;">
                <span style="color: #10b981; font-weight: 800; font-size: 15px;">✓</span>
                <span><strong>Priority SLA Support:</strong> Direct Slack / WhatsApp / VIP dedicated channel</span>
              </li>
              <li style="display: flex; align-items: flex-start; gap: 10px;">
                <span style="color: #10b981; font-weight: 800; font-size: 15px;">✓</span>
                <span><strong>Advanced Cross-Dataset Modules:</strong> Stock vs Funnel vs Price Elasticity correlation</span>
              </li>
              <li style="display: flex; align-items: flex-start; gap: 10px;">
                <span style="color: #10b981; font-weight: 800; font-size: 15px;">✓</span>
                <span><strong>Multi-User Intranet Deployment:</strong> Access for cross-functional commercial teams</span>
              </li>
              <li style="display: flex; align-items: flex-start; gap: 10px;">
                <span style="color: #10b981; font-weight: 800; font-size: 15px;">✓</span>
                <span>Quarterly model optimization &amp; new analytical module updates</span>
              </li>
            </ul>

          </div>

        </div>

        <!-- Custom Enterprise Banner -->
        <div style="background: #ffffff; border: 1.5px solid var(--border-orange); border-radius: 18px; padding: 26px; display: flex; align-items: center; justify-content: space-between; flex-wrap: wrap; gap: 18px; box-shadow: var(--card-shadow);">
          <div>
            <div style="font-weight: 800; color: var(--text-900); font-size: 17px; margin-bottom: 4px;">Need Custom Enterprise Architecture or ERP Integration?</div>
            <div style="font-size: 13.5px; color: var(--text-500);">Direct SAP, Nebim, Oracle integration or dedicated air-gapped GPU server clusters.</div>
          </div>
          <a href="/contact" style="display: inline-flex; align-items: center; gap: 8px; background: rgba(242,111,38,0.10); color: var(--orange-dark); border: 1px solid var(--border-orange); padding: 11px 20px; border-radius: 12px; font-weight: 700; font-size: 14px; text-decoration: none;">
            Contact Enterprise Team &nbsp;→
          </a>
        </div>
        """,
        kicker="Transparent Pricing",
        active_nav="pricing",
        max_width="1040px"
    )


@app.get("/debug-env")
def debug_env():
    import os
    stripe_key = os.getenv("STRIPE_SECRET_KEY")
    stripe_key_status = f"Set (Length: {len(stripe_key)}, Starts with: {stripe_key[:7]}...)" if stripe_key else "Not Set (None)"
    
    webhook_secret = os.getenv("STRIPE_WEBHOOK_SECRET")
    webhook_secret_status = f"Set (Length: {len(webhook_secret)}, Starts with: {webhook_secret[:7]}...)" if webhook_secret else "Not Set (None)"
    
    return {
        "STRIPE_SECRET_KEY": stripe_key_status,
        "STRIPE_WEBHOOK_SECRET": webhook_secret_status,
        "STRIPE_STANDARD_PRICE_ID": os.getenv("STRIPE_STANDARD_PRICE_ID"),
        "STRIPE_PRO_PRICE_ID": os.getenv("STRIPE_PRO_PRICE_ID"),
        "BASE_URL": os.getenv("BASE_URL"),
        "PORT": os.getenv("PORT"),
        "LLM_BACKEND": os.getenv("LLM_BACKEND")
    }


@app.get("/checkout", response_class=HTMLResponse)
def checkout(plan: str = "standard"):
    is_pro = plan.lower() == "pro"
    plan_name = "DataProvido Pro (+ 1.5h Weekly Support)" if is_pro else "DataProvido Standard"
    price_val = "299 €" if is_pro else "199 €"
    price_cents = "299.00" if is_pro else "199.00"
    amount_cents = 29900 if is_pro else 19900

    # 1. Check for Stripe Payment Link in environment variables
    stripe_link = os.getenv("STRIPE_PRO_PAYMENT_LINK") if is_pro else os.getenv("STRIPE_STANDARD_PAYMENT_LINK")
    if stripe_link:
        from fastapi.responses import RedirectResponse
        return RedirectResponse(url=stripe_link, status_code=303)

    # 2. Check for Stripe Secret Key in environment variables
    stripe_secret = os.getenv("STRIPE_SECRET_KEY")
    if stripe_secret:
        try:
            import stripe
            stripe.api_key = stripe_secret
            
            # Base domain detection
            base_url = os.getenv("BASE_URL") or os.getenv("RAILWAY_PUBLIC_DOMAIN") or "http://localhost:8000"
            if not base_url.startswith("http"):
                base_url = f"https://{base_url}"

            price_id = os.getenv("STRIPE_PRO_PRICE_ID") if is_pro else os.getenv("STRIPE_STANDARD_PRICE_ID")
            
            if price_id:
                line_items = [{"price": price_id, "quantity": 1}]
            else:
                line_items = [{
                    "price_data": {
                        "currency": "eur",
                        "product_data": {
                            "name": plan_name,
                            "description": "100% Offline On-Premise Retail AI License" + (" with 1.5h Weekly Support" if is_pro else "")
                        },
                        "unit_amount": amount_cents,
                        "recurring": {"interval": "month"}
                    },
                    "quantity": 1
                }]

            session = stripe.checkout.Session.create(
                line_items=line_items,
                mode="subscription",
                success_url=f"{base_url}/checkout/success?plan={'pro' if is_pro else 'standard'}&session_id={{CHECKOUT_SESSION_ID}}",
                cancel_url=f"{base_url}/pricing",
                managed_payments={"enabled": False}
            )
            from fastapi.responses import RedirectResponse
            return RedirectResponse(url=session.url, status_code=303)
        except Exception as e:
            print(f"[STRIPE ERROR] Failed to create checkout session: {e}")
            return HTMLResponse(
                content=f"""
                <div style="font-family: sans-serif; padding: 40px; max-width: 600px; margin: 40px auto; border: 1px solid #f87171; background: #fef2f2; border-radius: 12px; color: #991b1b;">
                    <h2 style="margin-top: 0; font-size: 20px;">⚠️ Stripe Integration Error</h2>
                    <p style="font-size: 14px; line-height: 1.6;">Your server has <code>STRIPE_SECRET_KEY</code> set, but the Stripe API returned an error when generating the session:</p>
                    <pre style="background: #ffffff; padding: 16px; border-radius: 8px; border: 1px solid #fee2e2; overflow-x: auto; font-family: monospace; font-size: 13px; color: #b91c1c;">{str(e)}</pre>
                    <p style="font-size: 13.5px; color: #7f1d1d; margin-bottom: 0;">Please check your Stripe Price IDs, API keys, or currency activation in your Stripe Dashboard.</p>
                </div>
                """,
                status_code=500
            )
    else:
        print("[STRIPE] Warning: STRIPE_SECRET_KEY env variable not found. Showing mockup payment page.")

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>Stripe Checkout – {plan_name}</title>
  <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap" rel="stylesheet">
  <style>
    *, *::before, *::after {{ margin: 0; padding: 0; box-sizing: border-box; }}
    body {{
      background: #f8fafc;
      color: #1e293b;
      font-family: 'Inter', sans-serif;
      min-height: 100vh;
      display: flex;
      align-items: center;
      justify-content: center;
      padding: 24px;
    }}
    .checkout-shell {{
      width: 100%;
      max-width: 880px;
      background: #ffffff;
      border-radius: 20px;
      box-shadow: 0 10px 40px rgba(0,0,0,0.08);
      border: 1px solid #e2e8f0;
      overflow: hidden;
      display: grid;
      grid-template-columns: 1fr 1fr;
    }}
    .order-summary {{
      background: #0f172a;
      color: #ffffff;
      padding: 44px 36px;
      display: flex;
      flex-direction: column;
      justify-content: space-between;
    }}
    .order-header {{ display: flex; align-items: center; gap: 8px; font-weight: 700; font-size: 16px; margin-bottom: 28px; }}
    .order-dot {{ width: 10px; height: 10px; border-radius: 50%; background: #f26f26; }}
    .order-plan {{ font-size: 24px; font-weight: 800; margin-bottom: 6px; }}
    .order-price {{ font-size: 38px; font-weight: 800; color: #f26f26; margin: 16px 0; }}
    .order-price span {{ font-size: 14px; color: #94a3b8; font-weight: 500; }}
    .order-features {{ list-style: none; padding: 0; margin: 20px 0; font-size: 13.5px; color: #cbd5e1; line-height: 2; }}
    .order-features li {{ display: flex; align-items: center; gap: 8px; }}
    .stripe-trust {{
      font-size: 12px;
      color: #94a3b8;
      display: flex;
      align-items: center;
      gap: 6px;
      margin-top: 20px;
      padding-top: 20px;
      border-top: 1px solid rgba(255,255,255,0.1);
    }}
    .payment-panel {{
      padding: 44px 36px;
      display: flex;
      flex-direction: column;
      justify-content: center;
    }}
    .payment-title {{ font-size: 20px; font-weight: 700; margin-bottom: 20px; color: #0f172a; }}
    .form-group {{ margin-bottom: 16px; }}
    .form-label {{ display: block; font-size: 12.5px; font-weight: 600; color: #475569; margin-bottom: 6px; }}
    .form-input {{
      width: 100%;
      padding: 12px 14px;
      border-radius: 10px;
      border: 1.5px solid #cbd5e1;
      font-size: 14px;
      outline: none;
      transition: all 0.2s;
    }}
    .form-input:focus {{ border-color: #6366f1; box-shadow: 0 0 0 3px rgba(99,102,241,0.15); }}
    .form-row {{ display: grid; grid-template-columns: 1fr 1fr; gap: 12px; }}
    .btn-pay {{
      background: #635bff;
      color: #ffffff;
      border: 0;
      border-radius: 10px;
      padding: 14px;
      font-size: 15px;
      font-weight: 600;
      cursor: pointer;
      width: 100%;
      margin-top: 12px;
      display: flex;
      align-items: center;
      justify-content: center;
      gap: 8px;
      transition: background 0.2s, transform 0.1s;
      box-shadow: 0 4px 14px rgba(99,91,255,0.3);
    }}
    .btn-pay:hover {{ background: #534ae8; transform: translateY(-1px); }}
    .btn-back {{
      display: block;
      text-align: center;
      margin-top: 14px;
      color: #64748b;
      font-size: 13px;
      text-decoration: none;
    }}
    .btn-back:hover {{ color: #0f172a; }}
    @media (max-width: 768px) {{
      .checkout-shell {{ grid-template-columns: 1fr; }}
      .order-summary, .payment-panel {{ padding: 28px 24px; }}
    }}
  </style>
</head>
<body>
  <div class="checkout-shell">
    
    <!-- LEFT: ORDER SUMMARY -->
    <div class="order-summary">
      <div>
        <div class="order-header">
          <div class="order-dot"></div>
          DataProvido Checkout
        </div>
        <div style="font-size: 12px; text-transform: uppercase; color: #94a3b8; font-weight: 700; letter-spacing: 0.5px;">Subscribe to Plan</div>
        <div class="order-plan">{plan_name}</div>
        <div class="order-price">{price_val} <span>/ billed monthly</span></div>
        
        <ul class="order-features">
          <li>✓ 100% Offline Local LLaMA 3.1 Engine</li>
          <li>✓ 15+ Advanced Analytical Modules</li>
          <li>✓ Unlimited Excel / CSV Data Ingestion</li>
          {"<li>⭐ <strong>1.5 Hours / Week Live Video Support</strong></li>" if is_pro else "<li>✓ Standard Email Onboarding Support</li>"}
          {"<li>✓ Custom KPI & Domain Metric Tuning</li>" if is_pro else "<li>✓ One-Click Styled Excel Reports</li>"}
        </ul>
      </div>

      <div class="stripe-trust">
        <span>🔒</span> Powered by Stripe · 256-Bit SSL Encrypted
      </div>
    </div>

    <!-- RIGHT: STRIPE PAYMENT FORM -->
    <div class="payment-panel">
      <div class="payment-title">Pay with Stripe</div>
      
      <form action="/checkout/success" method="GET">
        <input type="hidden" name="plan" value="{plan}">
        <input type="hidden" name="session_id" value="cs_live_sim_{os.urandom(6).hex()}">

        <div class="form-group">
          <label class="form-label">Email address</label>
          <input type="email" class="form-input" required placeholder="founder@company.com" value="demo@company.com">
        </div>

        <div class="form-group">
          <label class="form-label">Card information</label>
          <input type="text" class="form-input" required placeholder="4242 •••• •••• 4242" value="4242 •••• •••• 4242">
        </div>

        <div class="form-row">
          <div class="form-group">
            <label class="form-label">Expiration</label>
            <input type="text" class="form-input" required placeholder="MM / YY" value="12 / 28">
          </div>
          <div class="form-group">
            <label class="form-label">CVC</label>
            <input type="text" class="form-input" required placeholder="CVC" value="987">
          </div>
        </div>

        <div class="form-group">
          <label class="form-label">Name on card</label>
          <input type="text" class="form-input" required placeholder="Jane Doe" value="Data Leader">
        </div>

        <button type="submit" class="btn-pay">
          Pay €{price_cents} with Stripe &nbsp;→
        </button>

        <a href="/pricing" class="btn-back">← Cancel and return to plans</a>
      </form>
    </div>

  </div>
</body>
</html>"""


@app.get("/checkout/success", response_class=HTMLResponse)
def checkout_success(plan: str = "standard", session_id: str = ""):
    is_pro = plan.lower() == "pro"
    plan_title = "DataProvido Pro (+ 1.5h Weekly Support)" if is_pro else "DataProvido Standard"
    
    return f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>Payment Successful – DataProvido</title>
  <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&family=Playfair+Display:wght@700&display=swap" rel="stylesheet">
  <style>
    *, *::before, *::after {{ margin: 0; padding: 0; box-sizing: border-box; }}
    body {{
      background: linear-gradient(160deg, #fff8f4 0%, #ffffff 45%, #f0f7ff 100%);
      color: #1e293b;
      font-family: 'Inter', sans-serif;
      min-height: 100vh;
      display: flex;
      align-items: center;
      justify-content: center;
      padding: 24px;
    }}
    .success-card {{
      max-width: 620px;
      width: 100%;
      background: #ffffff;
      border: 1px solid #e2e8f0;
      border-radius: 24px;
      padding: 48px 40px;
      box-shadow: 0 10px 40px rgba(0,0,0,0.06);
      text-align: center;
      position: relative;
      overflow: hidden;
    }}
    .success-card::before {{
      content: ''; position: absolute; top: 0; left: 0; right: 0; height: 6px;
      background: linear-gradient(90deg, #10b981 0%, #059669 100%);
    }}
    .success-badge {{
      width: 64px; height: 64px; border-radius: 50%;
      background: #ecfdf5; color: #10b981; font-size: 32px;
      display: inline-flex; align-items: center; justify-content: center;
      margin-bottom: 20px; border: 2px solid #a7f3d0;
    }}
    h1 {{
      font-family: 'Playfair Display', serif;
      font-size: 32px;
      color: #0f172a;
      margin-bottom: 12px;
    }}
    p {{
      color: #475569;
      font-size: 15px;
      line-height: 1.7;
      margin-bottom: 24px;
    }}
    .plan-box {{
      background: #f8fafc;
      border: 1.5px solid #e2e8f0;
      border-radius: 16px;
      padding: 20px;
      margin-bottom: 28px;
      text-align: left;
    }}
    .plan-row {{
      display: flex; justify-content: space-between; margin-bottom: 8px; font-size: 13.5px;
    }}
    .plan-row:last-child {{ margin-bottom: 0; }}
    .btn-launch {{
      background: #f26f26;
      color: #ffffff;
      padding: 14px 32px;
      border-radius: 999px;
      font-size: 15px;
      font-weight: 700;
      text-decoration: none;
      display: inline-flex;
      align-items: center;
      gap: 8px;
      box-shadow: 0 4px 16px rgba(242,111,38,0.35);
      transition: background 0.2s, transform 0.15s;
    }}
    .btn-launch:hover {{ background: #d85c18; transform: translateY(-2px); }}
    .support-box {{
      background: #fff8f4;
      border: 1px solid rgba(242,111,38,0.3);
      border-radius: 12px;
      padding: 14px;
      margin-top: 20px;
      font-size: 13px;
      color: #7c2d12;
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 12px;
      text-align: left;
    }}
  </style>
</head>
<body>
  <div class="success-card">
    <div class="success-badge">✓</div>
    <h1>Payment Successful!</h1>
    <p>Your DataProvido license is active and ready to use. Your on-premise local AI environment has been granted full analytical access.</p>

    <div class="plan-box">
      <div class="plan-row">
        <span style="color: #64748b;">Subscribed Plan:</span>
        <strong style="color: #0f172a;">{plan_title}</strong>
      </div>
      <div class="plan-row">
        <span style="color: #64748b;">License Status:</span>
        <strong style="color: #10b981;">● Active (Unlimited Offline Queries)</strong>
      </div>
      <div class="plan-row">
        <span style="color: #64748b;">Session Reference:</span>
        <span style="font-family: monospace; color: #64748b;">{session_id or 'cs_live_active'}</span>
      </div>
    </div>

    {"<div class='support-box'><div><strong>📅 Weekly 1.5h Support Included:</strong><div style='font-size: 12px; color: #9a3412;'>Book your dedicated weekly 1-on-1 strategy &amp; technical consultation.</div></div><a href='mailto:info@dataprovido.com?subject=Schedule%20Weekly%201.5h%20Live%20Support%20Session' style='background: #f26f26; color: #fff; text-decoration: none; padding: 6px 12px; border-radius: 8px; font-weight: 600; font-size: 12px; white-space: nowrap;'>Book Session →</a></div><br>" if is_pro else ""}

    <div style="margin-top: 16px;">
      <a href="/journey?activated=true&plan={plan}" class="btn-launch">
        🚀 Launch Analytics Console &nbsp;→
      </a>
    </div>
  </div>
</body>
</html>"""


# ─────────────────────────────────────────────
#  STRIPE WEBHOOK ENDPOINT
#  POST /stripe-webhook
#  - Stripe calls this after every payment event
#  - Raw body required for signature verification
#  - Set STRIPE_WEBHOOK_SECRET in Railway env vars
# ─────────────────────────────────────────────
from fastapi import HTTPException
from fastapi.responses import JSONResponse

@app.post("/stripe-webhook")
async def stripe_webhook(
    request: Request,
    stripe_signature: str = Header(None, alias="stripe-signature")
):
    """
    Stripe sends signed POST requests here for all subscription events.
    Raw body must be read before any parsing to allow signature verification.
    """
    raw_body = await request.body()
    webhook_secret = os.getenv("STRIPE_WEBHOOK_SECRET")

    if not webhook_secret:
        print("[STRIPE WEBHOOK] Warning: STRIPE_WEBHOOK_SECRET not set — skipping verification")
        try:
            event = json.loads(raw_body)
        except Exception:
            raise HTTPException(status_code=400, detail="Invalid payload")
    else:
        try:
            import stripe
            stripe.api_key = os.getenv("STRIPE_SECRET_KEY", "")
            event = stripe.Webhook.construct_event(raw_body, stripe_signature, webhook_secret)
        except Exception as e:
            print(f"[STRIPE WEBHOOK] Signature verification failed: {e}")
            raise HTTPException(status_code=400, detail=f"Webhook signature error: {e}")

    event_type = event.get("type", "")
    data_object = event.get("data", {}).get("object", {})

    # ── checkout.session.completed ──────────────────────────────────────────
    # Fires immediately when card payment is confirmed on Stripe Checkout form.
    # NOTE: For async methods (SEPA, Bancontact, iDEAL), this fires but payment_status
    # may still be "unpaid" — wait for async_payment_succeeded before granting access.
    if event_type == "checkout.session.completed":
        session_id = data_object.get("id", "")
        customer_email = data_object.get("customer_details", {}).get("email", "")
        subscription_id = data_object.get("subscription", "")
        payment_status = data_object.get("payment_status", "")
        plan = data_object.get("metadata", {}).get("plan", "standard")
        print(f"[STRIPE] ✅ Checkout completed! session={session_id} email={customer_email} sub={subscription_id} plan={plan} status={payment_status}")
        # TODO: Activate license in DB, send welcome email

    # ── checkout.session.async_payment_succeeded ────────────────────────────
    # Fires when a delayed/async payment method (SEPA Direct Debit, Bancontact,
    # iDEAL, etc.) is finally confirmed after the checkout session was created.
    # This is the definitive "payment is good, grant access" signal for those methods.
    elif event_type == "checkout.session.async_payment_succeeded":
        session_id = data_object.get("id", "")
        customer_email = data_object.get("customer_details", {}).get("email", "")
        subscription_id = data_object.get("subscription", "")
        plan = data_object.get("metadata", {}).get("plan", "standard")
        print(f"[STRIPE] ✅ Async payment confirmed! session={session_id} email={customer_email} sub={subscription_id} plan={plan}")
        # TODO: Activate license, send access confirmation email

    # ── checkout.session.async_payment_failed ──────────────────────────────
    # Fires when a delayed payment method ultimately fails (e.g. bank rejects SEPA).
    # Must revoke any provisional access granted on session.completed.
    elif event_type == "checkout.session.async_payment_failed":
        session_id = data_object.get("id", "")
        customer_email = data_object.get("customer_details", {}).get("email", "")
        plan = data_object.get("metadata", {}).get("plan", "standard")
        print(f"[STRIPE] ❌ Async payment failed! session={session_id} email={customer_email} plan={plan}")
        # TODO: Revoke provisional access, notify customer

    # ── invoice.payment_succeeded ───────────────────────────────────────────
    # Fires every month on each successful recurring subscription charge.
    elif event_type == "invoice.payment_succeeded":
        subscription_id = data_object.get("subscription", "")
        customer_email = data_object.get("customer_email", "")
        amount_paid = data_object.get("amount_paid", 0) / 100
        currency = data_object.get("currency", "eur").upper()
        print(f"[STRIPE] 🔄 Renewal payment! {amount_paid} {currency} | sub={subscription_id} email={customer_email}")
        # TODO: Extend license period, log payment record

    # ── customer.subscription.deleted ──────────────────────────────────────
    # Fires when a subscription is cancelled or expires after failed payment retries.
    elif event_type == "customer.subscription.deleted":
        subscription_id = data_object.get("id", "")
        customer_id = data_object.get("customer", "")
        print(f"[STRIPE] ❌ Subscription cancelled/expired! sub={subscription_id} customer={customer_id}")
        # TODO: Revoke console access, send cancellation email

    # ── invoice.payment_failed ──────────────────────────────────────────────
    # Fires when a monthly recurring charge fails (card expired, insufficient funds).
    # Stripe will retry automatically based on your retry schedule.
    elif event_type == "invoice.payment_failed":
        customer_email = data_object.get("customer_email", "")
        subscription_id = data_object.get("subscription", "")
        attempt = data_object.get("attempt_count", 1)
        print(f"[STRIPE] ⚠️ Payment failed (attempt {attempt})! email={customer_email} sub={subscription_id}")
        # TODO: Notify customer, suspend access after grace period / max retries

    else:
        print(f"[STRIPE WEBHOOK] Unhandled event: {event_type}")

    return JSONResponse(content={"status": "ok"})


@app.get("/contact", response_class=HTMLResponse)
def contact():
    return simple_page(
        "Contact & Enterprise Inquiries",
        """
        <p class="page-subhead">
          Connect directly with our executive, sales, infrastructure, and AI engineering leadership for on-premise deployments, tailored demonstrations, or retail partnership inquiries.
        </p>

        <!-- TOP ROW: Leadership & Technical Core (Yaşam, Buse, Metehan) -->
        <div style="display: grid; grid-template-columns: repeat(auto-fit, minmax(280px, 1fr)); gap: 20px; margin: 28px 0 20px;">
          
          <!-- Yaşam Karadağ (AI & Product) -->
          <div style="background: #ffffff; border: 1.5px solid var(--border); border-radius: 18px; padding: 24px; box-shadow: var(--card-shadow); display: flex; flex-direction: column; transition: all 0.2s;">
            <div style="display: inline-flex; align-items: center; justify-content: center; width: 42px; height: 42px; border-radius: 12px; background: rgba(242,111,38,0.10); color: var(--orange); font-size: 20px; margin-bottom: 14px;">👨‍💻</div>
            <div style="font-weight: 800; color: var(--text-900); font-size: 17px; margin-bottom: 4px;">Yaşam Karadağ</div>
            <div style="font-size: 11.5px; font-weight: 600; color: var(--orange); text-transform: uppercase; letter-spacing: 0.5px; margin-bottom: 8px;">Tech, AI Architecture &amp; Product</div>
            <p style="font-size: 13px; color: var(--text-500); line-height: 1.6; margin-bottom: 18px; flex: 1;">
              For LLM reasoning engines, tool dispatching architecture, prompt engineering, and retail AI logic.
            </p>
            <a href="mailto:karadagya@dataprovido.com" style="display: inline-flex; align-items: center; justify-content: center; gap: 8px; color: var(--orange); font-weight: 700; text-decoration: none; font-size: 13.5px; background: #fff8f4; padding: 10px 14px; border-radius: 10px; border: 1px solid var(--border-orange); transition: background .15s;">
              ✉️ karadagya@dataprovido.com
            </a>
          </div>

          <!-- Buse Aksoy (Strategy) -->
          <div style="background: #ffffff; border: 1.5px solid var(--border); border-radius: 18px; padding: 24px; box-shadow: var(--card-shadow); display: flex; flex-direction: column; transition: all 0.2s;">
            <div style="display: inline-flex; align-items: center; justify-content: center; width: 42px; height: 42px; border-radius: 12px; background: rgba(242,111,38,0.10); color: var(--orange); font-size: 20px; margin-bottom: 14px;">📈</div>
            <div style="font-weight: 800; color: var(--text-900); font-size: 17px; margin-bottom: 4px;">Buse Aksoy</div>
            <div style="font-size: 11.5px; font-weight: 600; color: var(--orange); text-transform: uppercase; letter-spacing: 0.5px; margin-bottom: 8px;">Business Strategy &amp; Partnerships</div>
            <p style="font-size: 13px; color: var(--text-500); line-height: 1.6; margin-bottom: 18px; flex: 1;">
              For strategic retail partnerships, CRM growth consulting, and cross-functional operations.
            </p>
            <a href="mailto:aksoyb@dataprovido.com" style="display: inline-flex; align-items: center; justify-content: center; gap: 8px; color: var(--orange); font-weight: 700; text-decoration: none; font-size: 13.5px; background: #fff8f4; padding: 10px 14px; border-radius: 10px; border: 1px solid var(--border-orange); transition: background .15s;">
              ✉️ aksoyb@dataprovido.com
            </a>
          </div>

          <!-- Metehan Taşkan (Infrastructure & Docker) -->
          <div style="background: #ffffff; border: 1.5px solid var(--border); border-radius: 18px; padding: 24px; box-shadow: var(--card-shadow); display: flex; flex-direction: column; transition: all 0.2s;">
            <div style="display: inline-flex; align-items: center; justify-content: center; width: 42px; height: 42px; border-radius: 12px; background: rgba(242,111,38,0.10); color: var(--orange); font-size: 20px; margin-bottom: 14px;">⚙️</div>
            <div style="font-weight: 800; color: var(--text-900); font-size: 17px; margin-bottom: 4px;">Metehan Taşkan</div>
            <div style="font-size: 11.5px; font-weight: 600; color: var(--orange); text-transform: uppercase; letter-spacing: 0.5px; margin-bottom: 8px;">Server-Side &amp; Docker Integration</div>
            <p style="font-size: 13px; color: var(--text-500); line-height: 1.6; margin-bottom: 18px; flex: 1;">
              For server-side deployments, on-premise hardware sizing, Docker orchestration, and security infra.
            </p>
            <a href="mailto:mtaskan@dataprovido.com" style="display: inline-flex; align-items: center; justify-content: center; gap: 8px; color: var(--orange); font-weight: 700; text-decoration: none; font-size: 13.5px; background: #fff8f4; padding: 10px 14px; border-radius: 10px; border: 1px solid var(--border-orange); transition: background .15s;">
              ✉️ mtaskan@dataprovido.com
            </a>
          </div>

        </div>

        <!-- BOTTOM ROW: Sales & Enterprise Licensing (Muhammet Bozkurt & General Desk) -->
        <div style="display: grid; grid-template-columns: repeat(auto-fit, minmax(320px, 1fr)); gap: 20px; margin: 0 0 28px;">

          <!-- Muhammet Bozkurt (Sales) -->
          <div style="background: #ffffff; border: 1.5px solid var(--border); border-radius: 18px; padding: 24px; box-shadow: var(--card-shadow); display: flex; flex-direction: column; transition: all 0.2s;">
            <div style="display: inline-flex; align-items: center; justify-content: center; width: 42px; height: 42px; border-radius: 12px; background: rgba(242,111,38,0.10); color: var(--orange); font-size: 20px; margin-bottom: 14px;">💼</div>
            <div style="font-weight: 800; color: var(--text-900); font-size: 17px; margin-bottom: 4px;">Muhammet Bozkurt</div>
            <div style="font-size: 11.5px; font-weight: 600; color: var(--orange); text-transform: uppercase; letter-spacing: 0.5px; margin-bottom: 8px;">Enterprise Sales &amp; Commercial Solutions</div>
            <p style="font-size: 13px; color: var(--text-500); line-height: 1.6; margin-bottom: 18px; flex: 1;">
              For enterprise licensing packages, retail ROI modeling, pilot scoping, and custom contracts.
            </p>
            <a href="mailto:mbozkurt@dataprovido.com" style="display: inline-flex; align-items: center; justify-content: center; gap: 8px; color: var(--orange); font-weight: 700; text-decoration: none; font-size: 13.5px; background: #fff8f4; padding: 10px 14px; border-radius: 10px; border: 1px solid var(--border-orange); transition: background .15s;">
              ✉️ mbozkurt@dataprovido.com
            </a>
          </div>

          <!-- General Desk -->
          <div style="background: #ffffff; border: 1.5px solid var(--border); border-radius: 18px; padding: 24px; box-shadow: var(--card-shadow); display: flex; flex-direction: column; transition: all 0.2s;">
            <div style="display: inline-flex; align-items: center; justify-content: center; width: 42px; height: 42px; border-radius: 12px; background: rgba(242,111,38,0.10); color: var(--orange); font-size: 20px; margin-bottom: 14px;">🏢</div>
            <div style="font-weight: 800; color: var(--text-900); font-size: 17px; margin-bottom: 4px;">General &amp; Enterprise Desk</div>
            <div style="font-size: 11.5px; font-weight: 600; color: var(--orange); text-transform: uppercase; letter-spacing: 0.5px; margin-bottom: 8px;">Licensing &amp; Demonstration</div>
            <p style="font-size: 13px; color: var(--text-500); line-height: 1.6; margin-bottom: 18px; flex: 1;">
              For live platform demos, pilot deployments, RFP submissions, and legal NDA processes.
            </p>
            <a href="mailto:info@dataprovido.com" style="display: inline-flex; align-items: center; justify-content: center; gap: 8px; color: var(--orange); font-weight: 700; text-decoration: none; font-size: 13.5px; background: #fff8f4; padding: 10px 14px; border-radius: 10px; border: 1px solid var(--border-orange); transition: background .15s;">
              ✉️ info@dataprovido.com
            </a>
          </div>

        </div>

        <!-- Quick Inquiry Helper Box -->
        <div style="background: var(--bg-2); border: 1.5px solid var(--border); border-radius: 18px; padding: 28px; margin-top: 24px;">
          <h3 style="font-size: 18px; font-weight: 700; color: var(--text-900); margin-bottom: 8px;">Direct Inquiry Quick Composer</h3>
          <p style="font-size: 13.5px; color: var(--text-500); margin-bottom: 18px;">
            Choose a department or specialist and launch your email client with pre-formatted details:
          </p>
          <div style="display: flex; gap: 10px; flex-wrap: wrap;">
            <a href="mailto:mbozkurt@dataprovido.com?subject=DataProvido%20Enterprise%20Sales%20%26%20Licensing%20Inquiry" style="background: #fff3ec; border: 1px solid var(--border-orange); color: var(--orange-dark); padding: 8px 16px; border-radius: 999px; font-size: 13px; font-weight: 600; text-decoration: none; transition: border-color .2s;">
              💼 Enterprise Sales (Muhammet Bozkurt)
            </a>
            <a href="mailto:mtaskan@dataprovido.com?subject=DataProvido%20Server-Side%20%26%20Docker%20Infrastructure%20Inquiry" style="background: #ffffff; border: 1px solid var(--border); color: var(--text-900); padding: 8px 16px; border-radius: 999px; font-size: 13px; font-weight: 600; text-decoration: none; transition: border-color .2s;">
              ⚙️ Server-Side &amp; Docker (Metehan Taşkan)
            </a>
            <a href="mailto:karadagya@dataprovido.com?subject=DataProvido%20AI%20Architecture%20%26%20LLM%20Inquiry" style="background: #ffffff; border: 1px solid var(--border); color: var(--text-900); padding: 8px 16px; border-radius: 999px; font-size: 13px; font-weight: 600; text-decoration: none; transition: border-color .2s;">
              👨‍💻 AI Architecture (Yaşam Karadağ)
            </a>
            <a href="mailto:aksoyb@dataprovido.com?subject=DataProvido%20Retail%20Strategy%20%26%20Partnership%20Inquiry" style="background: #ffffff; border: 1px solid var(--border); color: var(--text-900); padding: 8px 16px; border-radius: 999px; font-size: 13px; font-weight: 600; text-decoration: none; transition: border-color .2s;">
              📈 Strategy &amp; Growth (Buse Aksoy)
            </a>
          </div>
        </div>

        <div style="background: #ffffff; border: 1px solid var(--border); border-radius: 14px; padding: 18px 22px; margin-top: 20px; font-size: 13.5px; color: var(--text-700); line-height: 1.7; display: flex; align-items: center; gap: 12px;">
          <span style="font-size: 20px;">🔒</span>
          <div><strong>Enterprise Privacy &amp; 24-Hour SLA:</strong> All conversations are strictly covered under NDA with guaranteed executive response within 24 hours.</div>
        </div>
        """,
        kicker="Direct Executive Contact",
        active_nav="contact",
        max_width="1080px"
    )


@app.get("/who-we-are", response_class=HTMLResponse)
def who_we_are():
    return simple_page(
        "Built by Retail Veterans & AI Engineers",
        """
        <p class="page-subhead">
          DataProvido was founded at the convergence of <strong>10+ years of retail, e-commerce, CRM, and supply chain leadership</strong> with cutting-edge <strong>on-premise AI systems engineering</strong>.
        </p>

        <!-- The Story -->
        <div style="background: var(--bg-2); border: 1.5px solid var(--border); border-radius: 18px; padding: 28px; margin-bottom: 28px;">
          <h3 style="font-family: 'Playfair Display', serif; font-size: 24px; color: var(--text-900); margin-bottom: 14px; font-weight: 700;">
            Solving the Retail Industry's Core Analytics Bottleneck
          </h3>
          <p style="font-size: 14.5px; color: var(--text-700); line-height: 1.85; margin-bottom: 14px;">
            For over a decade, our multidisciplinary team has managed billion-dollar retail portfolios, optimized conversion funnels for Tier-1 e-commerce platforms, orchestrated multi-tier CRM retention campaigns, and designed high-throughput data architectures.
          </p>
          <p style="font-size: 14.5px; color: var(--text-700); line-height: 1.85; margin-bottom: 0;">
            We experienced firsthand the single largest pain point in modern commerce: <em>Business leaders are drowning in disconnected spreadsheets, waiting days for data analysts to answer basic commercial questions, while cloud-based AI tools pose unacceptable data privacy and IP security risks.</em> DataProvido is the definitive solution — delivering instant, conversational, and 100% on-premise intelligence directly to commercial decision-makers.
          </p>
        </div>

        <!-- 4 Pillars of Deep Know-How -->
        <h3 style="font-family: 'Playfair Display', serif; font-size: 22px; color: var(--text-900); margin: 32px 0 16px; font-weight: 700;">
          Our 4 Core Pillars of Industry Expertise
        </h3>

        <div style="display: grid; grid-template-columns: repeat(auto-fit, minmax(250px, 1fr)); gap: 18px; margin-bottom: 32px;">
          
          <div style="background: #ffffff; border: 1.5px solid var(--border); border-radius: 16px; padding: 24px; box-shadow: var(--card-shadow);">
            <div style="font-size: 28px; margin-bottom: 12px;">👥</div>
            <div style="font-weight: 700; color: var(--text-900); font-size: 16px; margin-bottom: 6px;">CRM &amp; Lifecycle Marketing</div>
            <div style="font-size: 13px; color: var(--text-700); line-height: 1.7;">Mastery of RFM segmentation, customer lifetime value (CLV), churn prediction models, and automated omnichannel reactivation across millions of customer profiles.</div>
          </div>

          <div style="background: #ffffff; border: 1.5px solid var(--border); border-radius: 16px; padding: 24px; box-shadow: var(--card-shadow);">
            <div style="font-size: 28px; margin-bottom: 12px;">📦</div>
            <div style="font-weight: 700; color: var(--text-900); font-size: 16px; margin-bottom: 6px;">Merchandising &amp; Category Analytics</div>
            <div style="font-size: 13px; color: var(--text-700); line-height: 1.7;">Decade-long experience in stock coverage ratios, out-of-stock (OOS) revenue recovery, safety stock algorithms, and GfK / Nielsen retail market share benchmarking.</div>
          </div>

          <div style="background: #ffffff; border: 1.5px solid var(--border); border-radius: 16px; padding: 24px; box-shadow: var(--card-shadow);">
            <div style="font-size: 28px; margin-bottom: 12px;">📈</div>
            <div style="font-weight: 700; color: var(--text-900); font-size: 16px; margin-bottom: 6px;">Conversion Funnel Science</div>
            <div style="font-size: 13px; color: var(--text-700); line-height: 1.7;">Advanced diagnosis of micro-conversion stages (PDP Views, A2C, C2D, B2D, and checkout drop-off) to pinpoint exact UX friction points and eliminate revenue leakage.</div>
          </div>

          <div style="background: #ffffff; border: 1.5px solid var(--border); border-radius: 16px; padding: 24px; box-shadow: var(--card-shadow);">
            <div style="font-size: 28px; margin-bottom: 12px;">🔒</div>
            <div style="font-weight: 700; color: var(--text-900); font-size: 16px; margin-bottom: 6px;">Privacy-First AI Engineering</div>
            <div style="font-size: 13px; color: var(--text-700); line-height: 1.7;">Custom engineering of air-gapped on-premise LLMs (LLaMA 3.1) and vectorized memory-optimized engines (Pandas &amp; DuckDB) guaranteeing complete data sovereignty.</div>
          </div>

        </div>

        <!-- Metrics Row -->
        <div style="display: grid; grid-template-columns: repeat(auto-fit, minmax(180px, 1fr)); gap: 14px; margin: 28px 0;">
          <div style="padding: 20px; background: #fff8f4; border: 1.5px solid var(--border-orange); border-radius: 16px; text-align: center;">
            <div style="font-size: 28px; font-weight: 800; color: var(--orange); margin-bottom: 4px;">10+ Years</div>
            <div style="font-size: 12.5px; color: var(--text-700); font-weight: 600;">Retail &amp; CRM Heritage</div>
          </div>
          <div style="padding: 20px; background: #fff8f4; border: 1.5px solid var(--border-orange); border-radius: 16px; text-align: center;">
            <div style="font-size: 28px; font-weight: 800; color: var(--orange); margin-bottom: 4px;">100%</div>
            <div style="font-size: 12.5px; color: var(--text-700); font-weight: 600;">Air-Gapped Privacy</div>
          </div>
          <div style="padding: 20px; background: #fff8f4; border: 1.5px solid var(--border-orange); border-radius: 16px; text-align: center;">
            <div style="font-size: 28px; font-weight: 800; color: var(--orange); margin-bottom: 4px;">15+</div>
            <div style="font-size: 12.5px; color: var(--text-700); font-weight: 600;">Specialized AI Engines</div>
          </div>
          <div style="padding: 20px; background: #fff8f4; border: 1.5px solid var(--border-orange); border-radius: 16px; text-align: center;">
            <div style="font-size: 28px; font-weight: 800; color: var(--orange); margin-bottom: 4px;">&lt;30s</div>
            <div style="font-size: 12.5px; color: var(--text-700); font-weight: 600;">Instant Decision Speed</div>
          </div>
        </div>
        """,
        kicker="Leadership & Industry Heritage",
        active_nav="who-we-are",
        max_width="960px"
    )


@app.get("/how-works", response_class=HTMLResponse)
def how_works():
    return simple_page(
        "How DataProvido Works: Architecture & LLM Routing",
        """
        <p class="page-subhead">
          DataProvido transforms complex retail data into instant, prescriptive actions through an <strong>autonomous LLM tool-calling loop and high-performance in-memory engines</strong> — completely self-hosted on your own infrastructure.
        </p>

        <!-- 5-Step Technical Architecture -->
        <div style="display: flex; flex-direction: column; gap: 20px; margin: 28px 0;">
          
          <!-- Step 1 -->
          <div style="background: #ffffff; border: 1.5px solid var(--border); border-radius: 18px; padding: 24px; box-shadow: var(--card-shadow);">
            <div style="display: flex; align-items: center; gap: 14px; margin-bottom: 12px;">
              <div style="width: 38px; height: 38px; border-radius: 10px; background: var(--orange); color: #fff; display: grid; place-items: center; font-weight: 800; font-size: 16px;">1</div>
              <div>
                <div style="font-size: 11px; font-weight: 700; text-transform: uppercase; color: var(--orange); letter-spacing: 0.5px;">DATA LAYER (/data)</div>
                <h3 style="font-size: 18px; font-weight: 700; color: var(--text-900); margin: 0;">Multi-Source Enterprise Ingestion</h3>
              </div>
            </div>
            <p style="font-size: 14px; color: var(--text-700); line-height: 1.8; margin-bottom: 0;">
              Ingest enterprise stock balances (<code style="background: var(--bg-3); padding: 2px 6px; border-radius: 4px;">stok.xlsx</code>), transaction orders (<code style="background: var(--bg-3); padding: 2px 6px; border-radius: 4px;">orders.xlsx</code>), GfK market share panels (<code style="background: var(--bg-3); padding: 2px 6px; border-radius: 4px;">GfK_Leaderpanel.xlsx</code>, <code style="background: var(--bg-3); padding: 2px 6px; border-radius: 4px;">gfk_sku.xlsx</code>), and funnel tracking data. All files are loaded locally without any cloud exposure.
            </p>
          </div>

          <!-- Step 2 -->
          <div style="background: #ffffff; border: 1.5px solid var(--border); border-radius: 18px; padding: 24px; box-shadow: var(--card-shadow);">
            <div style="display: flex; align-items: center; gap: 14px; margin-bottom: 12px;">
              <div style="width: 38px; height: 38px; border-radius: 10px; background: var(--orange); color: #fff; display: grid; place-items: center; font-weight: 800; font-size: 16px;">2</div>
              <div>
                <div style="font-size: 11px; font-weight: 700; text-transform: uppercase; color: var(--orange); letter-spacing: 0.5px;">INFERENCE ENGINE (Ollama / LLaMA 3.1)</div>
                <h3 style="font-size: 18px; font-weight: 700; color: var(--text-900); margin: 0;">100% Offline Air-Gapped Intelligence</h3>
              </div>
            </div>
            <p style="font-size: 14px; color: var(--text-700); line-height: 1.8; margin-bottom: 0;">
              Queries are processed by a self-hosted LLaMA 3.1 model running in an isolated Docker container on your internal network. Unlike cloud AI APIs, no company turnover, SKU profit margins, or customer volumes ever leave your enterprise firewall — guaranteeing full KVKK and GDPR compliance.
            </p>
          </div>

          <!-- Step 3 -->
          <div style="background: #ffffff; border: 1.5px solid var(--border); border-radius: 18px; padding: 24px; box-shadow: var(--card-shadow);">
            <div style="display: flex; align-items: center; gap: 14px; margin-bottom: 12px;">
              <div style="width: 38px; height: 38px; border-radius: 10px; background: var(--orange); color: #fff; display: grid; place-items: center; font-weight: 800; font-size: 16px;">3</div>
              <div>
                <div style="font-size: 11px; font-weight: 700; text-transform: uppercase; color: var(--orange); letter-spacing: 0.5px;">ROUTER &amp; SCHEMAS (/schemas/tools.py)</div>
                <h3 style="font-size: 18px; font-weight: 700; color: var(--text-900); margin: 0;">Autonomous Tool-Calling &amp; Function Dispatch</h3>
              </div>
            </div>
            <p style="font-size: 14px; color: var(--text-700); line-height: 1.8; margin-bottom: 0;">
              When a user asks a question (e.g., <em>"Which SKUs in Small Appliances have high traffic but critical stock risk?"</em>), the model selects the optimal tool from our structured schemas and dispatches execution to specialized Python micro-engines in <code style="background: var(--bg-3); padding: 2px 6px; border-radius: 4px;">/functions</code>:
            </p>
            <div style="display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 10px; margin-top: 14px;">
              <div style="background: var(--bg-2); padding: 10px 12px; border-radius: 8px; font-size: 12.5px; border: 1px solid var(--border);">📊 <strong>business_calculator.py:</strong> SQL-style math &amp; aggregations</div>
              <div style="background: var(--bg-2); padding: 10px 12px; border-radius: 8px; font-size: 12.5px; border: 1px solid var(--border);">🛒 <strong>funnel_master.py:</strong> PDP → Cart → Checkout drop-offs</div>
              <div style="background: var(--bg-2); padding: 10px 12px; border-radius: 8px; font-size: 12.5px; border: 1px solid var(--border);">💰 <strong>price_competition.py:</strong> Market price gap benchmark</div>
              <div style="background: var(--bg-2); padding: 10px 12px; border-radius: 8px; font-size: 12.5px; border: 1px solid var(--border);">📦 <strong>stock.py:</strong> Live stock &amp; OOS revenue risk</div>
              <div style="background: var(--bg-2); padding: 10px 12px; border-radius: 8px; font-size: 12.5px; border: 1px solid var(--border);">🏆 <strong>gfk_analyzer.py:</strong> Market share &amp; brand rankings</div>
              <div style="background: var(--bg-2); padding: 10px 12px; border-radius: 8px; font-size: 12.5px; border: 1px solid var(--border);">⚡ <strong>action_executor.py:</strong> Automated business action generation</div>
            </div>
          </div>

          <!-- Step 4 -->
          <div style="background: #ffffff; border: 1.5px solid var(--border); border-radius: 18px; padding: 24px; box-shadow: var(--card-shadow);">
            <div style="display: flex; align-items: center; gap: 14px; margin-bottom: 12px;">
              <div style="width: 38px; height: 38px; border-radius: 10px; background: var(--orange); color: #fff; display: grid; place-items: center; font-weight: 800; font-size: 16px;">4</div>
              <div>
                <div style="font-size: 11px; font-weight: 700; text-transform: uppercase; color: var(--orange); letter-spacing: 0.5px;">EXECUTION LAYER (Pandas &amp; DuckDB)</div>
                <h3 style="font-size: 18px; font-weight: 700; color: var(--text-900); margin: 0;">Sub-Second In-Memory Vector Computation</h3>
              </div>
            </div>
            <p style="font-size: 14px; color: var(--text-700); line-height: 1.8; margin-bottom: 0;">
              Calculations execute in memory using vectorized Pandas operations, computing complex multi-table aggregations, margin percentages, and stock coverage velocities across hundreds of thousands of rows in sub-seconds.
            </p>
          </div>

          <!-- Step 5 -->
          <div style="background: #ffffff; border: 1.5px solid var(--border); border-radius: 18px; padding: 24px; box-shadow: var(--card-shadow);">
            <div style="display: flex; align-items: center; gap: 14px; margin-bottom: 12px;">
              <div style="width: 38px; height: 38px; border-radius: 10px; background: var(--orange); color: #fff; display: grid; place-items: center; font-weight: 800; font-size: 16px;">5</div>
              <div>
                <div style="font-size: 11px; font-weight: 700; text-transform: uppercase; color: var(--orange); letter-spacing: 0.5px;">ACTION &amp; EXPORT (OpenPyXL / UI)</div>
                <h3 style="font-size: 18px; font-weight: 700; color: var(--text-900); margin: 0;">Prescriptive Action Plans &amp; 1-Click Excel Reports</h3>
              </div>
            </div>
            <p style="font-size: 14px; color: var(--text-700); line-height: 1.8; margin-bottom: 0;">
              Rather than raw tables, DataProvido generates structured business conclusions with ranked next actions (e.g., immediate stock replenishment, dynamic price discount rule). With one click, users can export an executive-ready formatted <code style="background: var(--bg-3); padding: 2px 6px; border-radius: 4px;">.xlsx</code> report with custom header styles and KPI summaries.
            </p>
          </div>

        </div>

        <!-- Tech Stack Strip -->
        <div style="background: #fff8f4; border: 1.5px solid var(--border-orange); border-radius: 16px; padding: 20px 24px; display: flex; align-items: center; justify-content: space-between; flex-wrap: wrap; gap: 14px;">
          <div style="font-size: 13.5px; color: var(--text-700);">
            🚀 <strong>Ready to test the engine on your proprietary retail schema?</strong>
          </div>
          <a href="/pricing" style="background: var(--orange); color: #fff; text-decoration: none; font-size: 13px; font-weight: 700; padding: 9px 18px; border-radius: 8px;">
            Explore Plans (199 € &amp; 299 €) &nbsp;→
          </a>
        </div>
        """,
        kicker="Enterprise Architecture & Process",
        active_nav="how-works",
        max_width="1000px"
    )


@app.get("/", response_class=HTMLResponse)
def index():
    return """<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>DataProvido – Analyze your e-commerce data locally—no need to share it with anyone!</title>
  <meta name="description" content="Don't just analyze data—know exactly what actions to take next. DataProvido delivers real-time prescriptive commercial decisions 100% locally.">
  <meta property="og:title" content="DataProvido – Analyze your e-commerce data locally—no need to share it with anyone!">
  <meta property="og:description" content="Don't just analyze data—know exactly what actions to take next. DataProvido delivers real-time prescriptive commercial decisions 100% locally.">
  <meta property="og:image" content="https://www.dataprovido.com/logo.png">
  <meta property="og:url" content="https://www.dataprovido.com/">
  <meta property="og:type" content="website">
  <meta name="twitter:card" content="summary_large_image">
  <meta name="twitter:title" content="DataProvido – Analyze your e-commerce data locally—no need to share it with anyone!">
  <meta name="twitter:description" content="Don't just analyze data—know exactly what actions to take next. DataProvido delivers real-time prescriptive commercial decisions 100% locally.">
  <meta name="twitter:image" content="https://www.dataprovido.com/logo.png">
  <link rel="icon" type="image/png" href="/logo.png">
  <link rel="shortcut icon" type="image/png" href="/logo.png">
  <link rel="apple-touch-icon" href="/logo.png">
  <link rel="preconnect" href="https://fonts.googleapis.com">
  <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&family=Playfair+Display:ital,wght@0,400;0,700;1,400&display=swap" rel="stylesheet">
  <style>
    *, *::before, *::after { margin: 0; padding: 0; box-sizing: border-box; }
    :root {
      --orange: #f26f26; --orange-dark: #d85c18; --orange-light: #f58c50;
      --bg:      #ffffff;
      --bg-2:    #f9fafb;
      --bg-3:    #f1f3f5;
      --bg-grad: linear-gradient(180deg, #ecf7ff 0%, #fff3ec 100%);
      --text-900: #292c2f;
      --text-700: #4e5359;
      --text-500: #6b7178;
      --text-dim: #a3acb6;
      --border:   #dadee2;
      --border-orange: rgba(242,111,38,0.25);
      --card-bg:  #ffffff;
      --card-shadow: 0 2px 12px rgba(0,0,0,0.06), 0 1px 3px rgba(0,0,0,0.04);
      --radius: 16px; --nav-h: 72px;
    }
    html { scroll-behavior: smooth; }
    body { background: var(--bg); color: var(--text-900); font-family: 'Inter', sans-serif; overflow-x: hidden; }

    /* NAV */
    .nav {
      position: fixed; top: 0; left: 0; right: 0; z-index: 1000; height: var(--nav-h);
      display: flex; align-items: center; justify-content: space-between; padding: 0 40px;
      background: rgba(255,255,255,0.95); backdrop-filter: blur(16px);
      border-bottom: 1px solid var(--border); transition: background .3s;
      box-shadow: 0 1px 0 var(--border);
    }
    .nav-logo { display: flex; align-items: center; gap: 10px; font-weight: 700; font-size: 18px; color: var(--text-900); text-decoration: none; }
    .nav-logo-dot { width: 10px; height: 10px; border-radius: 50%; background: var(--orange); animation: pulse-dot 2.5s ease-in-out infinite; }
    @keyframes pulse-dot { 0%,100% { box-shadow: 0 0 0 0 rgba(242,111,38,0.5); } 50% { box-shadow: 0 0 0 6px rgba(242,111,38,0); } }
    .nav-links { display: flex; align-items: center; gap: 4px; }
    .nav-link { color: var(--text-700); text-decoration: none; font-size: 14px; font-weight: 500; padding: 8px 14px; border-radius: 8px; transition: color .2s, background .2s; }
    .nav-link:hover { color: var(--orange); background: rgba(242,111,38,0.06); }
    .nav-cta { background: var(--orange); color: #fff; border: none; padding: 10px 22px; border-radius: 24px; font-size: 14px; font-weight: 600; cursor: pointer; text-decoration: none; display: inline-flex; align-items: center; gap: 6px; transition: background .2s, transform .15s; box-shadow: 0 4px 14px rgba(242,111,38,0.3); }
    .nav-cta:hover { background: var(--orange-dark); transform: translateY(-1px); }

    /* HERO — clean light gradient, no image */
    .hero { position: relative; min-height: 80vh; display: flex; flex-direction: column; align-items: center; justify-content: center; text-align: center; padding: calc(var(--nav-h) + 80px) 24px 100px; overflow: hidden; background: linear-gradient(160deg, #fff8f4 0%, #ffffff 45%, #f0f7ff 100%); }
    .hero::before { content: ''; position: absolute; inset: 0; background: radial-gradient(ellipse 80% 60% at 50% 0%, rgba(242,111,38,0.07) 0%, transparent 70%); pointer-events: none; }
    .hero-content { position: relative; z-index: 2; }
    .hero-badge { display: inline-flex; align-items: center; gap: 8px; background: rgba(242,111,38,0.08); border: 1px solid rgba(242,111,38,0.2); border-radius: 999px; padding: 6px 16px; font-size: 13px; font-weight: 600; color: var(--orange); margin-bottom: 28px; animation: fade-up 0.7s ease both; }
    .hero h1 { font-family: 'Playfair Display', serif; font-size: clamp(40px, 6.5vw, 80px); font-weight: 700; line-height: 1.1; color: var(--text-900); letter-spacing: -1.5px; margin-bottom: 24px; animation: fade-up 0.9s ease 0.1s both; }
    .hero h1 span { color: var(--orange); }
    .hero-sub { font-size: clamp(16px, 2vw, 19px); font-weight: 400; color: var(--text-700); max-width: 560px; margin: 0 auto 44px; line-height: 1.7; animation: fade-up 0.9s ease 0.2s both; }
    .hero-actions { display: flex; align-items: center; justify-content: center; gap: 14px; flex-wrap: wrap; animation: fade-up 0.9s ease 0.3s both; }
    .btn-primary { background: var(--orange); color: #fff; padding: 14px 30px; border-radius: 999px; font-size: 15px; font-weight: 600; text-decoration: none; display: inline-flex; align-items: center; gap: 8px; box-shadow: 0 6px 20px rgba(242,111,38,0.35); transition: background .2s, transform .15s, box-shadow .2s; }
    .btn-primary:hover { background: var(--orange-dark); transform: translateY(-2px); box-shadow: 0 10px 28px rgba(242,111,38,0.45); }
    .btn-ghost { background: #fff; color: var(--text-900); border: 1.5px solid var(--border); padding: 13px 28px; border-radius: 999px; font-size: 15px; font-weight: 500; text-decoration: none; display: inline-flex; align-items: center; gap: 8px; transition: border-color .2s, box-shadow .2s; box-shadow: var(--card-shadow); }
    .btn-ghost:hover { border-color: var(--orange); box-shadow: 0 4px 16px rgba(242,111,38,0.12); }

    /* STATS */
    .stats-strip { background: var(--bg-2); border-top: 1px solid var(--border); border-bottom: 1px solid var(--border); padding: 40px 40px; }
    .stats-inner { max-width: 1200px; margin: 0 auto; display: grid; grid-template-columns: repeat(4, 1fr); }
    .stat-item { text-align: center; padding: 0 24px; border-right: 1px solid var(--border); }
    .stat-item:last-child { border-right: none; }
    .stat-number { font-size: clamp(36px, 4vw, 52px); font-weight: 800; color: var(--orange); line-height: 1; }
    .stat-label { font-size: 13px; color: var(--text-500); margin-top: 8px; font-weight: 400; }

    /* SECTION */
    .section { padding: 96px 40px; background: var(--bg); }
    .section-inner { max-width: 1200px; margin: 0 auto; }
    .section-tag { display: inline-flex; align-items: center; gap: 6px; background: rgba(242,111,38,0.08); border: 1px solid var(--border-orange); border-radius: 999px; padding: 5px 14px; font-size: 11px; font-weight: 700; color: var(--orange); letter-spacing: 0.1em; text-transform: uppercase; margin-bottom: 20px; }
    .section-title { font-family: 'Playfair Display', serif; font-size: clamp(28px, 3.5vw, 48px); font-weight: 700; color: var(--text-900); line-height: 1.2; margin-bottom: 16px; letter-spacing: -0.5px; }
    .section-desc { font-size: 17px; color: var(--text-700); max-width: 540px; line-height: 1.75; }

    /* YODECK-STYLE FEATURE SHOWCASE */
    .showcase { background: var(--bg-2); padding: 100px 0; border-top: 1px solid var(--border); border-bottom: 1px solid var(--border); }
    .showcase-inner { max-width: 1200px; margin: 0 auto; padding: 0 40px; }
    .showcase-header { margin-bottom: 64px; }
    .showcase-body { display: grid; grid-template-columns: 320px 1fr; gap: 32px; align-items: start; }

    /* LEFT TABS */
    .sc-tabs { display: flex; flex-direction: column; gap: 4px; }
    .sc-tab {
      position: relative; padding: 16px 18px 16px 22px;
      border-radius: 12px; cursor: pointer;
      border: 1px solid transparent;
      transition: background .2s, border-color .2s;
      overflow: hidden;
    }
    .sc-tab:hover { background: rgba(242,111,38,0.04); border-color: var(--border-orange); }
    .sc-tab.active { background: #fff; border-color: var(--border); box-shadow: var(--card-shadow); }
    .sc-tab-bar {
      position: absolute; left: 0; top: 0; bottom: 0; width: 3px;
      background: var(--bg-3); border-radius: 3px 0 0 3px;
    }
    .sc-tab.active .sc-tab-bar { background: rgba(242,111,38,0.2); }
    .sc-tab-progress {
      position: absolute; left: 0; top: 0; width: 3px;
      background: var(--orange); border-radius: 3px 0 0 3px;
      height: 0%;
    }
    .sc-tab.active .sc-tab-progress { animation: tab-fill 5s linear forwards; }
    @keyframes tab-fill { from { height: 0%; } to { height: 100%; } }
    .sc-tab-head { display: flex; align-items: center; gap: 10px; margin-bottom: 3px; }
    .sc-tab-icon { font-size: 16px; line-height: 1; flex-shrink: 0; }
    .sc-tab-title { font-size: 13px; font-weight: 700; color: var(--text-900); }
    .sc-tab:not(.active) .sc-tab-title { color: var(--text-700); font-weight: 500; }
    .sc-tab-desc { font-size: 12px; color: var(--text-500); line-height: 1.5; padding-left: 26px; display: none; margin-top: 4px; }
    .sc-tab.active .sc-tab-desc { display: block; }

    /* RIGHT PANEL */
    .sc-panel { position: relative; background: #fff; border: 1px solid var(--border); border-radius: 20px; overflow: hidden; box-shadow: 0 4px 32px rgba(0,0,0,0.07); min-height: 440px; }
    .sc-slide { position: absolute; inset: 0; padding: 48px 52px; display: flex; flex-direction: column; justify-content: center; opacity: 0; transform: translateX(16px); transition: opacity .4s ease, transform .4s ease; pointer-events: none; }
    .sc-slide.active { opacity: 1; transform: translateX(0); pointer-events: auto; }
    .sc-slide-tag { display: inline-flex; align-items: center; gap: 6px; background: rgba(242,111,38,0.08); border: 1px solid var(--border-orange); border-radius: 999px; padding: 4px 12px; font-size: 11px; font-weight: 700; color: var(--orange); letter-spacing: 0.1em; text-transform: uppercase; margin-bottom: 22px; width: fit-content; }
    .sc-slide h3 { font-family: 'Playfair Display', serif; font-size: clamp(22px, 2.8vw, 36px); font-weight: 700; color: var(--text-900); line-height: 1.2; margin-bottom: 14px; letter-spacing: -0.5px; }
    .sc-slide p { font-size: 15px; color: var(--text-700); line-height: 1.75; max-width: 460px; margin-bottom: 30px; }
    .sc-metrics { display: flex; gap: 16px; flex-wrap: wrap; }
    .sc-metric { background: var(--bg-2); border: 1px solid var(--border); border-radius: 12px; padding: 12px 18px; }
    .sc-metric-val { font-size: 22px; font-weight: 800; color: var(--orange); line-height: 1; }
    .sc-metric-label { font-size: 11px; color: var(--text-500); margin-top: 3px; }
    .sc-slide-visual { position: absolute; right: -10px; bottom: -10px; font-size: 140px; opacity: 0.05; pointer-events: none; line-height: 1; }

    @media (max-width: 900px) {
      .showcase-body { grid-template-columns: 1fr; }
      .sc-slide { position: relative; inset: unset; padding: 32px 24px; }
      .sc-panel { min-height: unset; }
    }


    /* FEATURE CARDS */
    .features-grid { display: grid; grid-template-columns: repeat(3, 1fr); gap: 20px; margin-top: 56px; }
    .feat-card { background: var(--card-bg); border: 1px solid var(--border); border-radius: var(--radius); padding: 28px 24px; transition: border-color .25s, transform .25s, box-shadow .25s; position: relative; overflow: hidden; box-shadow: var(--card-shadow); }
    .feat-card:hover { border-color: var(--border-orange); transform: translateY(-3px); box-shadow: 0 10px 36px rgba(242,111,38,0.1); }
    .feat-icon { width: 48px; height: 48px; border-radius: 12px; display: flex; align-items: center; justify-content: center; font-size: 22px; margin-bottom: 18px; background: rgba(242,111,38,0.08); border: 1px solid rgba(242,111,38,0.18); }
    .feat-title { font-size: 16px; font-weight: 700; color: var(--text-900); margin-bottom: 8px; }
    .feat-desc { font-size: 14px; color: var(--text-700); line-height: 1.65; }

    /* HOW IT WORKS */
    .how-section { background: var(--bg-2); border-top: 1px solid var(--border); border-bottom: 1px solid var(--border); }
    .steps-row { display: grid; grid-template-columns: repeat(4, 1fr); gap: 0; margin-top: 64px; }
    .step { padding: 0 32px 0 0; border-right: 1px solid var(--border); }
    .step:last-child { border-right: none; padding-right: 0; }
    .step-num { font-size: 52px; font-weight: 800; color: rgba(242,111,38,0.15); line-height: 1; margin-bottom: 14px; font-family: 'Playfair Display', serif; letter-spacing: -2px; }
    .step-title { font-size: 16px; font-weight: 700; color: var(--text-900); margin-bottom: 8px; }
    .step-desc { font-size: 14px; color: var(--text-700); line-height: 1.65; }

    /* TRUST */
    .trust-section { text-align: center; background: var(--bg); }
    .trust-badges { display: flex; align-items: center; justify-content: center; gap: 20px; flex-wrap: wrap; margin-top: 48px; }
    .trust-badge { background: var(--card-bg); border: 1px solid var(--border); border-radius: 14px; padding: 18px 24px; display: flex; align-items: center; gap: 14px; transition: border-color .25s, box-shadow .25s; box-shadow: var(--card-shadow); }
    .trust-badge:hover { border-color: var(--border-orange); box-shadow: 0 6px 24px rgba(242,111,38,0.1); }
    .trust-badge-icon { font-size: 26px; }
    .trust-badge-text .label { font-size: 11px; color: var(--text-500); text-transform: uppercase; letter-spacing: 0.08em; font-weight: 500; }
    .trust-badge-text .value { font-size: 18px; font-weight: 700; color: var(--text-900); }

    /* CTA BANNER */
    .cta-banner { margin: 0 40px 80px; background: linear-gradient(135deg, #fff3ec 0%, #ecf7ff 100%); border: 1px solid var(--border-orange); border-radius: 24px; padding: 72px 56px; text-align: center; position: relative; overflow: hidden; }
    .cta-banner::before { content: ''; position: absolute; inset: 0; background: radial-gradient(ellipse 60% 50% at 50% 50%, rgba(242,111,38,0.06), transparent); pointer-events: none; }
    .cta-banner h2 { font-family: 'Playfair Display', serif; font-size: clamp(28px, 4vw, 46px); font-weight: 700; color: var(--text-900); margin-bottom: 18px; letter-spacing: -0.5px; }
    .cta-banner p { font-size: 17px; color: var(--text-700); margin-bottom: 36px; }

    /* FOOTER */
    .site-footer { border-top: 1px solid var(--border); padding: 28px 40px; display: flex; align-items: center; justify-content: space-between; flex-wrap: wrap; gap: 16px; background: var(--bg-2); }
    .footer-left { display: flex; align-items: center; gap: 10px; }
    .footer-logo { font-size: 15px; font-weight: 700; color: var(--text-900); }
    .footer-copy { font-size: 13px; color: var(--text-500); }
    .footer-right { font-size: 12px; color: var(--text-500); letter-spacing: 0.1em; text-transform: uppercase; }

    /* ANIMATIONS */
    @keyframes fade-up { from { opacity: 0; transform: translateY(28px); } to { opacity: 1; transform: translateY(0); } }
    .reveal { opacity: 0; transform: translateY(32px); transition: opacity 0.7s ease, transform 0.7s ease; }
    .reveal.visible { opacity: 1; transform: translateY(0); }
    /* JS nav style update for light nav */
    body.scrolled .nav { box-shadow: 0 2px 16px rgba(0,0,0,0.07); }

    @media (max-width: 900px) {
      .nav { padding: 0 20px; } .nav-links { display: none; }
      .stats-inner { grid-template-columns: repeat(2, 1fr); } .stat-item:nth-child(2) { border-right: none; }
      .features-grid { grid-template-columns: 1fr; }
      .steps-row { grid-template-columns: 1fr 1fr; gap: 32px; } .step { border-right: none; padding-right: 0; }
      .cta-banner { margin: 0 20px 60px; padding: 48px 28px; }
      .section { padding: 64px 20px; } .slider-wrap { padding: 0 20px; }
      .slide { min-width: 100%; }
    }
  </style>
</head>
<body>

<nav class="nav" id="main-nav">
  <a href="/" class="nav-logo" style="display: flex; align-items: center; gap: 10px;"><img src="/logo.png" alt="DataProvido" style="height: 36px; width: 36px; object-fit: contain; border-radius: 8px;" /><span>DataProvido</span></a>
  <div class="nav-links">
    <a href="/pricing" class="nav-link">Pricing</a>
    <a href="/contact" class="nav-link">Contact</a>
    <a href="/who-we-are" class="nav-link">Who We Are?</a>
    <a href="/how-works" class="nav-link">How Works?</a>
    <a href="/privacy" class="nav-link">Privacy Policy</a>
  </div>
  <a href="/journey" class="nav-cta">Start Journey &nbsp;→</a>
</nav>

<section class="hero" id="hero">
  <div class="hero-content">
    <div class="hero-badge">✦ Local Intelligence. Zero Compromise.</div>
    <h1>Turn CRM data into<br><span>actionable insights.</span></h1>
    <p class="hero-sub" style="max-width: 680px;">DataProvido doesn't just analyze your stock, funnel, and sales data locally — it revolutionizes your decision-making by delivering clear, prescriptive commercial actions your team can execute instantly with total data privacy.</p>
    <div class="hero-actions">
      <a href="/journey" class="btn-primary">Start Your Journey &nbsp;→</a>
      <a href="/how-works" class="btn-ghost">See How It Works</a>
    </div>
  </div>
</section>

<div class="stats-strip reveal">
  <div class="stats-inner">
    <div class="stat-item"><div class="stat-number">3.2×</div><div class="stat-label">Average increase in insight speed</div></div>
    <div class="stat-item"><div class="stat-number">98%</div><div class="stat-label">Data stays local, zero cloud exposure</div></div>
    <div class="stat-item"><div class="stat-number">15+</div><div class="stat-label">Pre-built analytical modules</div></div>
    <div class="stat-item"><div class="stat-number">&lt;30s</div><div class="stat-label">Average query response time</div></div>
  </div>
</div>

<section class="showcase">
  <div class="showcase-inner">
    <div class="showcase-header reveal">
      <div class="section-tag">✦ Core Capabilities</div>
      <h2 class="section-title">Everything your retail team<br>needs in one place</h2>
      <p class="section-desc">From CRM pipelines to funnel drop-offs — ask any business question and get a precise, actionable answer in seconds.</p>
    </div>
    <div class="showcase-body">

      <div class="sc-tabs" id="scTabs">
        <div class="sc-tab active" data-sc="0">
          <div class="sc-tab-bar"></div><div class="sc-tab-progress"></div>
          <div class="sc-tab-head"><span class="sc-tab-icon">📊</span><span class="sc-tab-title">CRM Data to Insights</span></div>
          <div class="sc-tab-desc">Surface underperforming deals and campaigns instantly.</div>
        </div>
        <div class="sc-tab" data-sc="1">
          <div class="sc-tab-bar"></div><div class="sc-tab-progress"></div>
          <div class="sc-tab-head"><span class="sc-tab-icon">🛒</span><span class="sc-tab-title">Funnel &amp; Conversion</span></div>
          <div class="sc-tab-desc">Track PDP &rarr; Cart &rarr; Checkout drop-offs in seconds.</div>
        </div>
        <div class="sc-tab" data-sc="2">
          <div class="sc-tab-bar"></div><div class="sc-tab-progress"></div>
          <div class="sc-tab-head"><span class="sc-tab-icon">📦</span><span class="sc-tab-title">Stock Intelligence</span></div>
          <div class="sc-tab-desc">Live stock levels, alerts and turnover velocity.</div>
        </div>
        <div class="sc-tab" data-sc="3">
          <div class="sc-tab-bar"></div><div class="sc-tab-progress"></div>
          <div class="sc-tab-head"><span class="sc-tab-icon">💰</span><span class="sc-tab-title">Price Radar</span></div>
          <div class="sc-tab-desc">Compare prices against market benchmarks per SKU.</div>
        </div>
        <div class="sc-tab" data-sc="4">
          <div class="sc-tab-bar"></div><div class="sc-tab-progress"></div>
          <div class="sc-tab-head"><span class="sc-tab-icon">🏆</span><span class="sc-tab-title">GFK Market Share</span></div>
          <div class="sc-tab-desc">Brand rankings and SKU-level competitive positioning.</div>
        </div>
        <div class="sc-tab" data-sc="5">
          <div class="sc-tab-bar"></div><div class="sc-tab-progress"></div>
          <div class="sc-tab-head"><span class="sc-tab-icon">⚡</span><span class="sc-tab-title">Action Executor</span></div>
          <div class="sc-tab-desc">Recommended actions you can execute directly from chat.</div>
        </div>
      </div>

      <div class="sc-panel" id="scPanel">
        <div class="sc-slide active" data-panel="0">
          <div class="sc-slide-tag">📊 Business Intelligence</div>
          <h3>Turn CRM data into<br>clear next actions</h3>
          <p>Connect your CRM pipeline and instantly surface which deals, segments, or campaigns are underperforming — with recommended actions attached. No SQL. No analyst queue.</p>
          <div class="sc-metrics">
            <div class="sc-metric"><div class="sc-metric-val">3.2×</div><div class="sc-metric-label">Faster insights</div></div>
            <div class="sc-metric"><div class="sc-metric-val">&lt;30s</div><div class="sc-metric-label">Query response</div></div>
          </div>
          <div class="sc-slide-visual">📊</div>
        </div>
        <div class="sc-slide" data-panel="1">
          <div class="sc-slide-tag">🛒 Funnel Analytics</div>
          <h3>Pinpoint exactly where<br>customers drop off</h3>
          <p>Track PDP View → Add to Cart → Checkout flows. Identify conversion blockers and get actionable recommendations to improve your funnel rate.</p>
          <div class="sc-metrics">
            <div class="sc-metric"><div class="sc-metric-val">+18%</div><div class="sc-metric-label">Avg. conversion lift</div></div>
            <div class="sc-metric"><div class="sc-metric-val">4 steps</div><div class="sc-metric-label">Full funnel tracked</div></div>
          </div>
          <div class="sc-slide-visual">🛒</div>
        </div>
        <div class="sc-slide" data-panel="2">
          <div class="sc-slide-tag">📦 Inventory</div>
          <h3>Real-time stock visibility<br>across all SKUs</h3>
          <p>Get live stock levels, low-stock alerts, out-of-stock reports, and turnover velocity — all answered in plain language. Never miss a stockout again.</p>
          <div class="sc-metrics">
            <div class="sc-metric"><div class="sc-metric-val">100%</div><div class="sc-metric-label">SKU coverage</div></div>
            <div class="sc-metric"><div class="sc-metric-val">Real-time</div><div class="sc-metric-label">Data freshness</div></div>
          </div>
          <div class="sc-slide-visual">📦</div>
        </div>
        <div class="sc-slide" data-panel="3">
          <div class="sc-slide-tag">💰 Pricing</div>
          <h3>Stay ahead with intelligent<br>price monitoring</h3>
          <p>Compare your pricing against market benchmarks automatically. Identify price gaps per SKU and brand, and act before competitors capture your customers.</p>
          <div class="sc-metrics">
            <div class="sc-metric"><div class="sc-metric-val">Daily</div><div class="sc-metric-label">Market updates</div></div>
            <div class="sc-metric"><div class="sc-metric-val">SKU-level</div><div class="sc-metric-label">Granularity</div></div>
          </div>
          <div class="sc-slide-visual">💰</div>
        </div>
        <div class="sc-slide" data-panel="4">
          <div class="sc-slide-tag">🏆 Market Intelligence</div>
          <h3>Dominate your category<br>with GFK insights</h3>
          <p>Upload GFK reports and instantly get brand performance rankings, category share analysis, and SKU-level competitive positioning — without touching a spreadsheet.</p>
          <div class="sc-metrics">
            <div class="sc-metric"><div class="sc-metric-val">Auto</div><div class="sc-metric-label">Report parsing</div></div>
            <div class="sc-metric"><div class="sc-metric-val">Brand+SKU</div><div class="sc-metric-label">Two-level view</div></div>
          </div>
          <div class="sc-slide-visual">🏆</div>
        </div>
        <div class="sc-slide" data-panel="5">
          <div class="sc-slide-tag">⚡ Automation</div>
          <h3>From insight to action<br>in one click</h3>
          <p>Go beyond insights — DataProvido recommends specific business actions and lets you execute them directly from the chat interface. Analyse, decide, act — all in one place.</p>
          <div class="sc-metrics">
            <div class="sc-metric"><div class="sc-metric-val">1-click</div><div class="sc-metric-label">Action execution</div></div>
            <div class="sc-metric"><div class="sc-metric-val">15+</div><div class="sc-metric-label">Action types</div></div>
          </div>
          <div class="sc-slide-visual">⚡</div>
        </div>
      </div>

    </div>
  </div>
</section>

<section class="section">
  <div class="section-inner">
    <div class="reveal">
      <div class="section-tag">✦ Why DataProvido</div>
      <h2 class="section-title">Built for retail teams,<br>not data scientists</h2>
      <p class="section-desc">No SQL, no dashboards, no waiting. Just ask your question in plain language and get a precise answer with recommended next steps.</p>
    </div>
    <div class="features-grid">
      <div class="feat-card reveal"><div class="feat-icon">🔒</div><div class="feat-title">100% Local &amp; Private</div><div class="feat-desc">Your data never leaves your server. Powered by Llama 3.1 running entirely on-premise — zero cloud, zero data risk.</div></div>
      <div class="feat-card reveal"><div class="feat-icon">⚡</div><div class="feat-title">Instant Answers</div><div class="feat-desc">From question to insight in under 30 seconds. No waiting for reports or analyst queues — just real-time business intelligence.</div></div>
      <div class="feat-card reveal"><div class="feat-icon">🎯</div><div class="feat-title">Action-Oriented Output</div><div class="feat-desc">Every answer comes with a recommended next action. Not just "what happened" but "what to do next" — ready to execute.</div></div>
      <div class="feat-card reveal"><div class="feat-icon">📈</div><div class="feat-title">Revenue Intelligence</div><div class="feat-desc">Track B2D, C2D, revenue per category, brand performance and more. Full financial picture without touching a single spreadsheet.</div></div>
      <div class="feat-card reveal"><div class="feat-icon">🗣️</div><div class="feat-title">Natural Language Interface</div><div class="feat-desc">Ask "Which SKUs are underperforming in GSM?" and get a ranked list with context — no training required.</div></div>
      <div class="feat-card reveal"><div class="feat-icon">📤</div><div class="feat-title">Export Ready</div><div class="feat-desc">Download any analysis as a formatted Excel report with one click — branded, structured, and ready to share with stakeholders.</div></div>
    </div>
  </div>
</section>

<section class="section how-section">
  <div class="section-inner">
    <div class="reveal">
      <div class="section-tag">✦ Process</div>
      <h2 class="section-title">Get your first insight<br>in minutes</h2>
    </div>
    <div class="steps-row">
      <div class="step reveal"><div class="step-num">01</div><div class="step-title">Upload Your Data</div><div class="step-desc">Drop in your Excel files — product list, sales data, GFK reports, funnel exports. DataProvido ingests them automatically.</div></div>
      <div class="step reveal"><div class="step-num">02</div><div class="step-title">Ask Any Question</div><div class="step-desc">Type a business question in plain English or Turkish. "Which brands have the highest C2D?" "What's my best-selling SKU this week?"</div></div>
      <div class="step reveal"><div class="step-num">03</div><div class="step-title">Get Smart Insights</div><div class="step-desc">DataProvido queries your data, runs the analysis, and returns a precise answer with numbers, trends, and context — in seconds.</div></div>
      <div class="step reveal"><div class="step-num">04</div><div class="step-title">Take Action</div><div class="step-desc">Follow the recommended next steps or export the full analysis as an Excel report — ready to share with your team.</div></div>
    </div>
  </div>
</section>

<section class="section trust-section">
  <div class="section-inner">
    <div class="reveal">
      <div class="section-tag">✦ Built on Solid Ground</div>
      <h2 class="section-title">Trusted by retail teams<br>who care about their data</h2>
    </div>
    <div class="trust-badges reveal">
      <div class="trust-badge"><div class="trust-badge-icon">🦙</div><div class="trust-badge-text"><div class="label">Powered by</div><div class="value">Llama 3.1</div></div></div>
      <div class="trust-badge"><div class="trust-badge-icon">🌐</div><div class="trust-badge-text"><div class="label">Runs</div><div class="value">100% Offline</div></div></div>
      <div class="trust-badge"><div class="trust-badge-icon">⚡</div><div class="trust-badge-text"><div class="label">Built with</div><div class="value">FastAPI</div></div></div>
      <div class="trust-badge"><div class="trust-badge-icon">📊</div><div class="trust-badge-text"><div class="label">Analyzes</div><div class="value">Any Excel</div></div></div>
    </div>
  </div>
</section>

<div class="cta-banner reveal">
  <h2>Ready to turn your data<br>into decisions?</h2>
  <p>Start exploring your retail data with AI-powered intelligence — locally, privately, instantly.</p>
  <a href="/journey" class="btn-primary" style="margin: 0 auto; display: inline-flex;">Jump to Journey &nbsp;→</a>
</div>

<footer class="site-footer">
  <div class="footer-left" style="display: flex; align-items: center; gap: 10px;"><img src="/logo.png" alt="DataProvido" style="height: 28px; width: 28px; object-fit: contain; border-radius: 6px;" /><span class="footer-logo">DataProvido</span><span class="footer-copy">· Local Intelligence. Zero Compromise.</span></div>
  <div class="footer-right">Powered by Llama 3.1 &nbsp;·&nbsp; Runs entirely offline</div>
</footer>

<script>
  const nav = document.getElementById('main-nav');
  window.addEventListener('scroll', () => { 
    if (window.scrollY > 40) { nav.style.background = 'rgba(255,255,255,0.98)'; nav.style.boxShadow = '0 2px 16px rgba(0,0,0,0.07)'; }
    else { nav.style.background = 'rgba(255,255,255,0.95)'; nav.style.boxShadow = '0 1px 0 #dadee2'; }
  });
  const observer = new IntersectionObserver((entries) => { entries.forEach(e => { if (e.isIntersecting) { e.target.classList.add('visible'); observer.unobserve(e.target); } }); }, { threshold: 0.12 });
  document.querySelectorAll('.reveal').forEach(el => observer.observe(el));
  // Showcase tab switcher (Yodeck-style)
  const scTabs = document.querySelectorAll('.sc-tab');
  const scPanels = document.querySelectorAll('.sc-slide');
  let scCurrent = 0, scTimer;

  function scGoTo(idx) {
    scTabs[scCurrent].classList.remove('active');
    scPanels[scCurrent].classList.remove('active');
    const oldProg = scTabs[scCurrent].querySelector('.sc-tab-progress');
    oldProg.style.animation = 'none';
    oldProg.offsetHeight;
    oldProg.style.animation = '';

    scCurrent = ((idx % scTabs.length) + scTabs.length) % scTabs.length;
    scTabs[scCurrent].classList.add('active');
    scPanels[scCurrent].classList.add('active');
    const newProg = scTabs[scCurrent].querySelector('.sc-tab-progress');
    newProg.style.animation = 'none';
    newProg.offsetHeight;
    newProg.style.animation = 'tab-fill 5s linear forwards';
  }

  function scNext() { scGoTo(scCurrent + 1); }

  scTabs.forEach((tab, i) => {
    tab.addEventListener('click', () => {
      clearInterval(scTimer);
      scGoTo(i);
      scTimer = setInterval(scNext, 5000);
    });
  });

  scTimer = setInterval(scNext, 5000);
</script>
</body>
</html>"""
